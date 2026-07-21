"""
Export tests: Excel export for materials and inventory records.
"""
import pytest
from io import BytesIO


class TestMaterialsExport:
    """Materials Excel export tests."""

    def test_export_materials_excel(self, admin_client, sample_material):
        """Export materials should return a valid Excel file."""
        resp = admin_client.get("/api/materials/export-excel")
        assert resp.status_code == 200
        content_type = resp.headers.get('content-type', '')
        assert 'spreadsheet' in content_type or 'octet-stream' in content_type

    def test_export_materials_has_content(self, admin_client, sample_material):
        """Exported Excel should have actual content."""
        resp = admin_client.get("/api/materials/export-excel")
        assert resp.status_code == 200
        assert len(resp.content) > 0

        # Verify it's a valid xlsx
        from openpyxl import load_workbook
        wb = load_workbook(filename=BytesIO(resp.content))
        ws = wb.active
        # Should have at least header row + data row
        assert ws.max_row >= 2


class TestRecordsExport:
    """Inventory records Excel export tests."""

    def test_export_records_excel(self, admin_client):
        """Export records should return a valid Excel file."""
        resp = admin_client.get("/api/inventory/export-excel")
        assert resp.status_code == 200
        content_type = resp.headers.get('content-type', '')
        assert 'spreadsheet' in content_type or 'octet-stream' in content_type

    def test_export_records_has_content(self, admin_client):
        """Exported records Excel should have actual data."""
        resp = admin_client.get("/api/inventory/export-excel")
        assert resp.status_code == 200
        assert len(resp.content) > 0

        from openpyxl import load_workbook
        wb = load_workbook(filename=BytesIO(resp.content))
        ws = wb.active
        assert ws.max_row >= 1  # At least header

    def test_export_records_account_and_operator_columns(self, admin_client,
                                                          sample_material):
        """账号（DB operator）与 操作人（DB actual_operator）分列，操作人列携带姓名。"""
        resp = admin_client.post("/api/materials/stock-in", json={
            "product_name": sample_material['name'],
            "quantity": 7,
            "reason_category": "purchase",
            "warehouse_id": sample_material['warehouse_id'],
            "actual_operator": "张三",
        })
        assert resp.status_code == 200 and resp.json()['success'] is True

        resp = admin_client.get("/api/inventory/export-excel",
                                params={"product_name": sample_material['name']})
        assert resp.status_code == 200

        from openpyxl import load_workbook
        wb = load_workbook(filename=BytesIO(resp.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        # 账号 then 操作人（账号在前）。
        assert '账号' in headers and '操作人' in headers
        assert headers.index('操作人') == headers.index('账号') + 1

        acct_col = headers.index('账号') + 1
        op_col = headers.index('操作人') + 1
        acct_cells = [ws.cell(row=r, column=acct_col).value
                      for r in range(2, ws.max_row + 1)]
        op_cells = [ws.cell(row=r, column=op_col).value
                    for r in range(2, ws.max_row + 1)]
        # 账号 列不含括号姓名（不再是 "operator (张三)"）。
        assert all(not (v and '(张三)' in v) for v in acct_cells), acct_cells
        # 操作人 列携带姓名。
        assert any(v == '张三' for v in op_cells), op_cells

    def test_export_out_record_variant(self, admin_client, sample_material):
        """出库记录（batch_id 为 NULL）导出时规格应从被消耗批次带出，而非空。"""
        # 带规格入库
        resp = admin_client.post("/api/materials/stock-in", json={
            "product_name": sample_material['name'],
            "quantity": 5,
            "reason_category": "purchase",
            "warehouse_id": sample_material['warehouse_id'],
            "variant": "红",
        })
        assert resp.status_code == 200 and resp.json()['success'] is True
        # 全量出库
        resp = admin_client.post("/api/materials/stock-out", json={
            "product_name": sample_material['name'],
            "quantity": 5,
            "reason_category": "sell",
            "warehouse_id": sample_material['warehouse_id'],
            "variant": "红",
        })
        assert resp.status_code == 200 and resp.json()['success'] is True

        resp = admin_client.get("/api/inventory/export-excel",
                                params={"product_name": sample_material['name']})
        assert resp.status_code == 200
        from openpyxl import load_workbook
        wb = load_workbook(filename=BytesIO(resp.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        type_col = headers.index('记录类型') + 1
        variant_col = headers.index('规格') + 1
        out_variants = [ws.cell(row=r, column=variant_col).value
                        for r in range(2, ws.max_row + 1)
                        if ws.cell(row=r, column=type_col).value == '出库']
        assert out_variants and all(v == '红' for v in out_variants), out_variants
