"""
Import tests: Excel import preview, confirm execution, new SKU creation.
"""
import pytest
import os
from io import BytesIO


def get_test_data_path():
    """Get path to test data file."""
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, 'test', 'data', 'inventory_test_data.xlsx')


class TestImportPreview:
    """Excel import preview tests."""

    def test_preview_with_valid_file(self, admin_client):
        """Preview with a valid Excel file should succeed."""
        test_file = get_test_data_path()
        if not os.path.exists(test_file):
            pytest.skip("Test data file not found")

        with open(test_file, 'rb') as f:
            resp = admin_client.post(
                "/api/materials/import-excel/preview",
                files={"file": ("test.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            )
        assert resp.status_code == 200
        data = resp.json()
        assert data['success'] is True
        assert isinstance(data['preview'], list)
        assert 'total_in' in data
        assert 'total_out' in data

    def test_preview_with_invalid_file(self, admin_client):
        """Preview with invalid content should return error."""
        fake_content = b"This is not an Excel file"
        resp = admin_client.post(
            "/api/materials/import-excel/preview",
            files={"file": ("bad.xlsx", BytesIO(fake_content), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data['success'] is False

    def test_preview_with_generated_excel(self, admin_client, sample_material):
        """Preview with a programmatically generated Excel file."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        # Header row matching expected format
        ws.append(['Name', 'SKU', 'Category', 'Quantity', 'Unit', 'Safe Stock', 'Location'])
        # Data rows
        ws.append([sample_material['name'], sample_material['sku'], 'Test Category', 200, 'pcs', 20, 'A-01'])
        ws.append(['New Import Item', 'IMPORT-001', 'New Cat', 50, 'pcs', 10, 'Z-01'])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        resp = admin_client.post(
            "/api/materials/import-excel/preview",
            files={"file": ("import.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data['success'] is True

    def test_preview_rejects_records_export_template(self, admin_client):
        """Inventory records exports should not be accepted as inventory import templates."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(['物料名称', '规格', '物料编码', '商品类型', '记录类型', '数量', '批次', '联系方', '操作人', '原因类别', '备注', '时间'])
        ws.append(['测试物料', '', 'SKU-001', '分类', '入库', 10, 'BATCH-001', '', 'admin', '采购入库', '', '2026-04-13 17:17:27'])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        resp = admin_client.post(
            "/api/materials/import-excel/preview",
            files={"file": ("inventory_records.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        assert resp.status_code == 200
        data = resp.json()
        assert data['success'] is False
        assert '出入库记录导出文件' in data['message']

    def test_preview_inventory_export_with_empty_batch_column(self, admin_client, sample_material):
        """Old inventory exports with an empty batch column should import as a simple snapshot."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(['物料名称', '物料编码(SKU)', '分类', '单位', '安全库存', '批次号', '变体', '库存', '存放位置', '联系方'])
        row = [sample_material['name'], sample_material['sku'], 'Test Category', 'pcs', 20, None, None, 120, 'A-01', None]
        ws.append(row)
        ws.append(row)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        resp = admin_client.post(
            "/api/materials/import-excel/preview",
            files={"file": ("inventory.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        assert resp.status_code == 200
        data = resp.json()
        assert data['success'] is True
        assert data['is_batch_mode'] is False
        assert len(data['preview']) == 1
        assert '重复行' in data['message']


class TestSafeStockColumnMapping:
    """「安全库存」列必须真的进库 —— 它含 '库存' 二字，极易被通用数量列吃掉。"""

    # 当前导出/下载模板的表头，以及客户手上那份旧模板（'变体' 尚未改名 '规格'）。
    CANONICAL = ['物料名称', '规格', '物料编码(SKU)', '分类', '单位',
                 '安全库存', '批次号', '库存', '存放位置', '联系方']
    LEGACY = ['物料名称', '物料编码(SKU)', '分类', '单位', '安全库存',
              '批次号', '变体', '库存', '存放位置', '联系方']

    def _upload(self, admin_client, headers, row):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "库存数据"
        ws.append(headers)
        ws.append(row)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return admin_client.post(
            "/api/materials/import-excel/preview",
            files={"file": ("t.xlsx", buf,
                            "application/vnd.openxmlformats-officedocument."
                            "spreadsheetml.sheet")},
        )

    @pytest.mark.parametrize("layout", ["canonical", "legacy"])
    def test_safe_stock_reaches_the_database(self, admin_client, layout,
                                             default_warehouse_id):
        import uuid
        sku = f"SS-{uuid.uuid4().hex[:8].upper()}"
        if layout == "canonical":
            headers = self.CANONICAL
            row = ['安全库存物料', '', sku, '测试', '个', 42, '', 5, 'A-01', '']
        else:
            headers = self.LEGACY
            row = ['安全库存物料', sku, '测试', '个', 42, '', '', 5, 'A-01', '']

        resp = self._upload(admin_client, headers, row)
        assert resp.status_code == 200
        data = resp.json()
        assert data['success'] is True, data

        confirm = admin_client.post("/api/materials/import-excel/confirm", json={
            "changes": data['preview'],
            "reason_category": "purchase",
            "confirm_new_skus": True,
        })
        assert confirm.status_code == 200, confirm.text

        from database import get_db_connection
        conn = get_db_connection()
        try:
            row_db = conn.execute(
                "SELECT safe_stock, quantity FROM materials WHERE sku = ?", (sku,)
            ).fetchone()
        finally:
            conn.close()
        assert row_db is not None, "物料未创建"
        # 回归点：安全库存曾被通用「库存」分支吃掉，整列静默丢失。
        assert row_db['safe_stock'] == 42

    @pytest.mark.parametrize("layout", ["canonical", "legacy"])
    def test_quantity_is_not_taken_from_safe_stock_column(self, admin_client,
                                                          layout):
        """调整分支顺序不能反过来让安全库存冒充数量列。"""
        headers = self.CANONICAL if layout == "canonical" else self.LEGACY
        if layout == "canonical":
            row = ['数量归属物料', '', 'SS-QTY-1', '测试', '个', 42, '', 5, 'A-01', '']
        else:
            row = ['数量归属物料', 'SS-QTY-1', '测试', '个', 42, '', '', 5, 'A-01', '']
        data = self._upload(admin_client, headers, row).json()
        assert data['success'] is True, data
        item = data['preview'][0]
        assert item['import_quantity'] == 5
        assert item['safe_stock'] == 42


class TestDuplicateSkuWarning:
    """简化模式下同 SKU 多行会造虚假流水，预览必须在确认前说清楚。"""

    HEADERS = ['物料名称', '规格', '物料编码(SKU)', '分类', '单位',
               '安全库存', '批次号', '库存', '存放位置', '联系方']

    def _preview(self, admin_client, rows):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "库存数据"
        ws.append(self.HEADERS)
        for r in rows:
            ws.append(r)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return admin_client.post(
            "/api/materials/import-excel/preview",
            files={"file": ("t.xlsx", buf,
                            "application/vnd.openxmlformats-officedocument."
                            "spreadsheetml.sheet")},
        ).json()

    def test_duplicate_sku_in_simple_mode_warns(self, admin_client):
        """同一编码两行（用户拿多行表达多库位）→ 必须告警并指出批次模式。"""
        rows = [
            ['多库位物料', '', 'DUP-001', '测试', '个', None, '', 112, 'A1-2-03', ''],
            ['多库位物料', '', 'DUP-001', '测试', '个', None, '', 56, 'A1-2-04', ''],
        ]
        data = self._preview(admin_client, rows)
        assert data['success'] is True
        assert data['is_batch_mode'] is False
        warns = data['duplicate_sku_warnings']
        assert warns, "同 SKU 多行必须告警"
        joined = "".join(warns)
        assert 'DUP-001' in joined
        assert '批次号' in joined, "必须指出正解是走批次模式"

    def test_unique_skus_produce_no_warning(self, admin_client):
        rows = [
            ['物料甲', '', 'UNI-001', '测试', '个', None, '', 10, 'A-01', ''],
            ['物料乙', '', 'UNI-002', '测试', '个', None, '', 20, 'A-02', ''],
        ]
        assert self._preview(admin_client, rows)['duplicate_sku_warnings'] == []

    def test_batch_mode_does_not_warn(self, admin_client):
        """批次模式下每行是独立批次，同 SKU 多行完全正常，不该报警。"""
        rows = [
            ['多批次物料', '', 'BAT-001', '测试', '个', None, 'B-A-1', 112, 'A1-2-03', ''],
            ['多批次物料', '', 'BAT-001', '测试', '个', None, 'B-A-2', 56, 'A1-2-04', ''],
        ]
        data = self._preview(admin_client, rows)
        assert data['is_batch_mode'] is True
        assert data['duplicate_sku_warnings'] == []


class TestImportConfirm:
    """Excel import confirmation tests."""

    def test_confirm_import_creates_records(self, admin_client, sample_material):
        """Confirming import should create inventory records."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(['Name', 'SKU', 'Category', 'Quantity', 'Unit', 'Safe Stock', 'Location'])
        # Set quantity higher than current (100) to trigger stock-in
        ws.append([sample_material['name'], sample_material['sku'], 'Test Category', 120, 'pcs', 20, 'A-01'])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Preview first
        preview_resp = admin_client.post(
            "/api/materials/import-excel/preview",
            files={"file": ("import.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        preview_data = preview_resp.json()
        if preview_data.get('success') and preview_data.get('preview'):
            # Confirm import with the changes from preview
            confirm_resp = admin_client.post("/api/materials/import-excel/confirm", json={
                "changes": preview_data['preview'],
                "reason_category": "purchase",
                "confirm_new_skus": False,
                "confirm_disable_missing_skus": False,
                "warehouse_id": sample_material['warehouse_id']
            })
            assert confirm_resp.status_code == 200
            data = confirm_resp.json()
            assert data['success'] is True

    def test_confirm_import_new_sku(self, admin_client, default_warehouse_id):
        """Import with new SKU should create new material."""
        from openpyxl import Workbook
        import uuid

        new_sku = f"NEW-{uuid.uuid4().hex[:6].upper()}"
        new_name = f"New Material {new_sku}"

        wb = Workbook()
        ws = wb.active
        ws.append(['Name', 'SKU', 'Category', 'Quantity', 'Unit', 'Safe Stock', 'Location'])
        ws.append([new_name, new_sku, 'Import Cat', 75, 'pcs', 15, 'X-01'])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Preview
        preview_resp = admin_client.post(
            "/api/materials/import-excel/preview",
            files={"file": ("import.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        preview_data = preview_resp.json()
        if preview_data.get('success'):
            # Check that new SKU is detected
            assert preview_data['total_new'] >= 1 or len(preview_data.get('new_skus', [])) >= 1

            # Confirm with new SKUs included
            changes = preview_data.get('preview', [])
            confirm_resp = admin_client.post("/api/materials/import-excel/confirm", json={
                "changes": changes,
                "reason_category": "purchase",
                "confirm_new_skus": True,
                "confirm_disable_missing_skus": False,
                "warehouse_id": default_warehouse_id
            })
            assert confirm_resp.status_code == 200


class TestImportUnitSanitization:
    """Imported cells must strip Excel formula residue (e.g. '=+VLOOKUP(...)').

    The whole-DB import path (``_import_tenant_database``) is exercised in
    ``test_import_fuzzy_invalidation.py`` (isolated to a throwaway tenant); here
    we cover the pure sanitizer and the Excel preview wiring.
    """

    def test_sanitize_import_text_drops_formula(self):
        """The shared sanitizer turns formula residue into None, keeps real text."""
        import sys
        sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'backend'))
        from app import _sanitize_import_text
        assert _sanitize_import_text("=+VLOOKUP(C1,[1]x!$B:$I,8,0)") is None
        assert _sanitize_import_text("=SUM(A1:A2)") is None
        assert _sanitize_import_text(None) is None
        assert _sanitize_import_text("   ") is None
        assert _sanitize_import_text("  Pcs ") == "Pcs"
        assert _sanitize_import_text("个") == "个"

    def test_excel_preview_strips_formula_unit(self, admin_client):
        """An Excel cell that is actually a leftover formula must not become the unit."""
        from openpyxl import Workbook
        import uuid

        sku = f"FX-{uuid.uuid4().hex[:6].upper()}"
        wb = Workbook()
        ws = wb.active
        ws.append(['Name', 'SKU', 'Category', 'Quantity', 'Unit', 'Safe Stock', 'Location'])
        ws.append([f'Formula Unit {sku}', sku, 'Cat', 10,
                   '=+VLOOKUP(C4153,[1]back!$B:$I,8,0)', 5, 'A-01'])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        resp = admin_client.post(
            "/api/materials/import-excel/preview",
            files={"file": ("f.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data['success'] is True
        item = next(p for p in data['preview'] if p['sku'] == sku)
        assert '=' not in (item['unit'] or ''), f"formula leaked into unit: {item['unit']!r}"
        assert item['unit'] == '个'  # sanitized -> default unit


class TestFormulaWarnings:
    """公式没被求值时，预览必须**明说**，不能静默把该列变成空值。

    背景：openpyxl 取的是公式的缓存结果。文件若没被 Excel/WPS 计算保存过，
    缓存为空 → 读成 None → 库位悄无声息变空。现场已经因为这条链踩过一次
    （当时是反向：没传 data_only，公式串原样写进了 1306 条物料的库位）。
    修完源头之后，剩下的风险就是"静默变空"，所以要在预览里拦一道。
    """

    @staticmethod
    def _book(location_value, formula=True):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(['物料名称', '物料编码(SKU)', '分类', '单位', '库存', '存放位置'])
        ws.append(['螺母A', 'T-F001', '未分类', 'Pcs', 3, location_value])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _preview(self, client, buf):
        r = client.post(
            "/api/materials/import-excel/preview",
            files={"file": ("f.xlsx", buf,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        assert r.status_code == 200, r.text
        return r.json()

    def test_uncalculated_formula_is_reported(self, admin_client):
        """openpyxl 写出的公式没有缓存值——正是"打开就变空"的那种文件。"""
        buf = self._book('=+IFERROR(VLOOKUP(B2,[1]Sheet1!$A:$B,2,FALSE),"临时库位")')
        data = self._preview(admin_client, buf)
        assert data['success'] is True
        warns = data.get('formula_warnings') or []
        assert warns, "公式取不到结果值却没有任何告警——客户只会看到库位莫名消失"
        joined = ' '.join(warns)
        assert '存放位置' in joined
        assert '第 2 行' in joined, f"要指出具体行号才能让客户去改：{joined}"

    def test_plain_value_produces_no_warning(self, admin_client):
        """正常字面量不能误报，否则告警会被当成噪音忽略。"""
        data = self._preview(admin_client, self._book('12-2'))
        assert data['success'] is True
        assert not (data.get('formula_warnings') or [])

    def test_empty_cell_is_not_a_formula_warning(self, admin_client):
        """空格子本来就是空，不该报成公式问题。"""
        data = self._preview(admin_client, self._book(None))
        assert data['success'] is True
        assert not (data.get('formula_warnings') or [])
