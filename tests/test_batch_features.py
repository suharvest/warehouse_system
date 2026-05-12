"""
Batch-aware features tests:
- Schema migration (LEGACY batches, location column)
- Batches API endpoint
- Excel export with batch columns
- Import preview (simple mode & batch mode)
- Import confirm with batch creation
- Stock-in with location
- Round-trip export → import consistency
"""
import pytest
from io import BytesIO
from openpyxl import Workbook, load_workbook


# ============ Fixtures ============

@pytest.fixture()
def material_with_batch(admin_client, material_with_legacy_batch):
    """Create a sample material with LEGACY batch + a new stock-in batch."""
    mat = material_with_legacy_batch
    resp = admin_client.post("/api/materials/stock-in", json={
        "product_name": mat['name'],
        "quantity": 50,
        "reason_category": "purchase",
        "location": "A区-01",
        "fuzzy": False,
        "warehouse_id": mat['warehouse_id'],
    })
    data = resp.json()
    assert data['success'] is True
    return {
        **mat,
        'quantity': mat['quantity'] + 50,
        'batch_no': data['batch']['batch_no'],
        'batch_id': data['batch']['batch_id'],
    }


@pytest.fixture()
def material_with_legacy_batch(admin_client, sample_material):
    """Ensure the sample material has a LEGACY batch (for materials created after init)."""
    from database import get_db_connection
    conn = get_db_connection()
    cursor = conn.cursor()
    # Check if already has a batch
    cursor.execute('SELECT COUNT(*) as c FROM batches WHERE material_id = ? AND is_exhausted = 0',
                   (sample_material['id'],))
    if cursor.fetchone()['c'] == 0:
        # Create LEGACY batch manually (simulating what init_database does)
        import datetime
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        batch_no = f"LEGACY-{sample_material['id']:04d}"
        # Already verified count==0 above, so plain INSERT is portable
        # across sqlite and MySQL (avoid sqlite-only `INSERT OR IGNORE`).
        # Look up the material's tenant_id/warehouse_id so the LEGACY batch
        # carries the same scope. Bug C fix added scope predicates to the
        # Excel FIFO query, so an orphan-scoped batch would be filtered out.
        cursor.execute('SELECT tenant_id, warehouse_id FROM materials WHERE id = ?',
                       (sample_material['id'],))
        scope = cursor.fetchone()
        cursor.execute('''
            INSERT INTO batches (batch_no, material_id, quantity, initial_quantity, location, created_at, tenant_id, warehouse_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (batch_no, sample_material['id'], sample_material['quantity'],
              sample_material['quantity'], 'A-01', now,
              scope['tenant_id'], scope['warehouse_id']))
        conn.commit()
    conn.close()
    return sample_material


@pytest.fixture()
def sample_contact(admin_client):
    """Create a sample contact for testing."""
    import uuid
    name = f"Test Supplier {uuid.uuid4().hex[:6]}"
    resp = admin_client.post("/api/contacts", json={
        "name": name,
        "is_supplier": True,
        "is_customer": False,
    })
    assert resp.status_code == 200
    return resp.json()


def _make_excel(rows, headers=None):
    """Helper: create an Excel file in memory from rows."""
    wb = Workbook()
    ws = wb.active
    if headers:
        ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _upload_preview(admin_client, excel_buf):
    """Helper: upload Excel for import preview."""
    resp = admin_client.post(
        "/api/materials/import-excel/preview",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    )
    assert resp.status_code == 200
    return resp.json()


# ============ Schema Migration Tests ============

class TestSchemaMigration:
    """Test that database migration creates LEGACY batches and location column."""

    def test_batches_table_has_location_column(self, admin_client):
        """batches table should have a location column after migration."""
        from db import get_engine
        from sqlalchemy import inspect
        insp = inspect(get_engine())
        columns = {c['name'] for c in insp.get_columns('batches')}
        assert 'location' in columns

    # Removed: test_legacy_batches_created_for_orphan_materials
    # 单一真相源切换后，sample_material fixture 直接插入 active batch，不再有
    # "孤儿 material.quantity > 0 / batches 为空" 的过渡场景。LEGACY 合成逻辑保留
    # 在 db 初始化代码以兼容旧库，但不再为它写单测（已无可触发路径）。


# ============ Batches API Tests ============

class TestBatchesAPI:
    """Test GET /api/materials/batches endpoint."""

    def test_get_batches_returns_list(self, admin_client, material_with_batch):
        """Should return a list of active batches for the material."""
        resp = admin_client.get(
            "/api/materials/batches",
            params={"name": material_with_batch['name']}
        )
        assert resp.status_code == 200
        data = resp.json()
        assert 'batches' in data
        assert 'total_quantity' in data
        assert isinstance(data['batches'], list)
        assert len(data['batches']) >= 1

    def test_get_batches_has_correct_fields(self, admin_client, material_with_batch):
        """Each batch should have batch_no, quantity, location, contact_name, created_at."""
        resp = admin_client.get(
            "/api/materials/batches",
            params={"name": material_with_batch['name']}
        )
        data = resp.json()
        batch = data['batches'][-1]  # The most recent one (stock-in)
        assert 'batch_no' in batch
        assert 'quantity' in batch
        assert 'location' in batch
        assert 'contact_name' in batch
        assert 'created_at' in batch

    def test_get_batches_includes_location(self, admin_client, material_with_batch):
        """Stock-in with location should show location on the batch."""
        resp = admin_client.get(
            "/api/materials/batches",
            params={"name": material_with_batch['name']}
        )
        data = resp.json()
        # Find the non-LEGACY batch (the one created via stock-in)
        non_legacy = [b for b in data['batches'] if not b['batch_no'].startswith('LEGACY-')]
        assert len(non_legacy) >= 1
        assert non_legacy[0]['location'] == 'A区-01'

    def test_get_batches_nonexistent_product(self, admin_client):
        """Should return 404 for nonexistent product."""
        resp = admin_client.get(
            "/api/materials/batches",
            params={"name": "NONEXISTENT_PRODUCT_12345"}
        )
        assert resp.status_code == 404


# ============ Stock-in with Location Tests ============

class TestStockInLocation:
    """Test that stock_in properly stores location on batch."""

    def test_stock_in_creates_batch_with_location(self, admin_client, sample_material):
        """Stock-in should create a batch with the specified location."""
        resp = admin_client.post("/api/materials/stock-in", json={
            "product_name": sample_material['name'],
            "quantity": 10,
            "reason_category": "purchase",
            "location": "B区-05",
            "fuzzy": False,
            "warehouse_id": sample_material['warehouse_id'],
        })
        data = resp.json()
        assert data['success'] is True
        batch_no = data['batch']['batch_no']

        # Verify batch has location in DB
        from database import get_db_connection
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT location FROM batches WHERE batch_no = ?', (batch_no,))
        row = cursor.fetchone()
        conn.close()
        assert row['location'] == 'B区-05'

    def test_stock_in_with_contact(self, admin_client, sample_material, sample_contact):
        """Stock-in with contact_id should associate the batch."""
        resp = admin_client.post("/api/materials/stock-in", json={
            "product_name": sample_material['name'],
            "quantity": 5,
            "reason_category": "purchase",
            "contact_id": sample_contact['id'],
            "fuzzy": False,
            "warehouse_id": sample_material['warehouse_id'],
        })
        data = resp.json()
        assert data['success'] is True

        # Verify batch has contact in DB
        from database import get_db_connection
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT contact_id FROM batches WHERE batch_no = ?', (data['batch']['batch_no'],))
        row = cursor.fetchone()
        conn.close()
        assert row['contact_id'] == sample_contact['id']


# ============ Excel Export Tests ============

class TestBatchAwareExport:
    """Test that Excel export includes batch columns."""

    def test_export_has_batch_columns(self, admin_client, sample_material):
        """Exported Excel should have 10 columns including batch_no, variant, location, contact."""
        resp = admin_client.get("/api/materials/export-excel")
        assert resp.status_code == 200

        wb = load_workbook(filename=BytesIO(resp.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert '批次号' in headers
        assert '存放位置' in headers
        assert '联系方' in headers
        assert len(headers) == 10

    def test_export_one_row_per_batch(self, admin_client, material_with_batch, material_with_legacy_batch):
        """A material with 2 batches (LEGACY + stock-in) should produce 2 rows."""
        # material_with_batch depends on sample_material which now also has a LEGACY batch
        resp = admin_client.get("/api/materials/export-excel")
        wb = load_workbook(filename=BytesIO(resp.content))
        ws = wb.active

        # Find SKU column index
        sku_col = None
        for idx, cell in enumerate(ws[1]):
            if cell.value and 'SKU' in cell.value:
                sku_col = idx
                break

        rows_for_material = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[sku_col] == material_with_batch['sku']:
                rows_for_material.append(row)

        # material_with_batch has LEGACY batch (from fixture) + stock-in batch = 2 rows
        assert len(rows_for_material) >= 2


# ============ Import Preview Tests ============

class TestImportPreviewSimpleMode:
    """Test import preview in simple mode (no batch column)."""

    def test_simple_mode_detected(self, admin_client, sample_material):
        """Excel without batch_no column should be detected as simple mode."""
        excel = _make_excel(
            [[sample_material['name'], sample_material['sku'], 'Test', '个', 20, 100]],
            headers=['物料名称', '物料编码(SKU)', '分类', '单位', '安全库存', '库存']
        )
        data = _upload_preview(admin_client, excel)
        assert data['success'] is True
        assert data['is_batch_mode'] is False

    def test_simple_mode_no_change(self, admin_client, sample_material):
        """Same quantity should produce operation=none."""
        excel = _make_excel(
            [[sample_material['name'], sample_material['sku'], 'Test', '个', 20, sample_material['quantity']]],
            headers=['物料名称', '物料编码(SKU)', '分类', '单位', '安全库存', '库存']
        )
        data = _upload_preview(admin_client, excel)
        assert data['preview'][0]['operation'] == 'none'

    def test_simple_mode_in_operation(self, admin_client, sample_material):
        """Higher quantity should produce operation=in."""
        excel = _make_excel(
            [[sample_material['name'], sample_material['sku'], 'Test', '个', 20, sample_material['quantity'] + 50]],
            headers=['物料名称', '物料编码(SKU)', '分类', '单位', '安全库存', '库存']
        )
        data = _upload_preview(admin_client, excel)
        assert data['preview'][0]['operation'] == 'in'
        assert data['total_in'] == 50

    def test_simple_mode_contact_resolution(self, admin_client, sample_material, sample_contact):
        """Contact name should be resolved to contact_id."""
        contact_name = sample_contact['name']
        excel = _make_excel(
            [[sample_material['name'], sample_material['sku'], 'Test', '个', 20, sample_material['quantity'], '', contact_name]],
            headers=['物料名称', '物料编码(SKU)', '分类', '单位', '安全库存', '库存', '存放位置', '联系方']
        )
        data = _upload_preview(admin_client, excel)
        assert data['preview'][0]['contact_name'] == contact_name
        assert data['preview'][0]['contact_id'] == sample_contact['id']

    def test_simple_mode_new_contact_detected(self, admin_client, sample_material):
        """Unknown contact name should appear in new_contacts."""
        excel = _make_excel(
            [[sample_material['name'], sample_material['sku'], 'Test', '个', 20, sample_material['quantity'], '', 'New Supplier XYZ']],
            headers=['物料名称', '物料编码(SKU)', '分类', '单位', '安全库存', '库存', '存放位置', '联系方']
        )
        data = _upload_preview(admin_client, excel)
        assert 'New Supplier XYZ' in data['new_contacts']


class TestImportPreviewBatchMode:
    """Test import preview in batch mode (with batch column)."""

    def test_batch_mode_detected(self, admin_client, material_with_batch):
        """Excel with batch_no column should be detected as batch mode."""
        excel = _make_excel(
            [[material_with_batch['name'], material_with_batch['sku'], 'Test', '个', 20,
              material_with_batch['batch_no'], 50, 'A区-01', '']],
            headers=['物料名称', '物料编码(SKU)', '分类', '单位', '安全库存', '批次号', '库存', '存放位置', '联系方']
        )
        data = _upload_preview(admin_client, excel)
        assert data['success'] is True
        assert data['is_batch_mode'] is True

    def test_batch_mode_existing_no_change(self, admin_client, material_with_batch):
        """Same batch quantity should produce operation=none."""
        excel = _make_excel(
            [[material_with_batch['name'], material_with_batch['sku'], 'Test', '个', 20,
              material_with_batch['batch_no'], 50, 'A区-01', '']],
            headers=['物料名称', '物料编码(SKU)', '分类', '单位', '安全库存', '批次号', '库存', '存放位置', '联系方']
        )
        data = _upload_preview(admin_client, excel)
        assert data['preview'][0]['operation'] == 'none'

    def test_batch_mode_existing_qty_change(self, admin_client, material_with_batch):
        """Changed batch quantity should produce in/out operation."""
        excel = _make_excel(
            [[material_with_batch['name'], material_with_batch['sku'], 'Test', '个', 20,
              material_with_batch['batch_no'], 80, 'A区-01', '']],
            headers=['物料名称', '物料编码(SKU)', '分类', '单位', '安全库存', '批次号', '库存', '存放位置', '联系方']
        )
        data = _upload_preview(admin_client, excel)
        item = data['preview'][0]
        assert item['operation'] == 'in'
        assert item['difference'] == 30

    def test_batch_mode_new_batch(self, admin_client, material_with_batch):
        """Empty batch_no should be treated as new batch."""
        excel = _make_excel(
            [[material_with_batch['name'], material_with_batch['sku'], 'Test', '个', 20,
              '', 25, 'B区-02', '']],
            headers=['物料名称', '物料编码(SKU)', '分类', '单位', '安全库存', '批次号', '库存', '存放位置', '联系方']
        )
        data = _upload_preview(admin_client, excel)
        item = data['preview'][0]
        assert item['is_batch_new'] is True
        assert item['operation'] == 'in'

    def test_batch_mode_unknown_batch_no_treated_as_new(self, admin_client, material_with_batch):
        """Non-existent batch_no should be treated as a new batch import."""
        excel = _make_excel(
            [[material_with_batch['name'], material_with_batch['sku'], 'Test', '个', 20,
              'FAKE-BATCH-999', 50, '', '']],
            headers=['物料名称', '物料编码(SKU)', '分类', '单位', '安全库存', '批次号', '库存', '存放位置', '联系方']
        )
        data = _upload_preview(admin_client, excel)
        assert data['success'] is True
        assert len(data['preview']) == 1
        item = data['preview'][0]
        assert item['batch_no'] == 'FAKE-BATCH-999'
        assert item['is_batch_new'] is True
        assert item['operation'] == 'in'


# ============ Import Confirm Tests ============

class TestImportConfirmSimpleMode:
    """Test import confirm in simple mode creates batches."""

    def test_simple_mode_in_creates_batch(self, admin_client, sample_material):
        """Simple mode in-operation should create a new batch."""
        from database import get_db_connection

        # Count batches before
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) as c FROM batches WHERE material_id = ?', (sample_material['id'],))
        before = cursor.fetchone()['c']
        conn.close()

        resp = admin_client.post("/api/materials/import-excel/confirm", json={
            "changes": [{
                "sku": sample_material['sku'],
                "name": sample_material['name'],
                "category": "Test",
                "unit": "个",
                "safe_stock": 20,
                "location": "C区-01",
                "current_quantity": sample_material['quantity'],
                "import_quantity": sample_material['quantity'] + 30,
                "difference": 30,
                "operation": "in",
            }],
            "reason_category": "purchase",
            "is_batch_mode": False,
            "warehouse_id": sample_material['warehouse_id'],
        })
        data = resp.json()
        assert data['success'] is True
        assert data['in_count'] == 1

        # Check batch was created
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) as c FROM batches WHERE material_id = ?', (sample_material['id'],))
        after = cursor.fetchone()['c']
        conn.close()
        assert after > before

    def test_simple_mode_out_uses_fifo(self, admin_client, material_with_legacy_batch):
        """Simple mode out-operation should consume batches via FIFO."""
        from database import get_db_connection

        mat = material_with_legacy_batch
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT quantity FROM materials WHERE id = ?', (mat['id'],))
        mat_qty = cursor.fetchone()['quantity']
        conn.close()

        if mat_qty < 5:
            pytest.skip("Not enough quantity to test out operation")

        resp = admin_client.post("/api/materials/import-excel/confirm", json={
            "changes": [{
                "sku": mat['sku'],
                "name": mat['name'],
                "category": "Test",
                "unit": "个",
                "safe_stock": 20,
                "location": "",
                "current_quantity": mat_qty,
                "import_quantity": mat_qty - 5,
                "difference": -5,
                "operation": "out",
            }],
            "reason_category": "sell",
            "is_batch_mode": False,
            "warehouse_id": mat['warehouse_id'],
        })
        data = resp.json()
        assert data['success'] is True
        assert data['out_count'] == 1

        # Check batch_consumptions were created
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT COUNT(*) as c FROM batch_consumptions bc
            JOIN inventory_records r ON bc.record_id = r.id
            WHERE r.material_id = ? AND r.reason_category = 'sell'
        ''', (mat['id'],))
        count = cursor.fetchone()['c']
        conn.close()
        assert count >= 1

    def test_simple_mode_new_contact_created(self, admin_client, sample_material):
        """Confirm with unknown contact_name should auto-create the contact."""
        from database import get_db_connection

        resp = admin_client.post("/api/materials/import-excel/confirm", json={
            "changes": [{
                "sku": sample_material['sku'],
                "name": sample_material['name'],
                "category": "Test",
                "unit": "个",
                "safe_stock": 20,
                "location": "",
                "current_quantity": None,
                "import_quantity": 10,
                "difference": 10,
                "operation": "in",
                "contact_name": "AutoCreated Supplier",
            }],
            "reason_category": "purchase",
            "is_batch_mode": False,
            "warehouse_id": sample_material['warehouse_id'],
        })
        data = resp.json()
        assert data['success'] is True

        # Verify contact was created
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, is_supplier FROM contacts WHERE name = 'AutoCreated Supplier'")
        contact = cursor.fetchone()
        conn.close()
        assert contact is not None
        assert contact['is_supplier'] == 1


class TestImportConfirmBatchMode:
    """Test import confirm in batch mode."""

    def test_batch_mode_new_batch_created(self, admin_client, material_with_batch):
        """New batch (empty batch_no) should create a batch record."""
        from database import get_db_connection

        # 单一真相源：库存来自 active batches 聚合
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            'SELECT COALESCE(SUM(quantity),0) AS q FROM batches '
            'WHERE material_id = ? AND is_exhausted = 0',
            (material_with_batch['id'],),
        )
        mat_qty = cursor.fetchone()['q']
        conn.close()

        resp = admin_client.post("/api/materials/import-excel/confirm", json={
            "changes": [{
                "sku": material_with_batch['sku'],
                "name": material_with_batch['name'],
                "category": "Test",
                "unit": "個",
                "safe_stock": 20,
                "location": "D区-01",
                "current_quantity": 0,
                "import_quantity": 25,
                "difference": 25,
                "operation": "in",
                "is_batch_new": True,
            }],
            "reason_category": "purchase",
            "is_batch_mode": True,
            "warehouse_id": material_with_batch['warehouse_id'],
        })
        data = resp.json()
        assert data['success'] is True
        assert data['in_count'] == 1

        # Verify material quantity increased — read active batch sum
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            'SELECT COALESCE(SUM(quantity),0) AS q FROM batches '
            'WHERE material_id = ? AND is_exhausted = 0',
            (material_with_batch['id'],),
        )
        new_qty = cursor.fetchone()['q']
        conn.close()
        assert new_qty == mat_qty + 25

    def test_batch_mode_existing_batch_updated(self, admin_client, material_with_batch):
        """Existing batch with changed quantity should be updated."""
        from database import get_db_connection

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT quantity FROM batches WHERE batch_no = ?', (material_with_batch['batch_no'],))
        old_batch_qty = cursor.fetchone()['quantity']
        conn.close()

        resp = admin_client.post("/api/materials/import-excel/confirm", json={
            "changes": [{
                "sku": material_with_batch['sku'],
                "name": material_with_batch['name'],
                "category": "Test",
                "unit": "個",
                "safe_stock": 20,
                "location": "A区-01",
                "current_quantity": old_batch_qty,
                "import_quantity": old_batch_qty + 10,
                "difference": 10,
                "operation": "in",
                "batch_no": material_with_batch['batch_no'],
            }],
            "reason_category": "purchase",
            "is_batch_mode": True,
            "warehouse_id": material_with_batch['warehouse_id'],
        })
        data = resp.json()
        assert data['success'] is True

        # Verify batch quantity updated
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT quantity FROM batches WHERE batch_no = ?', (material_with_batch['batch_no'],))
        new_batch_qty = cursor.fetchone()['quantity']
        conn.close()
        assert new_batch_qty == old_batch_qty + 10


# ============ Round-trip Tests ============

class TestExportImportRoundtrip:
    """Test export → import round-trip consistency."""

    def test_roundtrip_no_changes(self, admin_client, material_with_batch):
        """Export then immediately import the same file should show all 'none' for batch items."""
        # Export
        resp = admin_client.get("/api/materials/export-excel")
        assert resp.status_code == 200

        # Import preview
        data = _upload_preview(admin_client, BytesIO(resp.content))
        assert data['success'] is True
        assert data['is_batch_mode'] is True

        # Find our specific material's items — they should be 'none'
        our_items = [item for item in data['preview'] if item['sku'] == material_with_batch['sku']]
        assert len(our_items) >= 1
        for item in our_items:
            assert item['operation'] == 'none', \
                f"Expected 'none' for {item['sku']} batch={item.get('batch_no')}, got '{item['operation']}' (diff={item['difference']})"


class TestBatchesWarehouseAccess:
    """Test warehouse-scoped access control on batches endpoint."""

    def test_cross_warehouse_batches_access_denied(self, admin_client, default_warehouse_id):
        """User with access to warehouse A cannot access batches from warehouse B."""
        from database import get_db_connection
        import uuid

        conn = get_db_connection()
        cursor = conn.cursor()

        # Create a second warehouse
        wh_slug = f"wh-{uuid.uuid4().hex[:6]}"
        cursor.execute(
            'INSERT INTO warehouses (slug, name, is_default) VALUES (?, ?, 0)',
            (wh_slug, f"Second Warehouse {wh_slug}")
        )
        wh_b_id = cursor.lastrowid
        conn.commit()

        # Create a non-admin user with access only to default warehouse
        username = f"operator-{uuid.uuid4().hex[:6]}"
        password = "OpPass123!"
        cursor.execute(
            '''INSERT INTO users (username, password_hash, role, display_name, created_at)
               VALUES (?, ?, 'operate', ?, CURRENT_TIMESTAMP)''',
            (username, 'dummy_hash_for_test', username)
        )
        user_id = cursor.lastrowid
        cursor.execute(
            'INSERT INTO user_warehouses (user_id, warehouse_id) VALUES (?, ?)',
            (user_id, default_warehouse_id)
        )
        conn.commit()

        # Create material in warehouse B
        sku = f"XWH-{uuid.uuid4().hex[:8].upper()}"
        name = f"CrossWH Material {sku}"
        cursor.execute(
            '''INSERT INTO materials (name, sku, category, quantity, unit, warehouse_id)
               VALUES (?, ?, 'Test', 100, 'pcs', ?)''',
            (name, sku, wh_b_id)
        )
        mat_id = cursor.lastrowid

        # Create a batch in warehouse B
        batch_no = f"BATCH-XWH-{uuid.uuid4().hex[:6]}"
        cursor.execute(
            '''INSERT INTO batches (batch_no, material_id, quantity, initial_quantity, warehouse_id, created_at)
               VALUES (?, ?, 50, 50, ?, CURRENT_TIMESTAMP)''',
            (batch_no, mat_id, wh_b_id)
        )
        conn.commit()
        conn.close()

        # Login as restricted user (directly set real password hash)
        from app import hash_password
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            'UPDATE users SET password_hash = ? WHERE id = ?',
            (hash_password(password), user_id)
        )
        conn.commit()
        conn.close()

        # Login as restricted user
        from fastapi.testclient import TestClient
        import app as app_module
        c = TestClient(app_module.app)
        resp = c.post("/api/auth/login", json={"username": username, "password": password})
        assert resp.status_code == 200
        assert resp.json()['success'] is True

        # Try to access batches from warehouse B - should get 403
        resp = c.get("/api/materials/batches", params={
            "name": name,
            "warehouse_id": wh_b_id
        })
        assert resp.status_code == 403
        detail = resp.json().get('detail') or resp.json().get('error')
        assert "无权访问" in detail

        # Cleanup: delete test user and warehouse
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM user_warehouses WHERE user_id = ?', (user_id,))
        # Delete sessions first to satisfy FK on MySQL (sqlite ignores FK by default)
        cursor.execute('DELETE FROM sessions WHERE user_id = ?', (user_id,))
        cursor.execute('DELETE FROM users WHERE id = ?', (user_id,))
        cursor.execute('DELETE FROM batches WHERE material_id = ?', (mat_id,))
        cursor.execute('DELETE FROM materials WHERE id = ?', (mat_id,))
        cursor.execute('DELETE FROM warehouses WHERE id = ?', (wh_b_id,))
        conn.commit()
        conn.close()
