"""
仓库管理系统 FastAPI 后端
"""
import os
import logging
import secrets
import sqlite3
import httpx
from fastapi import FastAPI, Query, HTTPException, File, UploadFile, Request, Response, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from starlette.middleware.base import BaseHTTPMiddleware
from datetime import datetime, timedelta
from contextlib import contextmanager
from typing import Any, Dict, List, Optional
from pydantic import BaseModel, Field
from io import BytesIO
from functools import wraps
from enum import Enum, IntEnum

# 速率限制
from slowapi import Limiter
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

from database import (
    init_database, generate_mock_data, get_db_connection,
    has_admin_user, hash_password, verify_password,
    generate_session_token, generate_api_key, hash_api_key,
    generate_batch_no, needs_password_rehash,
    validate_username, validate_password_strength,
    REASON_CATEGORIES, REASON_CATEGORY_LABELS,
    get_deploy_mode, _is_sqlite,
)
from models import (
    DashboardStats, CategoryItem, WeeklyTrend, TopStock, LowStockItem,
    MaterialItem, ProductStats, ProductRecord,
    StockOperationRequest, StockOperationResponse, StockOperationProduct,
    ImportPreviewItem, ExcelImportPreviewResponse, ExcelImportConfirm,
    ExcelImportResponse, ManualRecordRequest, MissingSkuItem,
    PaginatedMaterialsResponse, PaginatedRecordsResponse, MaterialItemWithDisabled,
    InventoryRecordItem, PaginatedProductRecordsResponse,
    # Auth models
    AuthStatusResponse, UserInfo, SetupRequest, LoginRequest, LoginResponse,
    CreateUserRequest, UpdateUserRequest, UserListItem,
    # Registration models
    VerifyDeviceRequest, VerifyDeviceResponse, RegisterRequest, ResetPasswordRequest,
    CreateApiKeyRequest, ApiKeyStatusRequest, ApiKeyResponse, ApiKeyListItem,
    # Contact models
    CreateContactRequest, UpdateContactRequest, ContactItem, ContactListItem,
    PaginatedContactsResponse,
    # Batch models
    BatchInfo, BatchConsumption, StockInResponse, StockOutResponse,
    BatchMoveRequest, BatchMoveResponse,
    BatchDetailItem, BatchDetailResponse,
    # Operator model
    OperatorListItem,
    # Database management models
    DatabaseClearRequest, DatabaseOperationResponse,
    # MCP models
    CreateMCPConnectionRequest, UpdateMCPConnectionRequest,
    MCPConnectionItem, MCPConnectionResponse,
    # Fuzzy match models
    FuzzyMatchCandidate, FuzzyMatchResponse,
    # Tenant models
    TenantItem, CreateTenantRequest, UpdateTenantRequest,
    # Warehouse models
    WarehouseItem, CreateWarehouseRequest, UpdateWarehouseRequest,
    UserWarehouseAssignment,
    # R3: wire-format string enums
    RoleName, RecordType,
)
from fuzzy_match import FuzzyMatcher
from sqlalchemy import select, and_, or_, insert, update, delete, case, text, false
from sqlalchemy.exc import IntegrityError
from db import get_engine
from metadata import (
    warehouses as _t_warehouses,
    user_warehouses as _t_user_warehouses,
    users as _t_users,
    sessions as _t_sessions,
    tenants as _t_tenants,
    contacts as _t_contacts,
    materials as _t_materials,
    batches as _t_batches,
    inventory_records as _t_inventory_records,
    batch_consumptions as _t_batch_consumptions,
    api_keys as _t_api_keys,
    system_settings as _t_system_settings,
    erp_providers as _t_erp_providers,
    mcp_connections as _t_mcp_connections,
    # 5 个 face 表别名已随 routers/face.py 一起搬走（commit 3fd71ce）。
    # app.py 不再直接引用 face 表；如有其他路由确实需要，请直接 import 全名。
)
from sqlalchemy import func as _sa_func
import math
import uuid

# Excel处理
from openpyxl import Workbook, load_workbook

# MCP进程管理
from mcp_manager import MCPProcessManager

# Shared dependencies (extracted from app.py — Phase 1 split, task #5).
# Re-exported so existing route handlers in app.py continue to reference
# these as bare names without qualification.
from deps import (
    get_db,
    Role,
    Resource,
    Action,
    CurrentUser,
    get_current_user,
    require_permission,
    load_or_404,
    ROLE_LEVELS,
    _ACTION_TO_ROLE,
    audit_log,
    resolve_warehouse_id,
    check_warehouse_access,
    build_scope_predicates,
    build_authorized_scope_predicates,
    assert_row_in_scope,
    infer_single_writable_warehouse_id,
    ensure_contact_tenant,
    require_warehouse_id,
    resolve_tenant_id_for_write,
)

# ============================================
# 环境变量配置
# ============================================
# CORS配置：逗号分隔的域名列表，或 * 表示允许所有
CORS_ORIGINS = os.environ.get('CORS_ORIGINS', '*')
# 是否生成模拟数据（生产环境设为false）
INIT_MOCK_DATA = os.environ.get('INIT_MOCK_DATA', 'true').lower() == 'true'
# 是否启用安全响应头
ENABLE_SECURITY_HEADERS = os.environ.get('ENABLE_SECURITY_HEADERS', 'false').lower() == 'true'
# 是否启用审计日志
ENABLE_AUDIT_LOG = os.environ.get('ENABLE_AUDIT_LOG', 'true').lower() == 'true'
# Factory 设备代理 API
FACTORY_API_KEY = os.environ.get('FACTORY_API_KEY', '')
FACTORY_API_BASE_URL = os.environ.get('FACTORY_API_BASE_URL', 'https://watcher-agent-api.seeed.cc')
# Excel上传限制
MAX_UPLOAD_SIZE_MB = int(os.environ.get('MAX_UPLOAD_SIZE_MB', '10'))
MAX_IMPORT_ROWS = int(os.environ.get('MAX_IMPORT_ROWS', '10000'))
# 模糊匹配置信度阈值
FUZZY_CONFIDENT_SCORE = float(os.environ.get('FUZZY_CONFIDENT_SCORE', '80'))
FUZZY_CONFIDENT_GAP = float(os.environ.get('FUZZY_CONFIDENT_GAP', '10'))

# 配置日志
logging.basicConfig(
    level=os.environ.get('LOG_LEVEL', 'INFO').upper(),
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger('warehouse')

# ============================================
# 速率限制配置
# ============================================
limiter = Limiter(key_func=get_remote_address, enabled=os.environ.get('DISABLE_RATE_LIMIT', '0') != '1')

# 创建 FastAPI 应用
app = FastAPI(
    title="仓库管理系统 API",
    description="智能硬件仓库管理系统后端 API",
    version="2.0.0"
)


@app.get("/health")
async def health_check():
    return {"status": "ok"}

# 注册速率限制异常处理（带 CORS 头）
app.state.limiter = limiter

async def rate_limit_exceeded_handler(request: Request, exc: RateLimitExceeded):
    """自定义速率限制异常处理器，确保响应包含 CORS 头"""
    from starlette.responses import JSONResponse
    origin = request.headers.get("origin", "*")
    response = JSONResponse(
        status_code=429,
        content={"detail": f"Rate limit exceeded: {exc.detail}"}
    )
    # 添加 CORS 头
    response.headers["Access-Control-Allow-Origin"] = origin
    response.headers["Access-Control-Allow-Credentials"] = "true"
    return response

app.add_exception_handler(RateLimitExceeded, rate_limit_exceeded_handler)

# ============================================
# 安全头中间件
# ============================================
class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    """添加安全响应头"""
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        if ENABLE_SECURITY_HEADERS:
            response.headers['X-Content-Type-Options'] = 'nosniff'
            response.headers['X-Frame-Options'] = 'DENY'
            response.headers['X-XSS-Protection'] = '1; mode=block'
            response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
        return response

app.add_middleware(SecurityHeadersMiddleware)

# ============================================
# CORS 配置
# ============================================
def get_cors_origins():
    """解析CORS配置"""
    if CORS_ORIGINS == '*':
        return ['*']
    return [origin.strip() for origin in CORS_ORIGINS.split(',') if origin.strip()]

cors_origins = get_cors_origins()

# 自定义 CORS 中间件：正确处理通配符和 credentials
class DynamicCORSMiddleware(BaseHTTPMiddleware):
    """
    动态 CORS 中间件，解决以下问题：
    1. 当 allow_origins=['*'] 时，自动将 Access-Control-Allow-Origin 设为请求的 Origin
    2. 确保 credentials 模式下不返回通配符
    3. 正确处理预检请求（OPTIONS）
    """
    async def dispatch(self, request: Request, call_next):
        origin = request.headers.get("origin")

        # 处理预检请求
        if request.method == "OPTIONS":
            response = Response(status_code=204)
        else:
            response = await call_next(request)

        # 设置 CORS 头
        if origin:
            if CORS_ORIGINS == '*':
                # 通配符模式：使用请求的 Origin
                response.headers["Access-Control-Allow-Origin"] = origin
            elif origin in cors_origins:
                # 明确列表模式：只允许列表中的 Origin
                response.headers["Access-Control-Allow-Origin"] = origin
            else:
                # Origin 不在允许列表中，不设置 CORS 头（浏览器会拒绝）
                pass

        # 设置其他 CORS 头
        if "Access-Control-Allow-Origin" in response.headers:
            response.headers["Access-Control-Allow-Credentials"] = "true"
            response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS, PATCH"
            response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization, X-API-Key"
            response.headers["Access-Control-Max-Age"] = "86400"

        return response

# 使用自定义 CORS 中间件替代 FastAPI 的 CORSMiddleware
app.add_middleware(DynamicCORSMiddleware)

# Request 日志中间件 — 绕过 uvicorn.access 被吞的问题，每个请求都强制打一行
if os.environ.get('REQUEST_LOG', '1') != '0':
    import time as _time
    _req_logger = logging.getLogger('warehouse.request')

    @app.middleware("http")
    async def _log_requests(request, call_next):
        _t0 = _time.perf_counter()
        try:
            response = await call_next(request)
            _dt = (_time.perf_counter() - _t0) * 1000
            _req_logger.info(
                "%s %s -> %d  %.1fms  client=%s",
                request.method, request.url.path, response.status_code,
                _dt, request.client.host if request.client else '-'
            )
            return response
        except Exception:
            _dt = (_time.perf_counter() - _t0) * 1000
            _req_logger.exception(
                "%s %s -> EXC  %.1fms", request.method, request.url.path, _dt
            )
            raise

# ============================================
# 审计日志函数
# ============================================
# audit_log moved to deps.py (Phase 2 prep, task #6) — re-exported via the
# ``from deps import (...)`` block above.

# ============================================
# 初始化数据库
# ============================================
# NOTE (Phase 3g cleanup): schema management has moved to Alembic. The
# ``init_database()`` function is intentionally kept (sqlite-only DDL helper)
# because tests/conftest.py and a couple of legacy fixtures still call it
# directly. It is no longer invoked at FastAPI startup — see the
# ``_run_migrations`` startup hook below, which calls ``alembic upgrade head``
# and works on both SQLite and MySQL. ``generate_mock_data()`` was likewise
# moved into the startup hook so it runs *after* the schema exists.


# 数据库连接上下文管理器 — moved to backend/deps.py
# (re-exported via the top-of-file `from deps import ... get_db` below)


# FuzzyMatcher 全局实例
def get_fuzzy_matcher() -> FuzzyMatcher:
    """获取或创建 FuzzyMatcher 实例"""
    if not hasattr(app.state, 'fuzzy_matcher'):
        app.state.fuzzy_matcher = FuzzyMatcher(
            get_db_connection,
            confident_score=FUZZY_CONFIDENT_SCORE,
            confident_gap=FUZZY_CONFIDENT_GAP,
        )
    return app.state.fuzzy_matcher


# 自定义异常处理（保持响应格式兼容）
@app.exception_handler(HTTPException)
async def http_exception_handler(request, exc):
    return JSONResponse(
        status_code=exc.status_code,
        content={"error": exc.detail}
    )


# ============ 认证相关 ============

_VALID_ROLE_VALUES = {r.value for r in RoleName}  # {'admin', 'operate', 'view'}


# ROLE_LEVELS / Role / Resource / Action / CurrentUser / get_current_user /
# require_permission / load_or_404 — all moved to backend/deps.py
# (re-exported via the top-of-file ``from deps import ...`` below).
# Code preserved below has been deleted: continues at "仓库上下文辅助".


def _route_has_guard(dep) -> bool:
    """Recursively check a Dependant tree for a require_permission marker."""
    call = getattr(dep, "call", None)
    if call is not None and getattr(call, "__perm_marker__", False):
        return True
    for sub in getattr(dep, "dependencies", []) or []:
        if _route_has_guard(sub):
            return True
    return False


def _seed_base_data() -> None:
    """幂等补种：确保 tenant 1 和默认仓库存在。

    Docker 走纯 Alembic 路径，迁移只建表结构不插数据；
    本地 start.sh 走 init_database()，已有种子数据（INSERT OR IGNORE）。
    两条路径都调用此函数，冲突忽略保证幂等。
    """
    try:
        sqlite = _is_sqlite()
        ignore_kw = "OR IGNORE" if sqlite else "IGNORE"
        with get_engine().begin() as conn:
            conn.execute(text(
                f"INSERT {ignore_kw} INTO tenants (id, slug, name, is_active) "
                "VALUES (1, :slug, :name, 1)"
            ), {"slug": "default", "name": "默认租户"})
            conn.execute(text(
                f"INSERT {ignore_kw} INTO warehouses "
                "(id, slug, name, is_default, is_disabled, tenant_id) "
                "VALUES (1, :slug, :name, 1, 0, 1)"
            ), {"slug": "default", "name": "默认仓库"})
    except Exception as e:
        logger.warning(f"_seed_base_data() skipped: {e}")


def _validate_schema_matches_metadata() -> None:
    """启动时校验：SQLAlchemy metadata 声明的每一列都真实存在于 db。

    背景：曾踩坑两次——开发者在 metadata.py 给现有表加新列后忘写对应的
    Alembic 迁移，启动时 `alembic upgrade head` 因为没有新脚本是 no-op，
    问题被推迟到运行时第一次查询该列才报 `no such column` 500。

    这个校验把"模型加列但没迁移"的失败提到启动阶段，让 CI / 本地运行立刻崩，
    强制开发者补迁移再启动。比运行时 500 更早、更明确。
    """
    from sqlalchemy import inspect as sa_inspect
    from metadata import metadata as _md

    inspector = sa_inspect(get_engine())
    actual_tables = set(inspector.get_table_names())
    problems: list[str] = []

    for table in _md.sorted_tables:
        if table.name not in actual_tables:
            # 表整体缺失：通常是 Alembic 还没创建到此表。比缺列更严重。
            problems.append(f"表 {table.name} 在 db 中不存在")
            continue
        actual_cols = {c["name"] for c in inspector.get_columns(table.name)}
        expected_cols = {c.name for c in table.columns}
        missing = expected_cols - actual_cols
        if missing:
            problems.append(
                f"表 {table.name} 缺少列 {sorted(missing)}（metadata.py 已声明但 db 中没有，"
                f"很可能 metadata.py 加了列却忘写 Alembic 迁移）"
            )

    if problems:
        raise RuntimeError(
            "DB schema 与 metadata 不一致，拒绝启动：\n  - "
            + "\n  - ".join(problems)
            + "\n请补一个 Alembic 迁移：cd backend && alembic revision -m '...' "
              "然后在 upgrade() 里 op.add_column(...)，再 alembic upgrade head。"
        )


def _validate_deploy_mode_invariants() -> None:
    """启动时校验 DEPLOY_MODE 与 DB 状态自洽，否则拒绝启动。

    single_tenant 不变式：
      - 至多一个 active 租户（id=1）
      - 不存在 global admin（users.tenant_id IS NULL AND role='admin'）
    违反任何一条都说明部署元信息与数据不一致，继续启动会让 UI 各处判断分裂。
    """
    if get_deploy_mode() != 'single_tenant':
        return
    with get_engine().connect() as sa_conn:
        extra_tenants = sa_conn.execute(
            text("SELECT COUNT(*) FROM tenants WHERE is_active = 1 AND id != 1")
        ).scalar()
        if extra_tenants and extra_tenants > 0:
            raise RuntimeError(
                f'Refusing to start: DEPLOY_MODE=single_tenant but {extra_tenants} extra active '
                f'tenant(s) exist. Switch to multi_tenant or consolidate tenants.'
            )
        global_admins = sa_conn.execute(
            text("SELECT COUNT(*) FROM users WHERE role = 'admin' AND tenant_id IS NULL")
        ).scalar()
        if global_admins and global_admins > 0:
            raise RuntimeError(
                f'Refusing to start: DEPLOY_MODE=single_tenant but {global_admins} global admin '
                f'user(s) exist (users.tenant_id IS NULL). Either set DEPLOY_MODE=multi_tenant, '
                f'or bind them to a tenant: '
                f'UPDATE users SET tenant_id=1 WHERE role="admin" AND tenant_id IS NULL.'
            )


def _audit_routes() -> None:
    """Walk all FastAPI routes; warn for any route that has no auth guard.

    Exempts auth bootstrap endpoints, docs, health checks, and the SPA
    catch-all. Wrapped in ``try/except`` at the call-site so that audit
    bugs never block startup.
    """
    exempt_prefixes = (
        "/api/auth/setup",
        "/api/auth/login",
        "/api/auth/logout",
        "/api/auth/status",
        "/api/auth/register",        # 自助注册（multi_tenant）
        "/api/auth/reset-password",  # 凭设备ID重置密码
        "/api/system/mode",  # 部署元信息（single/multi tenant），前端在登录前就要据此渲染 UI
        "/factory/devices",  # 工厂设备代理使用 X-Factory-Key 做接口级鉴权
        "/docs",
        "/openapi.json",
        "/redoc",
        "/health",
    )
    unguarded: List[str] = []
    for route in app.routes:
        if not hasattr(route, "dependant"):
            continue
        path = getattr(route, "path", "")
        if any(path.startswith(p) for p in exempt_prefixes):
            continue
        if path == "/{path:path}":  # SPA catch-all
            continue
        if _route_has_guard(route.dependant):
            continue
        methods = ",".join(sorted(getattr(route, "methods", []) or []))
        func_name = getattr(getattr(route, "endpoint", None), "__name__", "?")
        unguarded.append(f"{methods} {path} ({func_name})")
    if unguarded:
        raise RuntimeError(
            "Unguarded routes detected at startup — every non-exempt route "
            "must use Depends(require_permission(Resource, Action)):\n  "
            + "\n  ".join(unguarded)
        )




# ============ 仓库上下文辅助 ============

# resolve_warehouse_id moved to deps.py (Phase 2 prep, task #6).


# Write-scope helpers (infer_single_writable_warehouse_id / ensure_contact_tenant /
# require_warehouse_id / resolve_tenant_id_for_write) moved to deps.py — Phase 3 prep
# (task #7) so future routers can import them without reaching back into app.py.
# build_warehouse_filter (legacy raw-SQL builder) had no callers and was dropped.
# check_warehouse_access / build_scope_predicates / build_authorized_scope_predicates
# previously moved to deps.py in commit d193d61.


# ============ 仓库管理 API ============

@app.get("/api/warehouses", response_model=List[WarehouseItem])
async def list_warehouses(
    include_disabled: bool = False,
    current_user: CurrentUser = Depends(require_permission(Resource.WAREHOUSES, Action.READ))
):
    """获取仓库列表 — Phase 2b: read via SQLAlchemy Core."""
    conds = list(build_scope_predicates(_t_warehouses, current_user.tenant_id, None))
    if current_user.role != RoleName.ADMIN:
        warehouse_ids = current_user.get_authorized_warehouses(None)
        conds.append(_t_warehouses.c.id.in_(warehouse_ids) if warehouse_ids else false())
    if not (include_disabled and current_user.role == RoleName.ADMIN):
        conds.append(_t_warehouses.c.is_disabled == 0)
    stmt = select(
        _t_warehouses.c.id, _t_warehouses.c.slug, _t_warehouses.c.name,
        _t_warehouses.c.address, _t_warehouses.c.is_default,
        _t_warehouses.c.is_disabled, _t_warehouses.c.created_at,
        _t_warehouses.c.tenant_id,
        _t_tenants.c.name.label('tenant_name'),
    ).select_from(
        _t_warehouses.outerjoin(_t_tenants, _t_warehouses.c.tenant_id == _t_tenants.c.id)
    )
    if conds:
        stmt = stmt.where(and_(*conds))
    stmt = stmt.order_by(_t_warehouses.c.is_default.desc(), _t_warehouses.c.id.asc())
    with get_engine().connect() as sa_conn:
        rows = sa_conn.execute(stmt).fetchall()
    return [WarehouseItem(
        id=r.id, slug=r.slug, name=r.name,
        address=r.address, is_default=bool(r.is_default),
        is_disabled=bool(r.is_disabled),
        created_at=(r.created_at.strftime('%Y-%m-%d %H:%M:%S')
                    if isinstance(r.created_at, datetime) else r.created_at),
        tenant_id=r.tenant_id,
        tenant_name=r.tenant_name,
    ) for r in rows]


# ---- warehouses CREATE/UPDATE/DELETE migrated to ResourceRouter (R2 phase 2) ----
# LIST stays as ``list_warehouses`` above (uses tenant join + scope predicates).
# GET-by-id was never part of the public API for warehouses, so we suppress
# the factory's GET registration via ``enable_get=False``.

_WAREHOUSE_OUT_COLUMNS = [
    _t_warehouses.c.id, _t_warehouses.c.slug, _t_warehouses.c.name,
    _t_warehouses.c.address, _t_warehouses.c.is_default,
    _t_warehouses.c.is_disabled, _t_warehouses.c.created_at,
    _t_warehouses.c.tenant_id,
]


def _warehouse_to_out(row) -> WarehouseItem:
    ca = row.created_at
    if isinstance(ca, datetime):
        ca = ca.strftime('%Y-%m-%d %H:%M:%S')
    return WarehouseItem(
        id=row.id, slug=row.slug, name=row.name,
        address=row.address, is_default=bool(row.is_default),
        is_disabled=bool(row.is_disabled),
        created_at=ca,
        tenant_id=row.tenant_id,
    )


def _warehouse_before_create(sa_conn, current_user, request: CreateWarehouseRequest):
    import re as _re
    if not _re.match(r'^[a-z0-9][a-z0-9\-]*$', request.slug):
        raise HTTPException(
            status_code=400,
            detail="仓库标识只能包含小写字母、数字和连字符，且不能以连字符开头"
        )
    target_tenant_id = request.tenant_id if current_user.tenant_id is None else current_user.tenant_id
    slug_filter = and_(
        _t_warehouses.c.slug == request.slug,
        _t_warehouses.c.tenant_id == target_tenant_id,
    )
    existing = sa_conn.execute(select(_t_warehouses.c.id).where(slug_filter)).first()
    if existing:
        raise HTTPException(status_code=400, detail="仓库标识已存在")


def _warehouse_values_for_create(sa_conn, current_user, request: CreateWarehouseRequest) -> dict:
    if current_user.tenant_id is None:
        if not request.tenant_id:
            raise HTTPException(status_code=400, detail="全局管理员创建仓库时必须指定 tenant_id")
        wh_tenant_id = request.tenant_id
        tenant_row = sa_conn.execute(
            select(_t_tenants.c.id).where(
                and_(_t_tenants.c.id == wh_tenant_id, _t_tenants.c.is_active == 1)
            )
        ).first()
        if not tenant_row:
            raise HTTPException(status_code=400, detail="租户不存在或已停用")
    else:
        if request.tenant_id is not None and request.tenant_id != current_user.tenant_id:
            raise HTTPException(status_code=403, detail="无权在其他租户下创建仓库")
        wh_tenant_id = current_user.tenant_id

    return {
        "slug": request.slug,
        "name": request.name,
        "address": request.address,
        "tenant_id": wh_tenant_id,
        "created_at": datetime.now(),
    }


def _warehouse_values_for_update(sa_conn, current_user, request: UpdateWarehouseRequest, row) -> dict:
    # ``row`` carries ``is_default`` because the registration sets
    # ``load_columns`` to include it — atomic with the scope check.
    values: dict = {}
    if request.name is not None:
        values['name'] = request.name
    if request.address is not None:
        values['address'] = request.address
    if request.is_disabled is not None:
        if row.is_default and request.is_disabled:
            raise HTTPException(status_code=400, detail="不能禁用默认仓库")
        values['is_disabled'] = 1 if request.is_disabled else 0
    return values


def _warehouse_before_delete(sa_conn, current_user, row):
    # ``row.is_default`` available via load_columns — atomic with scope check.
    if row.is_default:
        raise HTTPException(status_code=400, detail="不能删除默认仓库")
    n = sa_conn.execute(
        select(_sa_func.count()).select_from(_t_materials).where(
            and_(_t_materials.c.warehouse_id == row.id, _t_materials.c.is_disabled == 0)
        )
    ).scalar()
    if n and n > 0:
        raise HTTPException(status_code=400, detail="仓库内仍有物料，无法删除")
    sa_conn.execute(
        update(_t_warehouses).where(_t_warehouses.c.id == row.id).values(is_disabled=1)
    )


from resource_router import ResourceRouter as _ResourceRouterWH  # noqa: E402

_wh_router = _ResourceRouterWH(
    app=app,
    prefix="/api/warehouses",
    table=_t_warehouses,
    response_model=WarehouseItem,
    create_model=CreateWarehouseRequest,
    update_model=UpdateWarehouseRequest,
    permission_read=require_permission(Resource.WAREHOUSES, Action.READ),
    permission_write=require_permission(Resource.WAREHOUSES, Action.ADMIN),
    not_found_detail="仓库不存在",
    forbidden_detail="无权操作该仓库",
    to_out=_warehouse_to_out,
    values_for_create=_warehouse_values_for_create,
    values_for_update=_warehouse_values_for_update,
    before_create=_warehouse_before_create,
    before_delete=_warehouse_before_delete,
    list_handler=None,
    get_columns=_WAREHOUSE_OUT_COLUMNS,
    update_select_columns=_WAREHOUSE_OUT_COLUMNS,
    # Load id+tenant_id+is_default atomically so before_delete /
    # values_for_update can read is_default without a second SELECT.
    load_columns=[
        _t_warehouses.c.id, _t_warehouses.c.tenant_id, _t_warehouses.c.is_default,
    ],
    enable_get=False,
    delete_response={"success": True, "message": "仓库已禁用"},
)
_wh_router.register()


@app.get("/api/users/{user_id}/warehouses")
async def get_user_warehouses(
    user_id: int,
    current_user: CurrentUser = Depends(require_permission(Resource.USERS, Action.ADMIN))
):
    """获取用户授权的仓库列表 — Phase 2b: read via SQLAlchemy Core."""
    with get_engine().connect() as sa_conn:
        # 验证用户属于当前租户
        target_user = load_or_404(
            sa_conn, _t_users, user_id,
            columns=[_t_users.c.id, _t_users.c.tenant_id],
            not_found="用户不存在",
            tenant_id=current_user.tenant_id,
            forbidden="无权访问其他租户的用户",
        )
        stmt = select(
            _t_warehouses.c.id, _t_warehouses.c.slug, _t_warehouses.c.name,
        ).select_from(
            _t_user_warehouses.join(
                _t_warehouses, _t_user_warehouses.c.warehouse_id == _t_warehouses.c.id
            )
        ).where(_t_user_warehouses.c.user_id == user_id)
        rows = sa_conn.execute(stmt).fetchall()
    return {"warehouse_ids": [r.id for r in rows],
            "warehouses": [{"id": r.id, "slug": r.slug, "name": r.name} for r in rows]}


@app.put("/api/users/{user_id}/warehouses")
async def set_user_warehouses(
    user_id: int,
    request: UserWarehouseAssignment,
    current_user: CurrentUser = Depends(require_permission(Resource.USERS, Action.ADMIN))
):
    """设置用户授权的仓库列表"""
    with get_engine().begin() as sa_conn:
        # 验证用户属于当前租户
        target_user = load_or_404(
            sa_conn, _t_users, user_id,
            columns=[_t_users.c.id, _t_users.c.tenant_id],
            not_found="用户不存在",
            tenant_id=current_user.tenant_id,
            forbidden="无权访问其他租户的用户",
        )

        # 验证所有仓库ID存在
        for wh_id in request.warehouse_ids:
            if current_user.tenant_id is None:
                wh_check = sa_conn.execute(
                    select(_t_warehouses.c.id).where(_t_warehouses.c.id == wh_id)
                ).first()
            else:
                wh_check = sa_conn.execute(
                    select(_t_warehouses.c.id).where(
                        and_(_t_warehouses.c.id == wh_id,
                             _t_warehouses.c.tenant_id == current_user.tenant_id)
                    )
                ).first()
            if not wh_check:
                raise HTTPException(status_code=400, detail=f"仓库ID {wh_id} 不存在")

        # 替换授权
        sa_conn.execute(
            delete(_t_user_warehouses).where(_t_user_warehouses.c.user_id == user_id)
        )
        for wh_id in request.warehouse_ids:
            sa_conn.execute(
                insert(_t_user_warehouses).values(user_id=user_id, warehouse_id=wh_id)
            )
        return {"success": True, "message": "仓库授权已更新", "warehouse_ids": request.warehouse_ids}




# ============ Tenant Management APIs ============

@app.get("/api/tenants", response_model=List[TenantItem])
async def list_tenants(current_user: CurrentUser = Depends(require_permission(Resource.TENANTS, Action.ADMIN))):
    """获取租户列表（multi_tenant+admin）"""
    if get_deploy_mode() == "single_tenant":
        raise HTTPException(status_code=403, detail="单租户模式下不可用")

    # Phase 2f: SA Core read.
    stmt = select(
        _t_tenants.c.id, _t_tenants.c.slug, _t_tenants.c.name,
        _t_tenants.c.is_active, _t_tenants.c.created_at,
    )
    if current_user.tenant_id is not None:
        stmt = stmt.where(_t_tenants.c.id == current_user.tenant_id)
    stmt = stmt.order_by(_t_tenants.c.id.asc())
    with get_engine().connect() as sa_conn:
        rows = sa_conn.execute(stmt).fetchall()
    return [
        TenantItem(
            id=r.id, slug=r.slug, name=r.name,
            is_active=bool(r.is_active),
            created_at=(r.created_at.strftime('%Y-%m-%d %H:%M:%S')
                        if isinstance(r.created_at, datetime) else r.created_at),
        )
        for r in rows
    ]


@app.post("/api/tenants", response_model=TenantItem)
async def create_tenant(
    request: CreateTenantRequest,
    current_user: CurrentUser = Depends(require_permission(Resource.TENANTS, Action.ADMIN))
):
    """创建租户（仅全局 admin）"""
    if get_deploy_mode() == "single_tenant":
        raise HTTPException(status_code=403, detail="单租户模式下不可用")

    import re
    slug_pat = r"^[a-z0-9][a-z0-9\-]*$"
    if not re.match(slug_pat, request.slug):
        raise HTTPException(status_code=400, detail="租户标识只能包含小写字母、数字和连字符，且不能以连字符开头")

    if current_user.tenant_id is not None:
        raise HTTPException(status_code=403, detail="仅全局 admin 可创建租户")

    with get_engine().begin() as sa_conn:
        existing = sa_conn.execute(
            select(_t_tenants.c.id).where(_t_tenants.c.slug == request.slug)
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="租户标识已存在")

        created_at_dt = datetime.now()
        created_at = created_at_dt.strftime("%Y-%m-%d %H:%M:%S")
        result = sa_conn.execute(
            insert(_t_tenants).values(
                slug=request.slug, name=request.name, created_at=created_at_dt,
            )
        )
        tenant_id = result.inserted_primary_key[0]

        return TenantItem(
            id=tenant_id, slug=request.slug, name=request.name,
            is_active=True, created_at=created_at
        )


@app.put("/api/tenants/{tenant_id}", response_model=TenantItem)
async def update_tenant(
    tenant_id: int,
    request: UpdateTenantRequest,
    current_user: CurrentUser = Depends(require_permission(Resource.TENANTS, Action.ADMIN))
):
    """更新租户（仅全局 admin）"""
    if get_deploy_mode() == "single_tenant":
        raise HTTPException(status_code=403, detail="单租户模式下不可用")

    if current_user.tenant_id is not None:
        raise HTTPException(status_code=403, detail="仅全局 admin 可修改租户")

    with get_engine().begin() as sa_conn:
        tenant = sa_conn.execute(
            select(
                _t_tenants.c.id, _t_tenants.c.slug, _t_tenants.c.name,
                _t_tenants.c.is_active, _t_tenants.c.created_at,
            ).where(_t_tenants.c.id == tenant_id)
        ).first()
        if not tenant:
            raise HTTPException(status_code=404, detail="租户不存在")

        if tenant_id == 1:
            raise HTTPException(status_code=400, detail="不能修改默认租户")

        values = {}
        if request.name is not None:
            values["name"] = request.name
        if request.is_active is not None:
            values["is_active"] = 1 if request.is_active else 0

        if values:
            sa_conn.execute(
                update(_t_tenants).where(_t_tenants.c.id == tenant_id).values(**values)
            )

        r = sa_conn.execute(
            select(
                _t_tenants.c.id, _t_tenants.c.slug, _t_tenants.c.name,
                _t_tenants.c.is_active, _t_tenants.c.created_at,
            ).where(_t_tenants.c.id == tenant_id)
        ).first()
        return TenantItem(
            id=r.id, slug=r.slug, name=r.name,
            is_active=bool(r.is_active),
            created_at=(r.created_at.strftime('%Y-%m-%d %H:%M:%S')
                        if isinstance(r.created_at, datetime) else r.created_at),
        )


@app.delete("/api/tenants/{tenant_id}")
async def delete_tenant(
    tenant_id: int,
    current_user: CurrentUser = Depends(require_permission(Resource.TENANTS, Action.ADMIN))
):
    """停用租户（软删除，仅全局 admin）"""
    if get_deploy_mode() == "single_tenant":
        raise HTTPException(status_code=403, detail="单租户模式下不可用")

    if current_user.tenant_id is not None:
        raise HTTPException(status_code=403, detail="仅全局 admin 可停用租户")

    if tenant_id == 1:
        raise HTTPException(status_code=400, detail="不能停用默认租户")

    with get_engine().begin() as sa_conn:
        existing = sa_conn.execute(
            select(_t_tenants.c.id).where(_t_tenants.c.id == tenant_id)
        ).first()
        if not existing:
            raise HTTPException(status_code=404, detail="租户不存在")

        sa_conn.execute(
            update(_t_tenants).where(_t_tenants.c.id == tenant_id).values(is_active=0)
        )
        revoked_at_dt = datetime.now()
        users_subq = select(_t_users.c.id).where(_t_users.c.tenant_id == tenant_id)
        sa_conn.execute(
            update(_t_sessions).where(
                and_(
                    _t_sessions.c.user_id.in_(users_subq),
                    _t_sessions.c.revoked_at.is_(None),
                )
            ).values(revoked_at=revoked_at_dt)
        )
        return {"success": True, "message": "租户已停用"}


# ============ Auth APIs ============

@app.get("/api/auth/status", response_model=AuthStatusResponse)
async def get_auth_status(current_user: CurrentUser = Depends(get_current_user)):
    """获取认证状态"""
    initialized = has_admin_user()
    system_mode = "multi_tenant" if get_deploy_mode() == "multi_tenant" else "single_tenant"

    if current_user.is_guest:
        return AuthStatusResponse(
            initialized=initialized,
            logged_in=False,
            user=None,
            system_mode=system_mode,
        )

    return AuthStatusResponse(
        initialized=initialized,
        logged_in=True,
        user=UserInfo(
            id=current_user.id,
            username=current_user.username,
            display_name=current_user.display_name,
            role=current_user.role,
            tenant_id=current_user.tenant_id
        ),
        system_mode=system_mode,
    )


# ============ Registration APIs (multi_tenant self-service) ============

@app.post("/api/auth/register/verify-device", response_model=VerifyDeviceResponse)
@limiter.limit("10/hour")  # 防 device_id 暴力枚举（同 IP）；并发场景由工厂 API 自身限速兜底
async def register_verify_device(request: Request, body: VerifyDeviceRequest):
    """Step 1: 验证设备 ID 是否在工厂库中，以及是否已被注册。"""
    device_id = body.device_id.strip()
    if not device_id:
        return VerifyDeviceResponse(authorized=False, registered=False)

    # 先查本地：该设备是否已被某租户绑定（已注册直接返回，不走工厂 API）
    with get_engine().connect() as sa_conn:
        existing = sa_conn.execute(
            select(_t_tenants.c.name).where(_t_tenants.c.device_id == device_id)
        ).first()
    if existing:
        return VerifyDeviceResponse(authorized=True, registered=True, tenant_name=existing.name)

    # 未注册 → 调上游工厂 API 验证是否为合法设备
    if not FACTORY_API_KEY:
        return JSONResponse(
            status_code=503,
            content={"success": False, "message": "Device verification not configured"},
        )

    upstream_url = f"{FACTORY_API_BASE_URL}/factory/devices"
    authorized = False
    upstream_ok = False
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.get(
                upstream_url,
                params={"query": device_id, "page": 1, "pageSize": 1},
                headers={"X-Factory-Key": FACTORY_API_KEY},
            )
            if resp.status_code == 200:
                upstream_ok = True
                data = resp.json()
                if data.get("success") and data.get("data", {}).get("list"):
                    authorized = True
            elif resp.status_code == 401:
                logger.error(f"Factory API returned 401 — check FACTORY_API_KEY")
                raise HTTPException(status_code=502, detail="Device verification service misconfigured")
            else:
                logger.error(f"Factory API returned {resp.status_code}: {resp.text[:200]}")
                raise HTTPException(status_code=502, detail="Device verification service error")
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="Device verification service timeout")
    except httpx.RequestError:
        raise HTTPException(status_code=502, detail="Device verification service unavailable")

    if not authorized:
        return VerifyDeviceResponse(authorized=False, registered=False)

    return VerifyDeviceResponse(authorized=True, registered=False)


@app.post("/api/auth/register")
@limiter.limit("5/hour")  # 防批量注册（同 IP）：自助注册是低频操作，5/h 远超正常用户需求
async def register_tenant(request: Request, body: RegisterRequest, response: Response):
    """Step 2: 创建租户 + admin 账号 + 默认仓库，绑定设备 ID。"""
    if get_deploy_mode() != "multi_tenant":
        raise HTTPException(status_code=403, detail="自助注册仅在多租户模式下可用")

    if not has_admin_user():
        raise HTTPException(status_code=400, detail="系统尚未初始化，请先创建全局管理员")

    device_id = body.device_id.strip()
    username = body.username.strip()
    password = body.password.strip()

    if not device_id or not username or not password:
        raise HTTPException(status_code=400, detail="设备ID、用户名、密码均不能为空")
    err = validate_username(username)
    if err:
        raise HTTPException(status_code=400, detail=err)
    err = validate_password_strength(password)
    if err:
        raise HTTPException(status_code=400, detail=err)

    # ── 事务外 read-only 预检：device_id 已绑定就快速 409（避免白调一次工厂 API）──
    with get_engine().connect() as sa_conn:
        existing = sa_conn.execute(
            select(_t_tenants.c.id).where(_t_tenants.c.device_id == device_id)
        ).first()
    if existing:
        raise HTTPException(status_code=409, detail="该设备已注册，如需重置密码请使用找回密码功能")

    # ── 事务外：调工厂 API 验证设备（数十秒级 HTTP 调用，绝不能放在 DB 事务里持锁）──
    if not FACTORY_API_KEY:
        raise HTTPException(status_code=503, detail="Device verification not configured")
    upstream_url = f"{FACTORY_API_BASE_URL}/factory/devices"
    authorized = False
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.get(
                upstream_url,
                params={"query": device_id, "page": 1, "pageSize": 1},
                headers={"X-Factory-Key": FACTORY_API_KEY},
            )
            if resp.status_code == 200:
                data = resp.json()
                if data.get("success") and data.get("data", {}).get("list"):
                    authorized = True
            elif resp.status_code == 401:
                logger.error(f"Factory API returned 401 — check FACTORY_API_KEY")
                raise HTTPException(status_code=502, detail="Device verification service misconfigured")
            else:
                logger.error(f"Factory API returned {resp.status_code}: {resp.text[:200]}")
                raise HTTPException(status_code=502, detail="Device verification service error")
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="Device verification service timeout")
    except httpx.RequestError:
        raise HTTPException(status_code=502, detail="Device verification service unavailable")

    if not authorized:
        raise HTTPException(status_code=400, detail="设备未授权，请确认设备 ID 正确")

    # ── 事务外：CPU 密集 hash 与 token 生成 ──
    tenant_slug_base = f"tenant-{secrets.token_hex(4)}"
    password_hash = hash_password(password)
    token = generate_session_token()
    expires_at = datetime.now() + timedelta(hours=24)

    # ── 事务内：仅 DB 写入，依赖 tenants.device_id UNIQUE 兜底并发竞态 ──
    with get_engine().begin() as sa_conn:
        try:
            result = sa_conn.execute(
                insert(_t_tenants).values(
                    slug=tenant_slug_base,
                    name=f"租户-{device_id}",
                    device_id=device_id,
                )
            )
        except IntegrityError:
            # 两请求并发：另一个已经写成功；当前请求转为友好错误。
            raise HTTPException(status_code=409, detail="该设备已注册，如需重置密码请使用找回密码功能")
        tenant_id = result.inserted_primary_key[0]

        # 创建默认仓库
        wh_result = sa_conn.execute(
            insert(_t_warehouses).values(
                slug=f"wh-{tenant_slug_base}",
                name="默认仓库",
                is_default=1,
                tenant_id=tenant_id,
                created_at=datetime.now(),
            )
        )
        wh_id = wh_result.inserted_primary_key[0]

        # 创建 admin 用户
        result = sa_conn.execute(
            insert(_t_users).values(
                username=username,
                password_hash=password_hash,
                display_name=body.display_name or username,
                role=RoleName.ADMIN.value,
                tenant_id=tenant_id,
                created_by=None,
            )
        )
        user_id = result.inserted_primary_key[0]

        # 授权仓库
        sa_conn.execute(
            insert(_t_user_warehouses).values(user_id=user_id, warehouse_id=wh_id)
        )

        # 创建 session 自动登录
        sa_conn.execute(
            insert(_t_sessions).values(user_id=user_id, token=token, expires_at=expires_at)
        )

    response.set_cookie(
        key="session_token",
        value=token,
        httponly=True,
        max_age=86400,
        samesite="lax",
        secure=False,
    )

    return {
        "success": True,
        "message": "注册成功",
        "user": {
            "id": user_id,
            "username": username,
            "display_name": body.display_name or username,
            "role": RoleName.ADMIN.value,
            "tenant_id": tenant_id,
        },
    }


@app.post("/api/auth/reset-password")
@limiter.limit("5/hour")  # device_id 作为 possession factor 的恢复操作：限速防暴力枚举/批量接管
async def reset_password(request: Request, body: ResetPasswordRequest):
    """凭设备 ID + 管理员用户名重置密码。

    安全模型：device_id 视为物理持有因子（"谁手里有设备 → 谁能恢复 admin"）。
    物理保管是客户责任；本端做的补强是：
      - 限速（@limiter.limit 5/hour）防暴力枚举 device_id 或批量接管攻击
      - 每次尝试（成功/失败）写 audit log，含 IP 与 UA，便于事后追溯

    未来若客户上线邮箱/SMS 通道，应当在成功后异步推送通知给原 admin。
    """
    if get_deploy_mode() != "multi_tenant":
        raise HTTPException(status_code=403, detail="仅在多租户模式下可用")

    device_id = body.device_id.strip()
    username = body.username.strip()
    new_password = body.new_password.strip()

    # audit_log 公共字段：client IP / UA 取自 request（@limiter 已注入）
    client_ip = request.client.host if request.client else None
    user_agent = request.headers.get("user-agent", "")[:200]
    audit_base = {
        "device_id": device_id,
        "username_attempted": username,
        "client_ip": client_ip,
        "user_agent": user_agent,
    }

    if not device_id or not username or not new_password:
        audit_log("RESET_PASSWORD_FAIL", None, None, {**audit_base, "reason": "empty_field"})
        raise HTTPException(status_code=400, detail="设备ID、用户名、新密码均不能为空")
    err = validate_password_strength(new_password)
    if err:
        audit_log("RESET_PASSWORD_FAIL", None, None, {**audit_base, "reason": "weak_password"})
        raise HTTPException(status_code=400, detail=err)

    with get_engine().begin() as sa_conn:
        # 查租户
        tenant_row = sa_conn.execute(
            select(_t_tenants.c.id, _t_tenants.c.name)
            .where(_t_tenants.c.device_id == device_id)
        ).first()
        if not tenant_row:
            audit_log("RESET_PASSWORD_FAIL", None, None, {**audit_base, "reason": "unknown_device_id"})
            raise HTTPException(status_code=404, detail="未找到该设备对应的租户，请确认设备 ID 正确")

        # 查该租户下指定用户名的 admin
        admin_row = sa_conn.execute(
            select(_t_users.c.id, _t_users.c.username)
            .where(
                and_(
                    _t_users.c.tenant_id == tenant_row.id,
                    _t_users.c.username == username,
                    _t_users.c.role == RoleName.ADMIN.value,
                    _t_users.c.is_disabled == 0,
                )
            )
        ).first()
        if not admin_row:
            audit_log("RESET_PASSWORD_FAIL", None, None, {
                **audit_base, "reason": "unknown_admin", "tenant_id": tenant_row.id, "tenant_name": tenant_row.name,
            })
            raise HTTPException(status_code=404, detail="管理员用户名不正确，请确认后重试")

        # 重置密码
        new_hash = hash_password(new_password)
        sa_conn.execute(
            update(_t_users).where(_t_users.c.id == admin_row.id).values(password_hash=new_hash)
        )

    audit_log("RESET_PASSWORD_SUCCESS", admin_row.id, admin_row.username, {
        **audit_base, "tenant_id": tenant_row.id, "tenant_name": tenant_row.name,
    })
    return {"success": True, "message": f"租户「{tenant_row.name}」管理员 {admin_row.username} 密码已重置"}



@app.post("/api/auth/setup", response_model=LoginResponse)
async def setup_admin(request: SetupRequest, response: Response):
    """首次设置管理员账号"""
    if has_admin_user():
        raise HTTPException(status_code=400, detail="系统已初始化，无法重复设置")

    if len(request.password) < 4:
        raise HTTPException(status_code=400, detail="密码长度至少4位")

    # bcrypt + token gen 在事务外执行（CPU 密集 / 纯 Python）
    password_hash = hash_password(request.password)
    token = generate_session_token()
    expires_at = datetime.now() + timedelta(hours=24)
    # 全局 admin（tenant_id = NULL）仅在 multi_tenant 下；single_tenant 下 tenant_id = 1
    setup_tenant_id = None if get_deploy_mode() == 'multi_tenant' else 1

    with get_engine().begin() as sa_conn:
        result = sa_conn.execute(
            insert(_t_users).values(
                username=request.username,
                password_hash=password_hash,
                role=RoleName.ADMIN.value,
                display_name=request.display_name,
                tenant_id=setup_tenant_id,
                created_at=datetime.now(),
            )
        )
        user_id = result.inserted_primary_key[0]

        sa_conn.execute(
            insert(_t_sessions).values(
                user_id=user_id,
                token=token,
                expires_at=expires_at,
                created_at=datetime.now(),
            )
        )

    # 设置Cookie
    response.set_cookie(
        key="session_token",
        value=token,
        max_age=86400,  # 24小时
        httponly=True,
        samesite="lax"
    )

    return LoginResponse(
        success=True,
        message="管理员账号创建成功",
        user=UserInfo(
            id=user_id,
            username=request.username,
            display_name=request.display_name,
            role=RoleName.ADMIN.value,
            tenant_id=setup_tenant_id
        ),
        is_first_login=True,
    )


@app.post("/api/auth/login", response_model=LoginResponse)
@limiter.limit("5/minute")  # 登录接口速率限制：每分钟5次
async def login(request: Request, login_data: LoginRequest, response: Response):
    """用户登录"""
    user_stmt = select(
        _t_users.c.id, _t_users.c.username, _t_users.c.password_hash,
        _t_users.c.display_name, _t_users.c.role, _t_users.c.is_disabled,
        _t_users.c.tenant_id, _t_users.c.last_login_at,
        _t_tenants.c.is_active.label('tenant_is_active'),
    ).select_from(
        _t_users.outerjoin(_t_tenants, _t_users.c.tenant_id == _t_tenants.c.id)
    ).where(_t_users.c.username == login_data.username)

    with get_engine().connect() as sa_conn:
        user_rows = sa_conn.execute(user_stmt).all()

    if not user_rows:
        return LoginResponse(success=False, message="用户名或密码错误")

    password_matched = [
        u for u in user_rows
        if verify_password(login_data.password, u.password_hash)
    ]
    if len(password_matched) == 0:
        return LoginResponse(success=False, message="用户名或密码错误")

    enabled_matched = [u for u in password_matched if not u.is_disabled]
    if len(enabled_matched) == 0:
        return LoginResponse(success=False, message="用户已被禁用")

    matched = [
        u for u in enabled_matched
        if u.tenant_id is None or bool(u.tenant_is_active)
    ]
    if len(matched) == 0:
        return LoginResponse(success=False, message="租户已停用，请联系管理员")
    if len(matched) > 1:
        return LoginResponse(success=False, message="同名账号存在于多个租户，请联系管理员")
    user = matched[0]

    # 透明密码升级：如果使用旧的SHA256哈希，自动升级到bcrypt
    new_hash = None
    if needs_password_rehash(user.password_hash):
        new_hash = hash_password(login_data.password)

    # 创建新会话（允许同账号多端并发登录，不清理旧会话）
    token = generate_session_token()
    expires_at = datetime.now() + timedelta(hours=24)

    with get_engine().begin() as sa_conn:
        if new_hash is not None:
            sa_conn.execute(
                update(_t_users).where(_t_users.c.id == user.id).values(password_hash=new_hash)
            )
        sa_conn.execute(
            insert(_t_sessions).values(
                user_id=user.id,
                token=token,
                expires_at=expires_at,
                created_at=datetime.now(),
            )
        )

    if new_hash is not None:
        logger.info(f"Password upgraded to bcrypt for user: {user.username}")

    # 判断是否首次登录
    is_first_login = user.last_login_at is None

    # 更新 last_login_at
    with get_engine().begin() as sa_conn:
        sa_conn.execute(
            update(_t_users).where(_t_users.c.id == user.id).values(last_login_at=datetime.now())
        )

    # 审计日志
    audit_log("LOGIN", user.id, user.username, {"role": user.role})

    # 设置Cookie
    response.set_cookie(
        key="session_token",
        value=token,
        max_age=86400,
        httponly=True,
        samesite="lax"
    )

    return LoginResponse(
        success=True,
        message="登录成功",
        user=UserInfo(
            id=user.id,
            username=user.username,
            display_name=user.display_name,
            role=user.role,
            tenant_id=user.tenant_id
        ),
        is_first_login=is_first_login,
    )


@app.post("/api/auth/logout")
async def logout(response: Response, current_user: CurrentUser = Depends(get_current_user)):
    """用户登出"""
    if current_user.source == 'session' and current_user.id:
        with get_engine().begin() as sa_conn:
            sa_conn.execute(
                delete(_t_sessions).where(_t_sessions.c.user_id == current_user.id)
            )

    response.delete_cookie("session_token")
    return {"success": True, "message": "已登出"}


@app.get("/api/auth/me", response_model=UserInfo)
async def get_current_user_info(current_user: CurrentUser = Depends(require_permission(Resource.AUTH, Action.READ))):
    """获取当前用户信息"""
    if current_user.is_guest:
        raise HTTPException(status_code=401, detail="未登录")

    return UserInfo(
        id=current_user.id,
        username=current_user.username,
        display_name=current_user.display_name,
        role=current_user.role,
        tenant_id=current_user.tenant_id
    )


@app.get("/api/auth/warehouses")
async def get_my_warehouses(current_user: CurrentUser = Depends(require_permission(Resource.WAREHOUSES, Action.READ))):
    """获取当前用户可访问的仓库列表（含 tenant 信息，便于全局 admin 分组展示）"""
    with get_db() as conn:
        warehouses = current_user.get_authorized_warehouses(conn)
        cursor = conn.cursor()
        result = []
        for wh_id in warehouses:
            cursor.execute('''
                SELECT w.id, w.slug, w.name, w.is_default, w.tenant_id, t.name AS tenant_name
                FROM warehouses w
                LEFT JOIN tenants t ON w.tenant_id = t.id
                WHERE w.id = ?
            ''', (wh_id,))
            wh = cursor.fetchone()
            if wh:
                result.append({
                    "id": wh['id'],
                    "slug": wh['slug'],
                    "name": wh['name'],
                    "is_default": bool(wh['is_default']),
                    "tenant_id": wh['tenant_id'],
                    "tenant_name": wh['tenant_name'],
                })
        return {"warehouses": result}


# ============ User Management APIs ============

@app.get("/api/users", response_model=List[UserListItem])
async def list_users(current_user: CurrentUser = Depends(require_permission(Resource.USERS, Action.ADMIN))):
    """获取用户列表（仅管理员）。Phase 2c: SA Core reads."""
    deploy_mode_local = get_deploy_mode()
    user_stmt = select(
        _t_users.c.id, _t_users.c.username, _t_users.c.display_name,
        _t_users.c.role, _t_users.c.is_disabled, _t_users.c.created_at,
        _t_users.c.tenant_id,
    )
    if deploy_mode_local == 'multi_tenant':
        scope_preds = build_scope_predicates(_t_users, current_user.tenant_id, None)
        if scope_preds:
            user_stmt = user_stmt.where(*scope_preds)
    user_stmt = user_stmt.order_by(_t_users.c.created_at.desc())

    with get_engine().connect() as sa_conn:
        users_rows = sa_conn.execute(user_stmt).fetchall()
        result = []
        for row in users_rows:
            wh_stmt = select(
                _t_warehouses.c.id, _t_warehouses.c.name,
            ).select_from(
                _t_user_warehouses.join(
                    _t_warehouses, _t_user_warehouses.c.warehouse_id == _t_warehouses.c.id
                )
            ).where(_t_user_warehouses.c.user_id == row.id)
            wh_rows = sa_conn.execute(wh_stmt).fetchall()
            ca = row.created_at
            if isinstance(ca, datetime):
                ca = ca.strftime('%Y-%m-%d %H:%M:%S')
            result.append(UserListItem(
                id=row.id,
                username=row.username,
                display_name=row.display_name,
                role=row.role,
                is_disabled=bool(row.is_disabled),
                created_at=ca,
                tenant_id=row.tenant_id,
                warehouse_ids=[r.id for r in wh_rows],
                warehouse_names=[r.name for r in wh_rows],
            ))
    return result


# ---- users CREATE/UPDATE/DELETE migrated to ResourceRouter (R2 phase 2) ----
# LIST stays as ``list_users`` above because it joins user_warehouses to fill
# warehouse_ids / warehouse_names on each row.

_USER_OUT_COLUMNS = [
    _t_users.c.id, _t_users.c.username, _t_users.c.display_name,
    _t_users.c.role, _t_users.c.is_disabled, _t_users.c.created_at,
    _t_users.c.tenant_id,
]


def _user_to_out(row) -> UserListItem:
    ca = row.created_at
    if isinstance(ca, datetime):
        ca = ca.strftime('%Y-%m-%d %H:%M:%S')
    return UserListItem(
        id=row.id,
        username=row.username,
        display_name=row.display_name,
        role=row.role,
        is_disabled=bool(row.is_disabled),
        created_at=ca,
        tenant_id=row.tenant_id,
        # warehouse_ids / warehouse_names left as default None — matches the
        # original create_user / update_user response (LIST handler is the
        # only one that fills them).
    )


def _user_before_create(sa_conn, current_user, request: CreateUserRequest):
    if request.role not in _VALID_ROLE_VALUES:
        raise HTTPException(status_code=400, detail="无效的角色")
    if len(request.password) < 4:
        raise HTTPException(status_code=400, detail="密码长度至少4位")
    # 用户名唯一性按租户隔离检查
    target_tenant_id = request.tenant_id if request.tenant_id is not None else current_user.tenant_id
    existing = sa_conn.execute(
        select(_t_users.c.id).where(
            and_(_t_users.c.username == request.username, _t_users.c.tenant_id == target_tenant_id)
        )
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="用户名已存在")


def _user_values_for_create(sa_conn, current_user, request: CreateUserRequest) -> dict:
    # bcrypt is CPU-bound; we do it inside this hook (called inside begin()).
    # Original code did it outside the transaction — moving it inside is
    # acceptable here because the route still raises before any I/O.
    password_hash = hash_password(request.password)

    if current_user.tenant_id is None:
        if get_deploy_mode() == 'multi_tenant':
            new_tenant_id = request.tenant_id
            if new_tenant_id is None and request.role != 'admin':
                raise HTTPException(status_code=400, detail="全局用户必须是管理员角色")
        else:
            new_tenant_id = 1
    elif request.tenant_id is not None and request.tenant_id != current_user.tenant_id:
        raise HTTPException(status_code=403, detail="无权在其他租户下创建用户")
    else:
        new_tenant_id = current_user.tenant_id
    if new_tenant_id is not None:
        tenant_row = sa_conn.execute(
            select(_t_tenants.c.id).where(
                and_(_t_tenants.c.id == new_tenant_id, _t_tenants.c.is_active == 1)
            )
        ).first()
        if not tenant_row:
            raise HTTPException(status_code=400, detail="租户不存在或已停用")

    return {
        "username": request.username,
        "password_hash": password_hash,
        "role": request.role,
        "display_name": request.display_name,
        "created_by": current_user.id,
        "tenant_id": new_tenant_id,
        "created_at": datetime.now(),
    }


def _user_values_for_update(sa_conn, current_user, request: UpdateUserRequest, row) -> dict:
    user_id = row.id
    new_password_hash = None
    if request.password is not None:
        if len(request.password) < 4:
            raise HTTPException(status_code=400, detail="密码长度至少4位")
        new_password_hash = hash_password(request.password)

    values: dict = {}

    if request.username is not None:
        dup = sa_conn.execute(
            select(_t_users.c.id).where(
                and_(
                    _t_users.c.username == request.username,
                    _t_users.c.id != user_id,
                    _t_users.c.tenant_id == row.tenant_id,
                )
            )
        ).first()
        if dup:
            raise HTTPException(status_code=400, detail="用户名已存在")
        if len(request.username) < 2:
            raise HTTPException(status_code=400, detail="用户名长度至少2位")
        values['username'] = request.username

    if request.display_name is not None:
        values['display_name'] = request.display_name

    if request.role is not None:
        if request.role not in _VALID_ROLE_VALUES:
            raise HTTPException(status_code=400, detail="无效的角色")
        values['role'] = request.role

    if new_password_hash is not None:
        values['password_hash'] = new_password_hash

    if request.is_disabled is not None:
        values['is_disabled'] = 1 if request.is_disabled else 0

    # 密码变更或禁用用户时吊销所有会话 — done as a side effect, not a
    # column update, so it has to live here (we still have the conn).
    if request.password is not None or (request.is_disabled is not None and request.is_disabled):
        sa_conn.execute(
            update(_t_sessions)
            .where(and_(_t_sessions.c.user_id == user_id, _t_sessions.c.revoked_at.is_(None)))
            .values(revoked_at=datetime.now())
        )

    return values


def _user_before_delete(sa_conn, current_user, row):
    user_id = row.id
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="不能禁用自己")
    sa_conn.execute(
        update(_t_users).where(_t_users.c.id == user_id).values(is_disabled=1)
    )
    sa_conn.execute(
        update(_t_sessions)
        .where(and_(_t_sessions.c.user_id == user_id, _t_sessions.c.revoked_at.is_(None)))
        .values(revoked_at=datetime.now())
    )


def _user_after_commit(operation, sa_conn, current_user, row_id):
    # R5: only operator partition is affected by user writes
    get_fuzzy_matcher().invalidate_cache(entity_type="operator")


from resource_router import ResourceRouter as _ResourceRouterUser  # noqa: E402

_user_router = _ResourceRouterUser(
    app=app,
    prefix="/api/users",
    table=_t_users,
    response_model=UserListItem,
    create_model=CreateUserRequest,
    update_model=UpdateUserRequest,
    permission_read=require_permission(Resource.USERS, Action.ADMIN),
    permission_write=require_permission(Resource.USERS, Action.ADMIN),
    not_found_detail="用户不存在",
    forbidden_detail="无权操作其他租户的用户",
    to_out=_user_to_out,
    values_for_create=_user_values_for_create,
    values_for_update=_user_values_for_update,
    before_create=_user_before_create,
    before_delete=_user_before_delete,
    after_commit=_user_after_commit,
    list_handler=None,
    get_columns=_USER_OUT_COLUMNS,
    update_select_columns=_USER_OUT_COLUMNS,
    enable_get=False,
    delete_response={"success": True, "message": "用户已禁用"},
)
_user_router.register()


# ============ API Key Management APIs ============

@app.get("/api/api-keys", response_model=List[ApiKeyListItem])
async def list_api_keys(current_user: CurrentUser = Depends(require_permission(Resource.API_KEYS, Action.ADMIN))):
    """获取API密钥列表（仅管理员）— Phase 2f: SA Core read."""
    preds = [_t_api_keys.c.is_system == 0]
    preds.extend(build_scope_predicates(_t_api_keys, current_user.tenant_id, None))
    stmt = (
        select(
            _t_api_keys.c.id, _t_api_keys.c.name, _t_api_keys.c.role,
            _t_api_keys.c.is_disabled, _t_api_keys.c.created_at,
            _t_api_keys.c.last_used_at, _t_api_keys.c.warehouse_id,
            _t_warehouses.c.name.label('warehouse_name'),
        )
        .select_from(
            _t_api_keys.outerjoin(_t_warehouses, _t_api_keys.c.warehouse_id == _t_warehouses.c.id)
        )
        .where(and_(*preds))
        .order_by(_t_api_keys.c.created_at.desc())
    )
    with get_engine().connect() as sa_conn:
        rows = sa_conn.execute(stmt).fetchall()
    return [
        ApiKeyListItem(
            id=row.id,
            name=row.name,
            role=row.role,
            is_disabled=bool(row.is_disabled),
            created_at=(row.created_at.strftime('%Y-%m-%d %H:%M:%S')
                        if isinstance(row.created_at, datetime) else row.created_at),
            last_used_at=(row.last_used_at.strftime('%Y-%m-%d %H:%M:%S')
                          if isinstance(row.last_used_at, datetime) else row.last_used_at),
            warehouse_id=row.warehouse_id,
            warehouse_name=row.warehouse_name,
        )
        for row in rows
    ]


# ---- api-keys CREATE / DELETE migrated to ResourceRouter (R2 phase 2) ----
# LIST stays as ``list_api_keys`` above (it joins warehouses for
# warehouse_name and filters out is_system rows). There is no GET-by-id
# or PUT route on /api/api-keys/{id}; toggling status lives at
# /api/api-keys/{id}/status (handler ``toggle_api_key_status`` below).
#
# CREATE returns the plaintext key one-time via ``to_out_create`` —
# subsequent reads only ever see the prefix / hash.

# Per-request stash for "the plaintext key we just generated". Because
# values_for_create generates the key and to_out_create needs it back,
# and the factory's hook contract doesn't pass the request between them
# directly, we thread it through ``request._plaintext_api_key``. Pydantic
# v2 BaseModel allows arbitrary attribute set via ``object.__setattr__``.

def _apikey_before_create(sa_conn, current_user, request: CreateApiKeyRequest):
    if request.role not in _VALID_ROLE_VALUES:
        raise HTTPException(status_code=400, detail="无效的角色")


def _apikey_values_for_create(sa_conn, current_user, request: CreateApiKeyRequest) -> dict:
    api_key_plain = generate_api_key()
    key_hash = hash_api_key(api_key_plain)
    # Stash plaintext on the request so to_out_create can retrieve it.
    object.__setattr__(request, "_plaintext_api_key", api_key_plain)

    wh_id = request.warehouse_id
    if wh_id is not None:
        wh_id = resolve_warehouse_id(current_user, wh_id)

    if current_user.tenant_id is None and wh_id is None:
        raise HTTPException(
            status_code=400,
            detail="全局管理员创建 API Key 必须指定 warehouse_id（无法推导目标租户）"
        )
    if current_user.tenant_id is not None:
        key_tenant_id = current_user.tenant_id
    else:
        key_tenant_id = resolve_tenant_id_for_write(current_user, wh_id)

    return {
        "key_hash": key_hash,
        "name": request.name,
        "role": request.role,
        "user_id": current_user.id,
        "tenant_id": key_tenant_id,
        "warehouse_id": wh_id,
        "created_at": datetime.now(),
    }


def _apikey_to_out_create(row, *, request, sa_conn, current_user) -> ApiKeyResponse:
    plaintext = getattr(request, "_plaintext_api_key", None)
    ca = row.created_at
    if isinstance(ca, datetime):
        ca = ca.strftime('%Y-%m-%d %H:%M:%S')
    return ApiKeyResponse(
        id=row.id,
        name=row.name,
        role=row.role,
        key=plaintext,
        created_at=ca,
        # last_used_at is None on a freshly created key — included so the
        # factory's response_model serialises the same shape as the
        # existing handler (Pydantic emits the field with default None).
        last_used_at=None,
    )


def _apikey_to_out(row) -> ApiKeyResponse:
    # GET-by-id is not exposed; this is only here to satisfy the factory
    # signature requirement. PUT is also disabled. If either is ever
    # re-enabled, this will return the read-only shape (no plaintext).
    ca = row.created_at
    if isinstance(ca, datetime):
        ca = ca.strftime('%Y-%m-%d %H:%M:%S')
    return ApiKeyResponse(
        id=row.id, name=row.name, role=row.role, key=None,
        created_at=ca, last_used_at=None,
    )


def _apikey_values_for_update(sa_conn, current_user, request, row) -> dict:
    # Unreachable — PUT is disabled. Kept to satisfy required-hook contract.
    return {}


from resource_router import ResourceRouter as _ResourceRouterAK  # noqa: E402

_apikey_router = _ResourceRouterAK(
    app=app,
    prefix="/api/api-keys",
    table=_t_api_keys,
    response_model=ApiKeyResponse,
    create_model=CreateApiKeyRequest,
    update_model=CreateApiKeyRequest,  # placeholder, PUT disabled
    permission_read=require_permission(Resource.API_KEYS, Action.ADMIN),
    permission_write=require_permission(Resource.API_KEYS, Action.ADMIN),
    not_found_detail="API密钥不存在",
    forbidden_detail="无权操作其他租户的API密钥",
    to_out=_apikey_to_out,
    to_out_create=_apikey_to_out_create,
    values_for_create=_apikey_values_for_create,
    values_for_update=_apikey_values_for_update,
    before_create=_apikey_before_create,
    list_handler=None,
    enable_get=False,
    enable_put=False,
    hard_delete=True,
    delete_response={"success": True, "message": "API密钥已删除"},
)
_apikey_router.register()


@app.put("/api/api-keys/{key_id}/status")
async def toggle_api_key_status(
    key_id: int,
    request: ApiKeyStatusRequest,
    current_user: CurrentUser = Depends(require_permission(Resource.API_KEYS, Action.ADMIN))
):
    """切换API密钥状态（仅管理员）"""
    with get_engine().begin() as sa_conn:
        row = load_or_404(
            sa_conn, _t_api_keys, key_id,
            columns=[_t_api_keys.c.id, _t_api_keys.c.tenant_id],
            not_found="API密钥不存在",
            tenant_id=current_user.tenant_id,
            forbidden="无权操作其他租户的API密钥",
        )

        sa_conn.execute(
            update(_t_api_keys).where(_t_api_keys.c.id == key_id).values(
                is_disabled=1 if request.disabled else 0
            )
        )

        status_text = "已禁用" if request.disabled else "已启用"
        return {"success": True, "message": f"API密钥{status_text}"}


# ============ Database Management APIs ============

# 仓库相关表（导出/导入/清空时操作）
# 顺序很重要：先无依赖的表，再有外键依赖的表
# warehouses -> materials, contacts -> batches -> inventory_records -> batch_consumptions
WAREHOUSE_TABLES = ['warehouses', 'materials', 'contacts', 'batches', 'inventory_records', 'batch_consumptions']


def _table_columns(cursor, table: str) -> List[str]:
    cursor.execute(f"PRAGMA table_info({table})")
    return [row['name'] if isinstance(row, sqlite3.Row) else row[1] for row in cursor.fetchall()]


def _unique_warehouse_slug(cursor, base_slug: str, tenant_id: Optional[int] = None) -> str:
    slug = base_slug
    idx = 1
    while True:
        if tenant_id is not None:
            cursor.execute('SELECT 1 FROM warehouses WHERE slug = ? AND tenant_id = ?', (slug, tenant_id))
        else:
            cursor.execute('SELECT 1 FROM warehouses WHERE slug = ?', (slug,))
        if not cursor.fetchone():
            return slug
        idx += 1
        slug = f"{base_slug}-{idx}"


def _ensure_default_warehouse_for_tenant(cursor, tenant_id: int) -> int:
    cursor.execute('SELECT id FROM warehouses WHERE tenant_id = ? LIMIT 1', (tenant_id,))
    row = cursor.fetchone()
    if row:
        return row['id']
    base_slug = 'default' if tenant_id == 1 else f'tenant-{tenant_id}-default'
    slug = _unique_warehouse_slug(cursor, base_slug, tenant_id=tenant_id)
    cursor.execute(
        'INSERT INTO warehouses (slug, name, is_default, tenant_id, created_at) VALUES (?, ?, 1, ?, ?)',
        (slug, '默认仓库', tenant_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    )
    return cursor.lastrowid


def _export_rows_for_scope(cursor, table: str, tenant_id: Optional[int]):
    if tenant_id is None:
        cursor.execute(f"SELECT * FROM {table}")
        return cursor.fetchall()
    if table == 'batch_consumptions':
        cursor.execute('''
            SELECT bc.*
            FROM batch_consumptions bc
            LEFT JOIN inventory_records r ON bc.record_id = r.id
            LEFT JOIN batches b ON bc.batch_id = b.id
            WHERE r.tenant_id = ? OR b.tenant_id = ?
        ''', (tenant_id, tenant_id))
        return cursor.fetchall()
    columns = _table_columns(cursor, table)
    if 'tenant_id' in columns:
        cursor.execute(f"SELECT * FROM {table} WHERE tenant_id = ?", (tenant_id,))
    else:
        cursor.execute(f"SELECT * FROM {table}")
    return cursor.fetchall()


def _count_rows_for_scope(cursor, table: str, tenant_id: Optional[int]) -> int:
    if tenant_id is None:
        cursor.execute(f"SELECT COUNT(*) as count FROM {table}")
        return cursor.fetchone()['count']
    if table == 'batch_consumptions':
        cursor.execute('''
            SELECT COUNT(*) as count
            FROM batch_consumptions bc
            LEFT JOIN inventory_records r ON bc.record_id = r.id
            LEFT JOIN batches b ON bc.batch_id = b.id
            WHERE r.tenant_id = ? OR b.tenant_id = ?
        ''', (tenant_id, tenant_id))
        return cursor.fetchone()['count']
    columns = _table_columns(cursor, table)
    if 'tenant_id' in columns:
        cursor.execute(f"SELECT COUNT(*) as count FROM {table} WHERE tenant_id = ?", (tenant_id,))
    else:
        cursor.execute(f"SELECT COUNT(*) as count FROM {table}")
    return cursor.fetchone()['count']


def _clear_database_scope(cursor, tenant_id: Optional[int]) -> dict:
    details = {table: _count_rows_for_scope(cursor, table, tenant_id) for table in WAREHOUSE_TABLES}

    if tenant_id is None:
        cursor.execute('DELETE FROM batch_consumptions')
        cursor.execute('DELETE FROM inventory_records')
        cursor.execute('DELETE FROM batches')
        cursor.execute('DELETE FROM materials')
        cursor.execute('DELETE FROM contacts')
        cursor.execute('DELETE FROM user_warehouses')
        cursor.execute('UPDATE api_keys SET warehouse_id = NULL')
        cursor.execute('UPDATE mcp_connections SET warehouse_id = NULL')
        cursor.execute('DELETE FROM warehouses')
        cursor.execute('SELECT id FROM tenants WHERE is_active = 1 ORDER BY id')
        for row in cursor.fetchall():
            _ensure_default_warehouse_for_tenant(cursor, row['id'])
        return details

    cursor.execute('SELECT id FROM warehouses WHERE tenant_id = ?', (tenant_id,))
    wh_ids = [row['id'] for row in cursor.fetchall()]

    cursor.execute('''
        DELETE FROM batch_consumptions
        WHERE record_id IN (SELECT id FROM inventory_records WHERE tenant_id = ?)
           OR batch_id IN (SELECT id FROM batches WHERE tenant_id = ?)
    ''', (tenant_id, tenant_id))
    cursor.execute('DELETE FROM inventory_records WHERE tenant_id = ?', (tenant_id,))
    cursor.execute('DELETE FROM batches WHERE tenant_id = ?', (tenant_id,))
    cursor.execute('DELETE FROM materials WHERE tenant_id = ?', (tenant_id,))
    cursor.execute('DELETE FROM contacts WHERE tenant_id = ?', (tenant_id,))
    if wh_ids:
        placeholders = ','.join('?' for _ in wh_ids)
        cursor.execute(f'DELETE FROM user_warehouses WHERE warehouse_id IN ({placeholders})', wh_ids)
        cursor.execute(f'UPDATE api_keys SET warehouse_id = NULL WHERE warehouse_id IN ({placeholders})', wh_ids)
        cursor.execute(f'UPDATE mcp_connections SET warehouse_id = NULL WHERE warehouse_id IN ({placeholders})', wh_ids)
    cursor.execute('DELETE FROM warehouses WHERE tenant_id = ?', (tenant_id,))
    _ensure_default_warehouse_for_tenant(cursor, tenant_id)
    return details


def _insert_row_with_overrides(cursor, table: str, row, target_columns: set, overrides: dict = None, skip: set = None) -> int:
    overrides = overrides or {}
    skip = skip or set()
    source_columns = set(row.keys())
    columns = [col for col in source_columns if col in target_columns and col not in skip]
    for col in overrides:
        if col in target_columns and col not in columns:
            columns.append(col)
    values = [overrides[col] if col in overrides else row[col] for col in columns]
    placeholders = ','.join('?' for _ in columns)
    cursor.execute(f"INSERT INTO {table} ({','.join(columns)}) VALUES ({placeholders})", values)
    return cursor.lastrowid


def _import_tenant_database(cursor, import_cursor, available_tables: set, tenant_id: int) -> dict:
    details = {}
    _clear_database_scope(cursor, tenant_id)

    target_columns = {}
    for table in WAREHOUSE_TABLES:
        target_columns[table] = set(_table_columns(cursor, table))

    warehouse_map = {}
    contact_map = {}
    material_map = {}
    batch_map = {}
    record_map = {}

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if 'warehouses' in available_tables:
        import_cursor.execute('SELECT * FROM warehouses')
        wh_rows = import_cursor.fetchall()
    else:
        wh_rows = []
    for row in wh_rows:
        old_id = row['id'] if 'id' in row.keys() else None
        slug = row['slug'] if 'slug' in row.keys() and row['slug'] else f'tenant-{tenant_id}-warehouse'
        cursor.execute('SELECT 1 FROM warehouses WHERE slug = ? AND tenant_id = ?', (slug, tenant_id))
        if cursor.fetchone():
            slug = _unique_warehouse_slug(cursor, f'{slug}-t{tenant_id}', tenant_id=tenant_id)
        new_id = _insert_row_with_overrides(
            cursor, 'warehouses', row, target_columns['warehouses'],
            {'tenant_id': tenant_id, 'slug': slug, 'created_at': row['created_at'] if 'created_at' in row.keys() else now},
            {'id'}
        )
        if old_id is not None:
            warehouse_map[old_id] = new_id
    if not warehouse_map:
        default_id = _ensure_default_warehouse_for_tenant(cursor, tenant_id)
    else:
        default_id = next(iter(warehouse_map.values()))

    if 'contacts' in available_tables:
        import_cursor.execute('SELECT * FROM contacts')
        rows = import_cursor.fetchall()
    else:
        rows = []
    for row in rows:
        old_id = row['id'] if 'id' in row.keys() else None
        old_wh = row['warehouse_id'] if 'warehouse_id' in row.keys() else None
        new_id = _insert_row_with_overrides(
            cursor, 'contacts', row, target_columns['contacts'],
            {'tenant_id': tenant_id, 'warehouse_id': warehouse_map.get(old_wh, default_id)},
            {'id'}
        )
        if old_id is not None:
            contact_map[old_id] = new_id
    details['contacts'] = len(rows)

    if 'materials' in available_tables:
        import_cursor.execute('SELECT * FROM materials')
        rows = import_cursor.fetchall()
    else:
        rows = []
    for row in rows:
        old_id = row['id'] if 'id' in row.keys() else None
        old_wh = row['warehouse_id'] if 'warehouse_id' in row.keys() else None
        new_id = _insert_row_with_overrides(
            cursor, 'materials', row, target_columns['materials'],
            {'tenant_id': tenant_id, 'warehouse_id': warehouse_map.get(old_wh, default_id)},
            {'id'}
        )
        if old_id is not None:
            material_map[old_id] = new_id
    details['materials'] = len(rows)

    if 'batches' in available_tables:
        import_cursor.execute('SELECT * FROM batches')
        rows = import_cursor.fetchall()
    else:
        rows = []
    for row in rows:
        old_id = row['id'] if 'id' in row.keys() else None
        old_mat = row['material_id'] if 'material_id' in row.keys() else None
        if old_mat not in material_map:
            continue
        old_contact = row['contact_id'] if 'contact_id' in row.keys() else None
        old_wh = row['warehouse_id'] if 'warehouse_id' in row.keys() else None
        new_id = _insert_row_with_overrides(
            cursor, 'batches', row, target_columns['batches'],
            {
                'tenant_id': tenant_id,
                'warehouse_id': warehouse_map.get(old_wh, default_id),
                'material_id': material_map[old_mat],
                'contact_id': contact_map.get(old_contact) if old_contact else None,
            },
            {'id'}
        )
        if old_id is not None:
            batch_map[old_id] = new_id
    details['batches'] = len(batch_map)

    if 'inventory_records' in available_tables:
        import_cursor.execute('SELECT * FROM inventory_records')
        rows = import_cursor.fetchall()
    else:
        rows = []
    for row in rows:
        old_id = row['id'] if 'id' in row.keys() else None
        old_mat = row['material_id'] if 'material_id' in row.keys() else None
        if old_mat not in material_map:
            continue
        old_contact = row['contact_id'] if 'contact_id' in row.keys() else None
        old_batch = row['batch_id'] if 'batch_id' in row.keys() else None
        old_wh = row['warehouse_id'] if 'warehouse_id' in row.keys() else None
        new_id = _insert_row_with_overrides(
            cursor, 'inventory_records', row, target_columns['inventory_records'],
            {
                'tenant_id': tenant_id,
                'warehouse_id': warehouse_map.get(old_wh, default_id),
                'material_id': material_map[old_mat],
                'contact_id': contact_map.get(old_contact) if old_contact else None,
                'batch_id': batch_map.get(old_batch) if old_batch else None,
            },
            {'id'}
        )
        if old_id is not None:
            record_map[old_id] = new_id
    details['inventory_records'] = len(record_map)

    if 'batch_consumptions' in available_tables:
        import_cursor.execute('SELECT * FROM batch_consumptions')
        rows = import_cursor.fetchall()
    else:
        rows = []
    imported_consumptions = 0
    for row in rows:
        old_record = row['record_id'] if 'record_id' in row.keys() else None
        old_batch = row['batch_id'] if 'batch_id' in row.keys() else None
        if old_record not in record_map or old_batch not in batch_map:
            continue
        old_wh = row['warehouse_id'] if 'warehouse_id' in row.keys() else None
        _insert_row_with_overrides(
            cursor, 'batch_consumptions', row, target_columns['batch_consumptions'],
            {
                'record_id': record_map[old_record],
                'batch_id': batch_map[old_batch],
                'tenant_id': tenant_id,
                'warehouse_id': warehouse_map.get(old_wh, default_id) if old_wh is not None else default_id,
            },
            {'id'}
        )
        imported_consumptions += 1
    details['batch_consumptions'] = imported_consumptions
    details['warehouses'] = len(warehouse_map) if warehouse_map else 1
    return details


@app.get("/api/database/export")
def export_database(
    target_tenant_id: Optional[int] = None,
    current_user: CurrentUser = Depends(require_permission(Resource.SYSTEM, Action.ADMIN))
):
    """导出仓库数据为SQLite数据库文件（仅管理员）

    只导出仓库相关表：materials, inventory_records, batches, batch_consumptions, contacts
    不导出用户相关表：users, sessions, api_keys

    NOTE: This endpoint is sqlite-only by design — it streams a literal
    ``.db`` file built from ``sqlite_master`` DDL. On non-sqlite deployments
    (e.g. MySQL) it returns 400.
    """
    import tempfile
    import sqlite3
    import shutil

    if get_engine().dialect.name != 'sqlite':
        raise HTTPException(
            status_code=400,
            detail="DB export is only available on the sqlite-backed deployment",
        )

    if current_user.tenant_id is None:
        if target_tenant_id is None:
            raise HTTPException(status_code=400, detail="全局管理员导出时必须通过 ?target_tenant_id= 指定目标租户")
        export_tenant_id = target_tenant_id
    else:
        export_tenant_id = current_user.tenant_id

    # 获取当前数据库路径 — 与 database.py / db.py / alembic env.py 保持一致，
    # 默认锚定到项目根，避免 CWD 不同造成"两个 db" 的踩坑。
    _default_db = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        'warehouse.db',
    )
    db_path = os.environ.get('DATABASE_PATH', _default_db)

    # 创建临时文件
    temp_fd, temp_path = tempfile.mkstemp(suffix='.db')
    os.close(temp_fd)

    try:
        # 创建新的临时数据库
        temp_conn = sqlite3.connect(temp_path)
        temp_cursor = temp_conn.cursor()

        with get_db() as source_conn:
            source_cursor = source_conn.cursor()

            for table in WAREHOUSE_TABLES:
                # 获取表结构
                source_cursor.execute(f"SELECT sql FROM sqlite_master WHERE type='table' AND name=?", (table,))
                result = source_cursor.fetchone()
                if result and result['sql']:
                    # 创建表
                    temp_cursor.execute(result['sql'])

                    rows = _export_rows_for_scope(source_cursor, table, export_tenant_id)
                    if rows:
                        columns = [desc[0] for desc in source_cursor.description]
                        placeholders = ','.join(['?' for _ in columns])
                        insert_sql = f"INSERT INTO {table} ({','.join(columns)}) VALUES ({placeholders})"
                        for row in rows:
                            temp_cursor.execute(insert_sql, tuple(row[col] for col in columns))

        temp_conn.commit()
        temp_conn.close()

        # 读取临时文件内容
        with open(temp_path, 'rb') as f:
            db_content = f.read()

        # 创建 BytesIO 对象用于流式响应
        output = BytesIO(db_content)
        output.seek(0)

        filename = f"warehouse_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"

        if ENABLE_AUDIT_LOG:
            logger.info(f"[AUDIT] 用户 {current_user.username or 'unknown'} 导出了数据库")

        return StreamingResponse(
            output,
            media_type="application/octet-stream",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    finally:
        # 清理临时文件
        if os.path.exists(temp_path):
            os.unlink(temp_path)


@app.post("/api/database/import", response_model=DatabaseOperationResponse)
@limiter.limit("5/minute")
async def import_database(
    request: Request,
    file: UploadFile = File(...),
    target_tenant_id: Optional[int] = None,
    current_user: CurrentUser = Depends(require_permission(Resource.SYSTEM, Action.ADMIN))
):
    """导入仓库数据（仅管理员）

    从上传的SQLite数据库文件中导入仓库相关表的数据。
    会清空现有仓库数据后再导入。
    不影响用户相关表：users, sessions, api_keys

    NOTE: This endpoint is sqlite-only by design — it operates on a literal
    ``.db`` upload using ``sqlite_master``. On non-sqlite deployments it
    returns 400.
    """
    import tempfile
    import sqlite3

    if get_engine().dialect.name != 'sqlite':
        raise HTTPException(
            status_code=400,
            detail="DB import is only available on the sqlite-backed deployment",
        )

    # 读取上传的文件
    contents = await file.read()

    # 检查文件大小
    file_size_mb = len(contents) / (1024 * 1024)
    if file_size_mb > MAX_UPLOAD_SIZE_MB:
        raise HTTPException(status_code=400, detail=f"文件过大，最大允许 {MAX_UPLOAD_SIZE_MB}MB")

    # 保存到临时文件
    temp_fd, temp_path = tempfile.mkstemp(suffix='.db')
    try:
        os.write(temp_fd, contents)
        os.close(temp_fd)

        # 验证是否为有效的SQLite数据库
        try:
            import_conn = sqlite3.connect(temp_path)
            import_conn.row_factory = sqlite3.Row
            import_cursor = import_conn.cursor()

            # 检查必要的表是否存在
            import_cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            available_tables = {row[0] for row in import_cursor.fetchall()}

            # 至少需要 materials 表
            if 'materials' not in available_tables:
                raise HTTPException(status_code=400, detail="无效的数据库文件：缺少 materials 表")

        except sqlite3.DatabaseError:
            raise HTTPException(status_code=400, detail="无效的数据库文件格式")

        # 开始导入
        details = {}

        with get_db() as conn:
            cursor = conn.cursor()

            try:
                if current_user.tenant_id is not None:
                    details = _import_tenant_database(cursor, import_cursor, available_tables, current_user.tenant_id)
                else:
                    # 全局 admin 必须显式指定目标租户，防止意外覆盖所有租户数据
                    if target_tenant_id is None:
                        raise HTTPException(status_code=400, detail="全局管理员导入时必须通过 ?target_tenant_id= 指定目标租户")
                    details = _import_tenant_database(cursor, import_cursor, available_tables, target_tenant_id)

                conn.commit()

            except Exception as e:
                conn.rollback()
                import traceback
                logger.error(f"[ERROR] 数据库导入失败: {str(e)}")
                logger.error(traceback.format_exc())
                raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")

        import_conn.close()

        if ENABLE_AUDIT_LOG:
            logger.info(f"[AUDIT] 用户 {current_user.username or 'unknown'} 导入了数据库")

        wh_count = details.get('warehouses', 0)
        wh_info = f"，{wh_count} 仓库" if wh_count else ""
        message = f"导入成功：{details.get('materials', 0)} 物料，{details.get('inventory_records', 0)} 记录，{details.get('batches', 0)} 批次，{details.get('contacts', 0)} 联系方{wh_info}"

        return DatabaseOperationResponse(
            success=True,
            message=message,
            details=details
        )

    finally:
        # 清理临时文件
        if os.path.exists(temp_path):
            os.unlink(temp_path)


@app.post("/api/database/clear", response_model=DatabaseOperationResponse)
async def clear_database(
    request: DatabaseClearRequest,
    current_user: CurrentUser = Depends(require_permission(Resource.SYSTEM, Action.ADMIN))
):
    """清空仓库数据（仅管理员）

    清空仓库相关表：materials, inventory_records, batches, batch_consumptions, contacts
    不影响用户相关表：users, sessions, api_keys
    """
    if not request.confirm:
        raise HTTPException(status_code=400, detail="请确认清空操作")

    if get_engine().dialect.name != 'sqlite':
        raise HTTPException(
            status_code=400,
            detail="DB clear is only available on the sqlite-backed deployment",
        )

    if current_user.tenant_id is None:
        if request.target_tenant_id is None:
            raise HTTPException(status_code=400, detail="全局管理员必须显式指定 target_tenant_id")
        scope_tenant_id = request.target_tenant_id
    else:
        scope_tenant_id = current_user.tenant_id

    with get_db() as conn:
        cursor = conn.cursor()

        try:
            details = _clear_database_scope(cursor, scope_tenant_id)
            conn.commit()

        except Exception as e:
            conn.rollback()
            raise HTTPException(status_code=500, detail=f"清空失败: {str(e)}")

    if ENABLE_AUDIT_LOG:
        logger.info(f"[AUDIT] 用户 {current_user.username or 'unknown'} 清空了数据库")

    message = f"已清空：{details.get('materials', 0)} 物料，{details.get('inventory_records', 0)} 记录，{details.get('batches', 0)} 批次，{details.get('contacts', 0)} 联系方"

    return DatabaseOperationResponse(
        success=True,
        message=message,
        details=details
    )


# ============ Contact Management APIs ============

@app.get("/api/contacts")
async def list_contacts(
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=10, le=100, description="每页条数"),
    name: Optional[str] = Query(None, description="名称模糊搜索"),
    contact_type: Optional[str] = Query(None, description="类型: supplier/customer/all"),
    include_disabled: bool = Query(False, description="是否包含禁用的联系方"),
    format: Optional[str] = Query(None, description="brief时精简返回"),
    current_user: CurrentUser = Depends(require_permission(Resource.CONTACTS, Action.READ)),
):
    """获取联系方列表（分页）— 联系方为租户级，不按仓库过滤。Phase 2c: SA Core reads."""
    conds = list(build_scope_predicates(_t_contacts, current_user.tenant_id, None))
    if not include_disabled:
        conds.append(_t_contacts.c.is_disabled == 0)
    if name:
        conds.append(_t_contacts.c.name.like(f'%{name}%'))
    if contact_type == 'supplier':
        conds.append(_t_contacts.c.is_supplier == 1)
    elif contact_type == 'customer':
        conds.append(_t_contacts.c.is_customer == 1)

    where_clause = and_(*conds) if conds else None
    count_stmt = select(_sa_func.count()).select_from(_t_contacts)
    list_stmt = select(
        _t_contacts.c.id, _t_contacts.c.name, _t_contacts.c.address,
        _t_contacts.c.phone, _t_contacts.c.email, _t_contacts.c.is_supplier,
        _t_contacts.c.is_customer, _t_contacts.c.notes,
        _t_contacts.c.is_disabled, _t_contacts.c.created_at,
    )
    if where_clause is not None:
        count_stmt = count_stmt.where(where_clause)
        list_stmt = list_stmt.where(where_clause)
    offset = (page - 1) * page_size
    list_stmt = list_stmt.order_by(_t_contacts.c.name.asc()).limit(page_size).offset(offset)

    with get_engine().connect() as sa_conn:
        total = sa_conn.execute(count_stmt).scalar() or 0
        rows = sa_conn.execute(list_stmt).fetchall()

    if format == "brief":
        items = [{"id": row.id, "name": row.name} for row in rows]
    else:
        items = []
        for row in rows:
            ca = row.created_at
            if isinstance(ca, datetime):
                ca = ca.strftime('%Y-%m-%d %H:%M:%S')
            items.append(ContactItem(
                id=row.id,
                name=row.name,
                address=row.address,
                phone=row.phone,
                email=row.email,
                is_supplier=bool(row.is_supplier),
                is_customer=bool(row.is_customer),
                notes=row.notes,
                is_disabled=bool(row.is_disabled),
                created_at=ca,
            ))

    total_pages = math.ceil(total / page_size) if total > 0 else 1

    return {
        "items": items,
        "page": page,
        "page_size": page_size,
        "total": total,
        "total_pages": total_pages,
    }


@app.get("/api/contacts/suppliers", response_model=List[ContactListItem])
async def list_suppliers(
    current_user: CurrentUser = Depends(require_permission(Resource.CONTACTS, Action.READ)),
):
    """获取供应商列表（用于下拉选择）— 联系方为租户级。Phase 2c: SA Core read."""
    conds = [_t_contacts.c.is_supplier == 1, _t_contacts.c.is_disabled == 0]
    conds.extend(build_scope_predicates(_t_contacts, current_user.tenant_id, None))
    stmt = select(
        _t_contacts.c.id, _t_contacts.c.name,
        _t_contacts.c.is_supplier, _t_contacts.c.is_customer,
    ).where(and_(*conds)).order_by(_t_contacts.c.name.asc())
    with get_engine().connect() as sa_conn:
        rows = sa_conn.execute(stmt).fetchall()
    return [
        ContactListItem(
            id=row.id,
            name=row.name,
            is_supplier=bool(row.is_supplier),
            is_customer=bool(row.is_customer),
        )
        for row in rows
    ]


@app.get("/api/contacts/customers", response_model=List[ContactListItem])
async def list_customers(
    current_user: CurrentUser = Depends(require_permission(Resource.CONTACTS, Action.READ)),
):
    """获取客户列表（用于下拉选择）— 联系方为租户级。Phase 2c: SA Core read."""
    conds = [_t_contacts.c.is_customer == 1, _t_contacts.c.is_disabled == 0]
    conds.extend(build_scope_predicates(_t_contacts, current_user.tenant_id, None))
    stmt = select(
        _t_contacts.c.id, _t_contacts.c.name,
        _t_contacts.c.is_supplier, _t_contacts.c.is_customer,
    ).where(and_(*conds)).order_by(_t_contacts.c.name.asc())
    with get_engine().connect() as sa_conn:
        rows = sa_conn.execute(stmt).fetchall()
    return [
        ContactListItem(
            id=row.id,
            name=row.name,
            is_supplier=bool(row.is_supplier),
            is_customer=bool(row.is_customer),
        )
        for row in rows
    ]


@app.get("/api/operators", response_model=List[OperatorListItem])
async def get_operators_for_filter(
    current_user: CurrentUser = Depends(require_permission(Resource.USERS, Action.READ))
):
    """获取操作员列表（用于筛选下拉）- 返回所有有操作权限的用户。Phase 2c: SA Core read."""
    conds = [
        _t_users.c.is_disabled == 0,
        _t_users.c.role.in_([RoleName.OPERATE.value, RoleName.ADMIN.value]),
    ]
    conds.extend(build_scope_predicates(_t_users, current_user.tenant_id, None))
    stmt = select(
        _t_users.c.id, _t_users.c.username, _t_users.c.display_name,
    ).where(and_(*conds)).order_by(_t_users.c.display_name, _t_users.c.username)
    with get_engine().connect() as sa_conn:
        rows = sa_conn.execute(stmt).fetchall()
    return [
        OperatorListItem(
            user_id=row.id,
            username=row.username,
            display_name=row.display_name,
        )
        for row in rows
    ]


# ---- contacts CRUD migrated to ResourceRouter (R2 phase 1) ----
# GET / POST / PUT / DELETE on /api/contacts/{id} are registered by the
# factory below. LIST stays as ``list_contacts`` above because of its
# resource-specific filters (name / contact_type / include_disabled /
# format=brief) and paginated response shape. ``list_suppliers`` and
# ``list_customers`` are also kept above (they aren't standard CRUD).

_CONTACT_GET_COLUMNS = [
    _t_contacts.c.id, _t_contacts.c.name, _t_contacts.c.address,
    _t_contacts.c.phone, _t_contacts.c.email, _t_contacts.c.is_supplier,
    _t_contacts.c.is_customer, _t_contacts.c.notes,
    _t_contacts.c.is_disabled, _t_contacts.c.created_at, _t_contacts.c.tenant_id,
]
_CONTACT_OUT_COLUMNS = [
    _t_contacts.c.id, _t_contacts.c.name, _t_contacts.c.address,
    _t_contacts.c.phone, _t_contacts.c.email, _t_contacts.c.is_supplier,
    _t_contacts.c.is_customer, _t_contacts.c.notes,
    _t_contacts.c.is_disabled, _t_contacts.c.created_at,
]


def _contact_to_out(row) -> ContactItem:
    ca = row.created_at
    if isinstance(ca, datetime):
        ca = ca.strftime('%Y-%m-%d %H:%M:%S')
    return ContactItem(
        id=row.id,
        name=row.name,
        address=row.address,
        phone=row.phone,
        email=row.email,
        is_supplier=bool(row.is_supplier),
        is_customer=bool(row.is_customer),
        notes=row.notes,
        is_disabled=bool(row.is_disabled),
        created_at=ca,
    )


def _contact_before_create(sa_conn, current_user, request: CreateContactRequest):
    if not request.is_supplier and not request.is_customer:
        raise HTTPException(status_code=400, detail="必须选择供应商或客户至少一项")


def _contact_values_for_create(sa_conn, current_user, request: CreateContactRequest) -> dict:
    # Tenant resolution: tenant users write to their own tenant; global
    # admin must specify request.tenant_id and the tenant must exist + be active.
    if current_user.tenant_id is not None:
        contact_tenant_id = current_user.tenant_id
    else:
        if request.tenant_id is None:
            raise HTTPException(
                status_code=400,
                detail="全局管理员创建联系方必须指定 tenant_id"
            )
        tenant_row = sa_conn.execute(
            select(_t_tenants.c.id).where(
                and_(_t_tenants.c.id == request.tenant_id, _t_tenants.c.is_active == 1)
            )
        ).first()
        if not tenant_row:
            raise HTTPException(status_code=400, detail="租户不存在或已停用")
        contact_tenant_id = request.tenant_id

    return {
        "name": request.name,
        "address": request.address,
        "phone": request.phone,
        "email": request.email,
        "is_supplier": 1 if request.is_supplier else 0,
        "is_customer": 1 if request.is_customer else 0,
        "notes": request.notes,
        "warehouse_id": None,
        "tenant_id": contact_tenant_id,
        "created_at": datetime.now(),
    }


def _contact_values_for_update(sa_conn, current_user, request: UpdateContactRequest, row) -> dict:
    values: dict = {}
    if request.name is not None:
        values['name'] = request.name
    if request.address is not None:
        values['address'] = request.address
    if request.phone is not None:
        values['phone'] = request.phone
    if request.email is not None:
        values['email'] = request.email
    if request.is_supplier is not None:
        values['is_supplier'] = 1 if request.is_supplier else 0
    if request.is_customer is not None:
        values['is_customer'] = 1 if request.is_customer else 0
    if request.notes is not None:
        values['notes'] = request.notes
    if request.is_disabled is not None:
        values['is_disabled'] = 1 if request.is_disabled else 0
    return values


def _contact_after_commit(operation, sa_conn, current_user, row_id):
    # R5: only contact partition is affected
    get_fuzzy_matcher().invalidate_cache(entity_type="contact")


from resource_router import ResourceRouter as _ResourceRouter  # noqa: E402

_ResourceRouter(
    app=app,
    prefix="/api/contacts",
    table=_t_contacts,
    response_model=ContactItem,
    create_model=CreateContactRequest,
    update_model=UpdateContactRequest,
    permission_read=require_permission(Resource.CONTACTS, Action.READ),
    permission_write=require_permission(Resource.CONTACTS, Action.WRITE),
    not_found_detail="联系方不存在",
    forbidden_detail="无权访问该联系方",
    to_out=_contact_to_out,
    values_for_create=_contact_values_for_create,
    values_for_update=_contact_values_for_update,
    before_create=_contact_before_create,
    after_commit=_contact_after_commit,
    get_columns=_CONTACT_GET_COLUMNS,
    update_select_columns=_CONTACT_OUT_COLUMNS,
    delete_response={"success": True, "message": "联系方已禁用"},
).register()


# ============ Dashboard APIs ============

@app.get("/api/dashboard/stats", response_model=DashboardStats)
def get_dashboard_stats(
    warehouse_id: Optional[int] = Query(None),
    current_user: CurrentUser = Depends(require_permission(Resource.DASHBOARD, Action.READ))
):
    """获取仪表盘统计数据（排除禁用物料）— Phase 2e: SA Core read."""
    wh_id = resolve_warehouse_id(current_user, warehouse_id)
    m_scope = list(build_authorized_scope_predicates(_t_materials, current_user, wh_id))
    r_scope = list(build_authorized_scope_predicates(_t_inventory_records, current_user, wh_id))

    today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_start_s = today_start.strftime('%Y-%m-%d %H:%M:%S')
    yesterday_start_s = (datetime.now() - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0).strftime('%Y-%m-%d %H:%M:%S')

    # 单一真相源：active batches 聚合
    batch_sum = (
        select(
            _t_batches.c.material_id.label('material_id'),
            _sa_func.coalesce(_sa_func.sum(_t_batches.c.quantity), 0).label('qty'),
        )
        .where(_t_batches.c.is_exhausted == 0)
        .group_by(_t_batches.c.material_id)
        .subquery()
    )

    with get_engine().connect() as sa_conn:
        # 库存总量（排除禁用） — 用 active batches 聚合
        j_total = _t_materials.outerjoin(batch_sum, batch_sum.c.material_id == _t_materials.c.id)
        total_stock = sa_conn.execute(
            select(_sa_func.coalesce(_sa_func.sum(batch_sum.c.qty), 0))
            .select_from(j_total)
            .where(and_(_t_materials.c.is_disabled == 0, *m_scope))
        ).scalar() or 0

        # 今日入库量
        j = _t_inventory_records.join(_t_materials, _t_inventory_records.c.material_id == _t_materials.c.id)
        today_in = sa_conn.execute(
            select(_sa_func.sum(_t_inventory_records.c.quantity)).select_from(j)
            .where(and_(
                _t_inventory_records.c.type == RecordType.IN.value,
                _t_inventory_records.c.created_at >= today_start_s,
                _t_materials.c.is_disabled == 0,
                *r_scope,
            ))
        ).scalar() or 0

        today_out = sa_conn.execute(
            select(_sa_func.sum(_t_inventory_records.c.quantity)).select_from(j)
            .where(and_(
                _t_inventory_records.c.type == RecordType.OUT.value,
                _t_inventory_records.c.created_at >= today_start_s,
                _t_materials.c.is_disabled == 0,
                *r_scope,
            ))
        ).scalar() or 0

        # 库存预警 — 比较 active batches sum 与 safe_stock
        j_low = _t_materials.outerjoin(batch_sum, batch_sum.c.material_id == _t_materials.c.id)
        low_stock_count = sa_conn.execute(
            select(_sa_func.count()).select_from(j_low)
            .where(and_(
                _t_materials.c.safe_stock.is_not(None),
                _sa_func.coalesce(batch_sum.c.qty, 0) < _t_materials.c.safe_stock,
                _t_materials.c.is_disabled == 0,
                *m_scope,
            ))
        ).scalar() or 0

        # 物料种类数
        material_types = sa_conn.execute(
            select(_sa_func.count()).select_from(_t_materials)
            .where(and_(_t_materials.c.is_disabled == 0, *m_scope))
        ).scalar() or 0

        # 昨日入库
        yesterday_in = sa_conn.execute(
            select(_sa_func.sum(_t_inventory_records.c.quantity))
            .where(and_(
                _t_inventory_records.c.type == RecordType.IN.value,
                _t_inventory_records.c.created_at >= yesterday_start_s,
                _t_inventory_records.c.created_at < today_start_s,
                *r_scope,
            ))
        ).scalar() or 1

        yesterday_out = sa_conn.execute(
            select(_sa_func.sum(_t_inventory_records.c.quantity))
            .where(and_(
                _t_inventory_records.c.type == RecordType.OUT.value,
                _t_inventory_records.c.created_at >= yesterday_start_s,
                _t_inventory_records.c.created_at < today_start_s,
                *r_scope,
            ))
        ).scalar() or 1

        in_change = round(((today_in - yesterday_in) / yesterday_in * 100), 1) if yesterday_in > 0 else 0
        out_change = round(((today_out - yesterday_out) / yesterday_out * 100), 1) if yesterday_out > 0 else 0

        return DashboardStats(
            total_stock=total_stock,
            today_in=today_in,
            today_out=today_out,
            low_stock_count=low_stock_count,
            material_types=material_types,
            in_change=in_change,
            out_change=out_change
        )


@app.get("/api/dashboard/category-distribution", response_model=List[CategoryItem])
def get_category_distribution(
    warehouse_id: Optional[int] = Query(None),
    current_user: CurrentUser = Depends(require_permission(Resource.DASHBOARD, Action.READ))
):
    """获取库存类型分布 — Phase 2e: SA Core read."""
    wh_id = resolve_warehouse_id(current_user, warehouse_id)
    preds = list(build_authorized_scope_predicates(_t_materials, current_user, wh_id))
    # 单一真相源：active batches 聚合
    batch_sum = (
        select(
            _t_batches.c.material_id.label('material_id'),
            _sa_func.coalesce(_sa_func.sum(_t_batches.c.quantity), 0).label('qty'),
        )
        .where(_t_batches.c.is_exhausted == 0)
        .group_by(_t_batches.c.material_id)
        .subquery()
    )
    j_cat = _t_materials.outerjoin(batch_sum, batch_sum.c.material_id == _t_materials.c.id)
    total_col = _sa_func.sum(_sa_func.coalesce(batch_sum.c.qty, 0)).label('total')
    stmt = (
        select(_t_materials.c.category, total_col)
        .select_from(j_cat)
        .group_by(_t_materials.c.category)
        .order_by(total_col.desc())
    )
    if preds:
        stmt = stmt.where(and_(*preds))
    with get_engine().connect() as sa_conn:
        rows = sa_conn.execute(stmt).fetchall()
        return [CategoryItem(name=row.category, value=row.total) for row in rows]


@app.get("/api/dashboard/weekly-trend", response_model=WeeklyTrend)
def get_weekly_trend(
    warehouse_id: Optional[int] = Query(None),
    current_user: CurrentUser = Depends(require_permission(Resource.DASHBOARD, Action.READ))
):
    """获取近7天出入库趋势 — Phase 2e: SA Core read."""
    wh_id = resolve_warehouse_id(current_user, warehouse_id)
    scope_preds = list(build_authorized_scope_predicates(_t_inventory_records, current_user, wh_id))

    dates = []
    in_data = []
    out_data = []
    with get_engine().connect() as sa_conn:
        for i in range(6, -1, -1):
            date = datetime.now() - timedelta(days=i)
            date_start = date.replace(hour=0, minute=0, second=0, microsecond=0)
            date_end = date_start + timedelta(days=1)
            ds = date_start.strftime('%Y-%m-%d %H:%M:%S')
            de = date_end.strftime('%Y-%m-%d %H:%M:%S')

            dates.append(date.strftime('%m-%d'))

            in_preds = [
                _t_inventory_records.c.type == RecordType.IN.value,
                _t_inventory_records.c.created_at >= ds,
                _t_inventory_records.c.created_at < de,
            ] + scope_preds
            in_total = sa_conn.execute(
                select(_sa_func.sum(_t_inventory_records.c.quantity)).where(and_(*in_preds))
            ).scalar() or 0
            in_data.append(in_total)

            out_preds = [
                _t_inventory_records.c.type == RecordType.OUT.value,
                _t_inventory_records.c.created_at >= ds,
                _t_inventory_records.c.created_at < de,
            ] + scope_preds
            out_total = sa_conn.execute(
                select(_sa_func.sum(_t_inventory_records.c.quantity)).where(and_(*out_preds))
            ).scalar() or 0
            out_data.append(out_total)

    return WeeklyTrend(dates=dates, in_data=in_data, out_data=out_data)


@app.get("/api/dashboard/top-stock", response_model=TopStock)
def get_top_stock(
    warehouse_id: Optional[int] = Query(None),
    current_user: CurrentUser = Depends(require_permission(Resource.DASHBOARD, Action.READ))
):
    """获取库存TOP10 — Phase 2e: SA Core read."""
    wh_id = resolve_warehouse_id(current_user, warehouse_id)
    preds = list(build_authorized_scope_predicates(_t_materials, current_user, wh_id))
    batch_sum = (
        select(
            _t_batches.c.material_id.label('material_id'),
            _sa_func.coalesce(_sa_func.sum(_t_batches.c.quantity), 0).label('qty'),
        )
        .where(_t_batches.c.is_exhausted == 0)
        .group_by(_t_batches.c.material_id)
        .subquery()
    )
    j_top = _t_materials.outerjoin(batch_sum, batch_sum.c.material_id == _t_materials.c.id)
    qty_col = _sa_func.coalesce(batch_sum.c.qty, 0).label('qty')
    stmt = (
        select(_t_materials.c.name, qty_col, _t_materials.c.category)
        .select_from(j_top)
        .order_by(qty_col.desc())
        .limit(10)
    )
    if preds:
        stmt = stmt.where(and_(*preds))
    with get_engine().connect() as sa_conn:
        rows = sa_conn.execute(stmt).fetchall()
        names = [row.name for row in rows]
        quantities = [int(row.qty or 0) for row in rows]
        categories = [row.category for row in rows]
        return TopStock(names=names, quantities=quantities, categories=categories)


@app.get("/api/dashboard/low-stock-alert", response_model=List[LowStockItem])
def get_low_stock_alert(
    warehouse_id: Optional[int] = Query(None),
    current_user: CurrentUser = Depends(require_permission(Resource.DASHBOARD, Action.READ))
):
    """获取库存预警列表 — Phase 2e: SA Core read."""
    wh_id = resolve_warehouse_id(current_user, warehouse_id)
    batch_sum = (
        select(
            _t_batches.c.material_id.label('material_id'),
            _sa_func.coalesce(_sa_func.sum(_t_batches.c.quantity), 0).label('qty'),
        )
        .where(_t_batches.c.is_exhausted == 0)
        .group_by(_t_batches.c.material_id)
        .subquery()
    )
    qty_col = _sa_func.coalesce(batch_sum.c.qty, 0)
    preds = [
        _t_materials.c.safe_stock.is_not(None),
        qty_col < _t_materials.c.safe_stock,
        _t_materials.c.is_disabled == 0,
    ]
    preds.extend(build_authorized_scope_predicates(_t_materials, current_user, wh_id))
    j_lsa = _t_materials.outerjoin(batch_sum, batch_sum.c.material_id == _t_materials.c.id)
    stmt = (
        select(
            _t_materials.c.name, _t_materials.c.sku, _t_materials.c.category,
            qty_col.label('quantity'), _t_materials.c.safe_stock, _t_materials.c.location,
        )
        .select_from(j_lsa)
        .where(and_(*preds))
        .order_by((qty_col - _t_materials.c.safe_stock).asc())
        .limit(20)
    )
    with get_engine().connect() as sa_conn:
        return [
            LowStockItem(
                name=row.name,
                sku=row.sku,
                category=row.category,
                quantity=int(row.quantity or 0),
                safe_stock=row.safe_stock,
                location=row.location,
                shortage=row.safe_stock - int(row.quantity or 0)
            )
            for row in sa_conn.execute(stmt).fetchall()
        ]


# ============ Fuzzy Match & Search APIs ============

@app.get("/api/fuzzy-match", response_model=FuzzyMatchResponse)
def fuzzy_match_endpoint(
    q: str = Query(..., description="搜索文本"),
    entity_type: str = Query("all", description="实体类型: material/contact/operator/all"),
    top_k: int = Query(5, ge=1, le=50, description="返回前k个结果"),
    threshold: float = Query(50.0, ge=0, le=100, description="最低分数阈值"),
    warehouse_id: Optional[int] = Query(None, description="仓库ID"),
    current_user: CurrentUser = Depends(require_permission(Resource.SEARCH, Action.READ))
):
    """模糊匹配搜索（按当前租户/仓库范围过滤候选）"""
    wh_id = resolve_warehouse_id(current_user, warehouse_id)
    matcher = get_fuzzy_matcher()
    result = matcher.resolve(q, entity_type=entity_type,
                             tenant_id=current_user.tenant_id, warehouse_id=wh_id)
    candidates_raw = matcher.search(q, entity_type=entity_type, top_k=top_k, threshold=threshold,
                                    tenant_id=current_user.tenant_id, warehouse_id=wh_id)

    candidates = [FuzzyMatchCandidate(**c) for c in candidates_raw]
    best_match = FuzzyMatchCandidate(**result['best_match']) if result['best_match'] else None

    if result['confident'] and best_match:
        message = f"找到最佳匹配: {best_match.name} (置信度: {best_match.score})"
    elif candidates:
        message = f"找到 {len(candidates)} 个候选项，请确认选择"
    else:
        message = f"未找到与 '{q}' 匹配的结果"

    return FuzzyMatchResponse(
        query=q,
        candidates=candidates,
        best_match=best_match,
        confident=result['confident'],
        message=message
    )


@app.get("/api/search")
def unified_search(
    q: str = Query(None, description="搜索文本"),
    entity_type: str = Query("material", description="实体类型: material/contact/operator"),
    category: str = Query(None, description="分类（仅material）"),
    status: str = Query(None, description="状态（仅material，逗号分隔）"),
    contact_type: str = Query(None, description="联系方类型: supplier/customer"),
    fuzzy: bool = Query(True, description="是否开启模糊匹配"),
    format: str = Query(None, description="brief时只返回核心字段"),
    include_batches: bool = Query(False, description="是否附带批次列表（仅material）"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页条数"),
    warehouse_id: Optional[int] = Query(None, description="仓库ID"),
    current_user: CurrentUser = Depends(require_permission(Resource.SEARCH, Action.READ))
):
    """统一搜索端点"""
    wh_id = resolve_warehouse_id(current_user, warehouse_id)
    tenant_id = current_user.tenant_id
    # Phase 4: search helpers are SA Core internally; the legacy
    # cursor / scope_filter / scope_params positional args are kept on the
    # helper signatures for backwards compatibility but unused.
    if entity_type == "material":
        return _search_materials(None, q, category, status, fuzzy, format, include_batches,
                                 page, page_size, '', (),
                                 tenant_id=tenant_id, warehouse_id=wh_id)
    elif entity_type == "contact":
        # 联系方为租户级（无 wh 过滤）
        return _search_contacts(None, q, contact_type, fuzzy, format, page, page_size,
                                '', (), tenant_id=tenant_id)
    elif entity_type == "operator":
        # users 表无 warehouse_id 列，只按 tenant 过滤
        return _search_operators(None, q, fuzzy, format, page, page_size,
                                 '', (), tenant_id=tenant_id)
    else:
        raise HTTPException(status_code=400, detail=f"不支持的实体类型: {entity_type}")


def _search_materials(cursor, q, category, status, fuzzy, fmt, include_batches, page, page_size, wh_filter='', wh_params=(), tenant_id=None, warehouse_id=None):
    """搜索物料 — Phase 2d: SA Core read. ``cursor``/``wh_filter``/``wh_params`` retained for signature compatibility but unused."""
    # 获取匹配的 material IDs (fuzzy mode)
    matched_ids = None
    if q and fuzzy:
        matcher = get_fuzzy_matcher()
        results = matcher.search(q, entity_type="material", top_k=100, threshold=50.0,
                                 tenant_id=tenant_id, warehouse_id=warehouse_id)
        matched_ids = [r['entity_id'] for r in results]
        if not matched_ids:
            return {"items": [], "page": page, "page_size": page_size, "total": 0, "total_pages": 1}

    preds = [_t_materials.c.is_disabled == 0]
    preds.extend(build_scope_predicates(_t_materials, tenant_id, warehouse_id))
    if matched_ids is not None:
        preds.append(_t_materials.c.id.in_(matched_ids))
    elif q and not fuzzy:
        like = f'%{q}%'
        preds.append(or_(_t_materials.c.name.like(like), _t_materials.c.sku.like(like)))
    if category:
        preds.append(_t_materials.c.category == category)

    # 单一真相源：active batches 聚合作为 quantity
    batch_sum = (
        select(
            _t_batches.c.material_id.label('material_id'),
            _sa_func.coalesce(_sa_func.sum(_t_batches.c.quantity), 0).label('qty'),
        )
        .where(_t_batches.c.is_exhausted == 0)
        .group_by(_t_batches.c.material_id)
        .subquery()
    )
    qty_col = _sa_func.coalesce(batch_sum.c.qty, 0).label('quantity')
    cols = [
        _t_materials.c.id, _t_materials.c.name, _t_materials.c.sku,
        _t_materials.c.category, qty_col, _t_materials.c.unit,
        _t_materials.c.safe_stock, _t_materials.c.location, _t_materials.c.is_disabled,
    ]
    j_mat = _t_materials.outerjoin(batch_sum, batch_sum.c.material_id == _t_materials.c.id)
    base_stmt = select(*cols).select_from(j_mat).where(and_(*preds)).order_by(_t_materials.c.name.asc())

    status_filter = status.split(',') if status else None

    with get_engine().connect() as sa_conn:
        if status_filter:
            rows = sa_conn.execute(base_stmt).fetchall()
            all_items = []
            for row in rows:
                qty = row.quantity
                ss = row.safe_stock
                if ss is not None:
                    if qty >= ss:
                        item_status = 'normal'
                    elif qty >= ss * 0.5:
                        item_status = 'warning'
                    else:
                        item_status = 'danger'
                else:
                    item_status = 'normal'

                if item_status not in status_filter:
                    continue

                if fmt == "brief":
                    all_items.append({"id": row.id, "name": row.name, "sku": row.sku})
                else:
                    all_items.append({
                        "id": row.id, "name": row.name, "sku": row.sku,
                        "category": row.category, "quantity": qty, "unit": row.unit,
                        "safe_stock": ss, "location": row.location, "status": item_status,
                    })

            total = len(all_items)
            total_pages = math.ceil(total / page_size) if total > 0 else 1
            offset = (page - 1) * page_size
            items = all_items[offset:offset + page_size]
        else:
            count_stmt = select(_sa_func.count()).select_from(_t_materials).where(and_(*preds))
            total = sa_conn.execute(count_stmt).scalar() or 0

            offset = (page - 1) * page_size
            paged = base_stmt.limit(page_size).offset(offset)
            rows = sa_conn.execute(paged).fetchall()

            items = []
            for row in rows:
                qty = row.quantity
                ss = row.safe_stock
                if ss is not None:
                    if qty >= ss:
                        item_status = 'normal'
                    elif qty >= ss * 0.5:
                        item_status = 'warning'
                    else:
                        item_status = 'danger'
                else:
                    item_status = 'normal'

                if fmt == "brief":
                    items.append({"id": row.id, "name": row.name, "sku": row.sku})
                else:
                    items.append({
                        "id": row.id, "name": row.name, "sku": row.sku,
                        "category": row.category, "quantity": qty, "unit": row.unit,
                        "safe_stock": ss, "location": row.location, "status": item_status,
                    })

            total_pages = math.ceil(total / page_size) if total > 0 else 1

        # 批量加载批次信息（一次 SQL 替代 N 次 HTTP）
        if include_batches and items:
            material_ids = [item['id'] for item in items]
            batch_stmt = (
                select(
                    _t_batches.c.material_id, _t_batches.c.batch_no,
                    _t_batches.c.quantity, _t_batches.c.location,
                    _t_batches.c.variant,
                    _t_contacts.c.name.label('contact_name'),
                )
                .select_from(
                    _t_batches.outerjoin(_t_contacts, _t_batches.c.contact_id == _t_contacts.c.id)
                )
                .where(and_(
                    _t_batches.c.material_id.in_(material_ids),
                    _t_batches.c.is_exhausted == 0,
                    _t_batches.c.quantity > 0,
                ))
                .order_by(_t_batches.c.created_at.asc())
            )
            batches_by_material = {}
            for row in sa_conn.execute(batch_stmt).fetchall():
                batches_by_material.setdefault(row.material_id, []).append({
                    'batch_no': row.batch_no,
                    'quantity': row.quantity,
                    'location': row.location or '',
                    'variant': row.variant or '',
                    'contact_name': row.contact_name or '',
                })
            for item in items:
                item['batches'] = batches_by_material.get(item['id'], [])

    return {"items": items, "page": page, "page_size": page_size, "total": total, "total_pages": total_pages}


def _search_contacts(cursor, q, contact_type, fuzzy, fmt, page, page_size, scope_filter='', scope_params=(), tenant_id=None):
    """搜索联系方（租户级） — Phase 2d: SA Core read. ``cursor``/``scope_filter``/``scope_params`` retained for signature compatibility but unused."""
    matched_ids = None
    if q and fuzzy:
        matcher = get_fuzzy_matcher()
        # 联系方为租户级，不传 warehouse_id
        results = matcher.search(q, entity_type="contact", top_k=100, threshold=50.0,
                                 tenant_id=tenant_id)
        matched_ids = [r['entity_id'] for r in results]
        if not matched_ids:
            return {"items": [], "page": page, "page_size": page_size, "total": 0, "total_pages": 1}

    preds = [_t_contacts.c.is_disabled == 0]
    preds.extend(build_scope_predicates(_t_contacts, tenant_id, None))
    if matched_ids is not None:
        preds.append(_t_contacts.c.id.in_(matched_ids))
    elif q and not fuzzy:
        preds.append(_t_contacts.c.name.like(f'%{q}%'))

    if contact_type == 'supplier':
        preds.append(_t_contacts.c.is_supplier == 1)
    elif contact_type == 'customer':
        preds.append(_t_contacts.c.is_customer == 1)

    cols = [
        _t_contacts.c.id, _t_contacts.c.name, _t_contacts.c.address,
        _t_contacts.c.phone, _t_contacts.c.email,
        _t_contacts.c.is_supplier, _t_contacts.c.is_customer,
        _t_contacts.c.notes, _t_contacts.c.is_disabled, _t_contacts.c.created_at,
    ]
    where = and_(*preds)
    count_stmt = select(_sa_func.count()).select_from(_t_contacts).where(where)
    offset = (page - 1) * page_size
    page_stmt = (
        select(*cols).where(where)
        .order_by(_t_contacts.c.name.asc())
        .limit(page_size).offset(offset)
    )

    with get_engine().connect() as sa_conn:
        total = sa_conn.execute(count_stmt).scalar() or 0
        rows = sa_conn.execute(page_stmt).fetchall()

    if fmt == "brief":
        items = [{"id": row.id, "name": row.name} for row in rows]
    else:
        items = [{
            "id": row.id, "name": row.name, "address": row.address,
            "phone": row.phone, "email": row.email,
            "is_supplier": bool(row.is_supplier), "is_customer": bool(row.is_customer),
            "notes": row.notes, "is_disabled": bool(row.is_disabled),
            "created_at": (row.created_at.strftime('%Y-%m-%d %H:%M:%S')
                           if isinstance(row.created_at, datetime) else row.created_at),
        } for row in rows]

    total_pages = math.ceil(total / page_size) if total > 0 else 1
    return {"items": items, "page": page, "page_size": page_size, "total": total, "total_pages": total_pages}


def _search_operators(cursor, q, fuzzy, fmt, page, page_size, scope_filter='', scope_params=(), tenant_id=None):
    """搜索操作员（按 tenant 过滤；users 表无 warehouse_id） — Phase 2d: SA Core read. ``cursor``/``scope_filter``/``scope_params`` retained for signature compatibility but unused."""
    matched_ids = None
    if q and fuzzy:
        matcher = get_fuzzy_matcher()
        results = matcher.search(q, entity_type="operator", top_k=100, threshold=50.0,
                                 tenant_id=tenant_id)
        matched_ids = [r['entity_id'] for r in results]
        if not matched_ids:
            return {"items": [], "page": page, "page_size": page_size, "total": 0, "total_pages": 1}

    preds = [_t_users.c.is_disabled == 0]
    preds.extend(build_scope_predicates(_t_users, tenant_id, None))
    if matched_ids is not None:
        preds.append(_t_users.c.id.in_(matched_ids))
    elif q and not fuzzy:
        like = f'%{q}%'
        preds.append(or_(_t_users.c.username.like(like), _t_users.c.display_name.like(like)))

    where = and_(*preds)
    count_stmt = select(_sa_func.count()).select_from(_t_users).where(where)
    offset = (page - 1) * page_size
    page_stmt = (
        select(_t_users.c.id, _t_users.c.username, _t_users.c.display_name)
        .where(where)
        .order_by(_t_users.c.username.asc())
        .limit(page_size).offset(offset)
    )

    with get_engine().connect() as sa_conn:
        total = sa_conn.execute(count_stmt).scalar() or 0
        rows = sa_conn.execute(page_stmt).fetchall()

    if fmt == "brief":
        items = [{"id": row.id, "name": row.display_name or row.username} for row in rows]
    else:
        items = [{
            "id": row.id, "username": row.username,
            "display_name": row.display_name,
            "name": row.display_name or row.username,
        } for row in rows]

    total_pages = math.ceil(total / page_size) if total > 0 else 1
    return {"items": items, "page": page, "page_size": page_size, "total": total, "total_pages": total_pages}


# ============ Materials APIs ============

@app.get("/api/materials/all", response_model=List[MaterialItem])
def get_all_materials(
    warehouse_id: Optional[int] = Query(None, description="仓库ID"),
    current_user: CurrentUser = Depends(require_permission(Resource.MATERIALS, Action.READ))
):
    """获取所有库存（兼容旧API）— Phase 2e: SA Core read."""
    wh_id = resolve_warehouse_id(current_user, warehouse_id)
    preds = [_t_materials.c.is_disabled == 0]
    preds.extend(build_authorized_scope_predicates(_t_materials, current_user, wh_id))
    batch_sum = (
        select(
            _t_batches.c.material_id.label('material_id'),
            _sa_func.coalesce(_sa_func.sum(_t_batches.c.quantity), 0).label('qty'),
        )
        .where(_t_batches.c.is_exhausted == 0)
        .group_by(_t_batches.c.material_id)
        .subquery()
    )
    j_all = _t_materials.outerjoin(batch_sum, batch_sum.c.material_id == _t_materials.c.id)
    qty_col = _sa_func.coalesce(batch_sum.c.qty, 0).label('quantity')
    stmt = select(
        _t_materials.c.name, _t_materials.c.sku, _t_materials.c.category,
        qty_col, _t_materials.c.unit, _t_materials.c.safe_stock,
        _t_materials.c.location, _t_materials.c.is_disabled,
    ).select_from(j_all).where(and_(*preds)).order_by(_t_materials.c.name.asc())
    with get_engine().connect() as sa_conn:
        rows = sa_conn.execute(stmt).fetchall()

        result = []
        for row in rows:
            quantity = int(row.quantity or 0)
            safe_stock = row.safe_stock

            # 判断状态
            if safe_stock is not None:
                if quantity >= safe_stock:
                    status = 'normal'
                    status_text = '正常'
                elif quantity >= safe_stock * 0.5:
                    status = 'warning'
                    status_text = '偏低'
                else:
                    status = 'danger'
                    status_text = '告急'
            else:
                status = 'normal'
                status_text = '正常'

            result.append(MaterialItem(
                name=row.name,
                sku=row.sku,
                category=row.category,
                quantity=quantity,
                unit=row.unit,
                safe_stock=safe_stock,
                location=row.location,
                status=status,
                status_text=status_text
            ))

        return result


@app.get("/api/materials/list")
def get_materials_list(
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=10, le=100, description="每页条数"),
    name: Optional[str] = Query(None, description="名称/SKU模糊搜索"),
    category: Optional[str] = Query(None, description="分类"),
    status: Optional[str] = Query(None, description="状态(逗号分隔: normal,warning,danger,disabled)"),
    min_stock: Optional[int] = Query(None, description="最小库存过滤"),
    max_stock: Optional[int] = Query(None, description="最大库存过滤"),
    location: Optional[str] = Query(None, description="位置模糊匹配"),
    fuzzy: bool = Query(True, description="名称模糊匹配开关"),
    format: Optional[str] = Query(None, description="brief时精简返回"),
    warehouse_id: Optional[int] = Query(None, description="仓库ID"),
    group_by_sku: bool = Query(False, description="按SKU聚合：每个物料一行，批次/位置/变体合并展示"),
    current_user: CurrentUser = Depends(require_permission(Resource.MATERIALS, Action.READ))
):
    """获取物料列表（分页+筛选）— 默认一行一批次；group_by_sku=true 时一行一物料."""
    wh_id = resolve_warehouse_id(current_user, warehouse_id)

    status_filter = status.split(',') if status else None

    # Fuzzy name search
    fuzzy_ids = None
    if name and fuzzy:
        matcher = get_fuzzy_matcher()
        results = matcher.search(name, entity_type="material", top_k=100, threshold=50.0,
                                 tenant_id=current_user.tenant_id, warehouse_id=wh_id)
        fuzzy_ids = [r['entity_id'] for r in results]
        if not fuzzy_ids:
            return {"items": [], "page": page, "page_size": page_size, "total": 0, "total_pages": 1}

    # 构建物料筛选条件
    preds = list(build_authorized_scope_predicates(_t_materials, current_user, wh_id))

    if not status_filter or 'disabled' not in status_filter:
        preds.append(_t_materials.c.is_disabled == 0)

    if fuzzy_ids is not None:
        preds.append(_t_materials.c.id.in_(fuzzy_ids))
    elif name and not fuzzy:
        like = f'%{name}%'
        preds.append(or_(_t_materials.c.name.like(like), _t_materials.c.sku.like(like)))

    if category:
        preds.append(_t_materials.c.category == category)

    if location:
        loc_like = f'%{location}%'
        preds.append(or_(_t_batches.c.location.like(loc_like), _t_materials.c.location.like(loc_like)))

    # 查询：一行一批次（LEFT JOIN batches）
    # 单一真相源：total_quantity 用 active batches 聚合
    batch_sum = (
        select(
            _t_batches.c.material_id.label('material_id'),
            _sa_func.coalesce(_sa_func.sum(_t_batches.c.quantity), 0).label('qty'),
        )
        .where(_t_batches.c.is_exhausted == 0)
        .group_by(_t_batches.c.material_id)
        .subquery()
    )
    join_expr = (
        _t_materials
        .outerjoin(batch_sum, batch_sum.c.material_id == _t_materials.c.id)
        .outerjoin(
            _t_batches,
            and_(_t_batches.c.material_id == _t_materials.c.id, _t_batches.c.is_exhausted == 0),
        )
        .outerjoin(_t_contacts, _t_batches.c.contact_id == _t_contacts.c.id)
        .outerjoin(_t_warehouses, _t_materials.c.warehouse_id == _t_warehouses.c.id)
    )

    stmt = (
        select(
            _t_materials.c.id.label('material_id'),
            _t_materials.c.name,
            _t_materials.c.sku,
            _t_materials.c.category,
            _sa_func.coalesce(batch_sum.c.qty, 0).label('total_quantity'),
            _t_materials.c.unit,
            _t_materials.c.safe_stock,
            _t_materials.c.location.label('material_location'),
            _t_materials.c.is_disabled,
            _t_batches.c.batch_no,
            _t_batches.c.quantity.label('batch_quantity'),
            _t_batches.c.location.label('batch_location'),
            _t_batches.c.variant,
            _t_contacts.c.name.label('contact_name'),
            _t_materials.c.warehouse_id,
            _t_warehouses.c.name.label('warehouse_name'),
        )
        .select_from(join_expr)
        .where(and_(*preds) if preds else and_())
        .order_by(_t_materials.c.name.asc(), _t_batches.c.created_at.asc())
    )

    with get_engine().connect() as sa_conn:
        all_rows = sa_conn.execute(stmt).fetchall()

    # 应用状态筛选和库存范围筛选（在应用层做，因为状态是计算值）
    filtered = []
    for row in all_rows:
        total_qty = row.total_quantity
        safe_stock_val = row.safe_stock
        is_disabled = bool(row.is_disabled)

        if is_disabled:
            item_status = 'disabled'
        elif safe_stock_val is not None:
            if total_qty >= safe_stock_val:
                item_status = 'normal'
            elif total_qty >= safe_stock_val * 0.5:
                item_status = 'warning'
            else:
                item_status = 'danger'
        else:
            item_status = 'normal'

        if status_filter and item_status not in status_filter:
            continue
        if min_stock is not None and total_qty < min_stock:
            continue
        if max_stock is not None and total_qty > max_stock:
            continue

        filtered.append((row, item_status))

    # 聚合模式：按 material_id 合并多批次为单行
    if group_by_sku and format != "brief":
        grouped: Dict[int, Dict[str, Any]] = {}
        for row, item_status in filtered:
            mid = row.material_id
            g = grouped.get(mid)
            if g is None:
                g = {
                    'row': row,
                    'status': item_status,
                    'batch_nos': set(),
                    'locations': set(),
                    'variants': set(),
                }
                grouped[mid] = g
            if row.batch_no:
                g['batch_nos'].add(row.batch_no)
                # location/variant 仅统计活跃批次的非空值
                loc = (row.batch_location or '').strip()
                if loc:
                    g['locations'].add(loc)
                var = (row.variant or '').strip()
                if var:
                    g['variants'].add(var)
        filtered = [(g['row'], g['status'], g) for g in grouped.values()]
        # 保持物料名升序（query 已 order_by name）
        filtered.sort(key=lambda x: (x[0].name or '', x[0].sku or ''))
    else:
        filtered = [(r, s, None) for r, s in filtered]

    total = len(filtered)
    total_pages = math.ceil(total / page_size) if total > 0 else 1
    offset = (page - 1) * page_size
    page_rows = filtered[offset:offset + page_size]

    result = []
    for row, item_status, agg in page_rows:
        is_disabled = bool(row.is_disabled)
        status_text_map = {'normal': '正常', 'warning': '偏低', 'danger': '告急', 'disabled': '禁用'}

        # 无批次时返回 0（不再 fallback 到 materials.quantity 派生值）
        batch_qty = row.batch_quantity if row.batch_quantity is not None else 0
        batch_loc = row.batch_location if row.batch_location else (row.material_location or '')

        if format == "brief":
            result.append({"id": row.material_id, "name": row.name, "sku": row.sku})
        elif agg is not None:
            # 聚合行：一行一物料
            batch_nos = agg['batch_nos']
            locations = agg['locations']
            variants = agg['variants']
            loc_mixed = len(locations) > 1
            var_mixed = len(variants) > 1
            single_loc = next(iter(locations)) if len(locations) == 1 else ''
            single_var = next(iter(variants)) if len(variants) == 1 else ''
            single_batch_no = next(iter(batch_nos)) if len(batch_nos) == 1 else ''
            result.append(MaterialItemWithDisabled(
                name=row.name,
                sku=row.sku,
                category=row.category,
                quantity=row.total_quantity or 0,  # 聚合视图：用总库存
                unit=row.unit,
                safe_stock=row.safe_stock,
                location='' if loc_mixed else (single_loc or (row.material_location or '')),
                status=item_status,
                status_text=status_text_map.get(item_status, ''),
                is_disabled=is_disabled,
                batch_no=single_batch_no,
                contact_name='',
                total_quantity=row.total_quantity,
                variant='' if var_mixed else single_var,
                warehouse_id=row.warehouse_id,
                warehouse_name=row.warehouse_name,
                batch_count=len(batch_nos),
                location_mixed=loc_mixed,
                variant_mixed=var_mixed,
            ))
        else:
            result.append(MaterialItemWithDisabled(
                name=row.name,
                sku=row.sku,
                category=row.category,
                quantity=batch_qty,
                unit=row.unit,
                safe_stock=row.safe_stock,
                location=batch_loc,
                status=item_status,
                status_text=status_text_map.get(item_status, ''),
                is_disabled=is_disabled,
                batch_no=row.batch_no or '',
                contact_name=row.contact_name or '',
                total_quantity=row.total_quantity,
                variant=row.variant or '',
                warehouse_id=row.warehouse_id,
                warehouse_name=row.warehouse_name,
            ))

    return {
        "items": result,
        "page": page,
        "page_size": page_size,
        "total": total,
        "total_pages": total_pages,
    }


@app.get("/api/materials/categories", response_model=List[str])
def get_categories(
    warehouse_id: Optional[int] = Query(None, description="仓库ID"),
    current_user: CurrentUser = Depends(require_permission(Resource.MATERIALS, Action.READ))
):
    """获取所有物料分类 — Phase 2e: SA Core read."""
    wh_id = resolve_warehouse_id(current_user, warehouse_id)
    preds = list(build_authorized_scope_predicates(_t_materials, current_user, wh_id))
    stmt = select(_t_materials.c.category).distinct().order_by(_t_materials.c.category)
    if preds:
        stmt = stmt.where(and_(*preds))
    with get_engine().connect() as sa_conn:
        return [row.category for row in sa_conn.execute(stmt).fetchall()]


@app.get("/api/materials/product-stats", response_model=ProductStats)
def get_product_stats(
    name: str = Query(..., description="产品名称"),
    warehouse_id: Optional[int] = Query(None, description="仓库ID"),
    current_user: CurrentUser = Depends(require_permission(Resource.MATERIALS, Action.READ))
):
    """获取单个产品的统计数据 — Phase 2e: SA Core read."""
    if not name:
        raise HTTPException(status_code=400, detail="缺少产品名称参数")

    wh_id = resolve_warehouse_id(current_user, warehouse_id)
    m_scope = list(build_authorized_scope_predicates(_t_materials, current_user, wh_id))

    today = datetime.now().strftime('%Y-%m-%d')
    yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')

    with get_engine().connect() as sa_conn:
        # 查询产品基本信息（支持 name 或 SKU）
        m_stmt = select(
            _t_materials.c.id, _t_materials.c.name, _t_materials.c.sku,
            # 注：保留 quantity 列只为兼容（过渡期 cache），库存以下方 batches 聚合为准
            _t_materials.c.quantity, _t_materials.c.unit,
            _t_materials.c.safe_stock, _t_materials.c.location,
        ).where(and_(
            or_(_t_materials.c.name == name, _t_materials.c.sku == name),
            *m_scope,
        ))
        product = sa_conn.execute(m_stmt).fetchone()
        if not product:
            raise HTTPException(status_code=404, detail="产品不存在")

        material_id = product.id
        # 单一真相源：从 active batches 聚合得到当前库存（不读 materials.quantity）。
        # 这是 watcher 的入口，必须与 /api/materials/list 在有批次时的聚合行为一致。
        current_stock = int(sa_conn.execute(
            select(_sa_func.coalesce(_sa_func.sum(_t_batches.c.quantity), 0))
            .where(and_(
                _t_batches.c.material_id == material_id,
                _t_batches.c.is_exhausted == 0,
            ))
        ).scalar() or 0)
        unit = product.unit
        safe_stock = product.safe_stock

        def _sum_records(rtype: str, date_str: Optional[str] = None) -> int:
            preds = [
                _t_inventory_records.c.material_id == material_id,
                _t_inventory_records.c.type == rtype,
            ]
            if date_str is not None:
                preds.append(_sa_func.date(_t_inventory_records.c.created_at) == date_str)
            return sa_conn.execute(
                select(_sa_func.coalesce(_sa_func.sum(_t_inventory_records.c.quantity), 0))
                .where(and_(*preds))
            ).scalar() or 0

        today_in = _sum_records(RecordType.IN.value, today)
        yesterday_in = _sum_records(RecordType.IN.value, yesterday)
        today_out = _sum_records(RecordType.OUT.value, today)
        yesterday_out = _sum_records(RecordType.OUT.value, yesterday)
        total_in = _sum_records(RecordType.IN.value)
        total_out = _sum_records(RecordType.OUT.value)

        in_change = ((today_in - yesterday_in) / yesterday_in * 100) if yesterday_in > 0 else 0
        out_change = ((today_out - yesterday_out) / yesterday_out * 100) if yesterday_out > 0 else 0

        return ProductStats(
            name=product.name,
            sku=product.sku,
            current_stock=current_stock,
            unit=unit,
            safe_stock=safe_stock,
            location=product.location,
            today_in=today_in,
            today_out=today_out,
            in_change=round(in_change, 1),
            out_change=round(out_change, 1),
            total_in=total_in,
            total_out=total_out
        )


@app.get("/api/materials/batches")
def get_material_batches(
    name: str = Query(..., description="产品名称"),
    warehouse_id: Optional[int] = Query(None, description="仓库ID"),
    current_user: CurrentUser = Depends(require_permission(Resource.MATERIALS, Action.READ))
):
    """获取物料的活跃批次列表 — Phase 2f: SA Core read."""
    wh_id = resolve_warehouse_id(current_user, warehouse_id)
    if wh_id is not None:
        with get_db() as conn:
            check_warehouse_access(conn, current_user, wh_id)
    m_scope = list(build_authorized_scope_predicates(_t_materials, current_user, wh_id))
    b_scope = list(build_authorized_scope_predicates(_t_batches, current_user, wh_id))
    with get_engine().connect() as sa_conn:
        material = sa_conn.execute(
            select(_t_materials.c.id).where(and_(_t_materials.c.name == name, *m_scope))
        ).first()
        if not material:
            raise HTTPException(status_code=404, detail="产品不存在")
        rows = sa_conn.execute(
            select(
                _t_batches.c.batch_no, _t_batches.c.quantity, _t_batches.c.location,
                _t_batches.c.created_at, _t_batches.c.variant,
                _t_contacts.c.name.label('contact_name'),
            ).select_from(
                _t_batches.outerjoin(_t_contacts, _t_batches.c.contact_id == _t_contacts.c.id)
            ).where(and_(
                _t_batches.c.material_id == material.id,
                _t_batches.c.is_exhausted == 0,
                *b_scope,
            )).order_by(_t_batches.c.created_at.asc())
        ).fetchall()

    total_quantity = sum(b.quantity for b in rows)
    return {
        "batches": [
            {
                "batch_no": b.batch_no,
                "quantity": b.quantity,
                "location": b.location or '',
                "contact_name": b.contact_name or '',
                "created_at": (b.created_at.strftime('%Y-%m-%d %H:%M:%S')
                               if isinstance(b.created_at, datetime) else b.created_at),
                "variant": b.variant or '',
            }
            for b in rows
        ],
        "total_quantity": total_quantity,
    }


@app.get("/api/batches/by-no", response_model=BatchDetailResponse)
def get_batch_by_no(
    batch_no: str = Query(..., description="批次号（精确匹配）"),
    warehouse_id: Optional[int] = Query(None, description="仓库 ID（可选；不传则跨仓查租户内）"),
    include_exhausted: bool = Query(True, description="是否包含已耗尽批次（默认 True，方便溯源）"),
    current_user: CurrentUser = Depends(require_permission(Resource.MATERIALS, Action.READ))
):
    """按批次号查询批次详情。

    - 200 + success=true：找到（即使已耗尽也算找到，is_exhausted 区分）
    - 200 + success=false + error="batch_not_found"：作用域内确实没有该批次
    设计选择：MCP 上游需要按 batch_no 直接查询，不依赖产品名。
    error 走 success=false 而不是 HTTP 404，以便 MCP wrap 成 speak_failed 而非 transport 错误。
    """
    bn = (batch_no or '').strip()
    if not bn:
        raise HTTPException(status_code=400, detail="缺少 batch_no 参数")

    wh_id = resolve_warehouse_id(current_user, warehouse_id)
    if wh_id is not None:
        with get_db() as conn:
            check_warehouse_access(conn, current_user, wh_id)

    b_scope = list(build_authorized_scope_predicates(_t_batches, current_user, wh_id))
    preds = [_t_batches.c.batch_no == bn, *b_scope]
    if not include_exhausted:
        preds.append(_t_batches.c.is_exhausted == 0)

    with get_engine().connect() as sa_conn:
        row = sa_conn.execute(
            select(
                _t_batches.c.batch_no, _t_batches.c.quantity,
                _t_batches.c.initial_quantity, _t_batches.c.location,
                _t_batches.c.variant, _t_batches.c.is_exhausted,
                _t_batches.c.created_at,
                _t_materials.c.name.label('material_name'),
                _t_materials.c.sku.label('material_sku'),
                _t_materials.c.unit,
                _t_warehouses.c.name.label('warehouse_name'),
                _t_contacts.c.name.label('contact_name'),
            ).select_from(
                _t_batches
                .join(_t_materials, _t_batches.c.material_id == _t_materials.c.id)
                .outerjoin(_t_warehouses, _t_batches.c.warehouse_id == _t_warehouses.c.id)
                .outerjoin(_t_contacts, _t_batches.c.contact_id == _t_contacts.c.id)
            ).where(and_(*preds))
        ).first()

    if not row:
        return BatchDetailResponse(
            success=False,
            error="batch_not_found",
            message=f"未找到批次 '{bn}'",
        )

    created_at_str = (row.created_at.strftime('%Y-%m-%d %H:%M:%S')
                      if isinstance(row.created_at, datetime) else (row.created_at or ''))
    return BatchDetailResponse(
        success=True,
        batch=BatchDetailItem(
            batch_no=row.batch_no,
            quantity=row.quantity,
            initial_quantity=row.initial_quantity,
            location=row.location or None,
            variant=row.variant or None,
            is_exhausted=bool(row.is_exhausted),
            material_name=row.material_name,
            material_sku=row.material_sku or None,
            unit=row.unit,
            warehouse_name=row.warehouse_name or None,
            contact_name=row.contact_name or None,
            created_at=created_at_str,
        ),
        message=f"批次 '{bn}' 查询成功",
    )


@app.get("/api/materials/product-trend", response_model=WeeklyTrend)
def get_product_trend(
    name: str = Query(..., description="产品名称"),
    warehouse_id: Optional[int] = Query(None),
    current_user: CurrentUser = Depends(require_permission(Resource.MATERIALS, Action.READ))
):
    """获取单个产品的近7天趋势"""
    if not name:
        raise HTTPException(status_code=400, detail="缺少产品名称参数")

    wh_id = resolve_warehouse_id(current_user, warehouse_id)
    # Phase 2f: SA Core read.
    m_scope = list(build_authorized_scope_predicates(_t_materials, current_user, wh_id))
    r_scope = list(build_authorized_scope_predicates(_t_inventory_records, current_user, wh_id))

    with get_engine().connect() as sa_conn:
        product = sa_conn.execute(
            select(_t_materials.c.id).where(and_(_t_materials.c.name == name, *m_scope))
        ).first()
        if not product:
            raise HTTPException(status_code=404, detail="产品不存在")

        material_id = product.id

        dates = []
        in_data = []
        out_data = []
        for i in range(6, -1, -1):
            d = datetime.now() - timedelta(days=i)
            dates.append(d.strftime('%m-%d'))
            day = d.strftime('%Y-%m-%d')

            in_total = sa_conn.execute(
                select(_sa_func.coalesce(_sa_func.sum(_t_inventory_records.c.quantity), 0))
                .where(and_(
                    _t_inventory_records.c.material_id == material_id,
                    _t_inventory_records.c.type == RecordType.IN.value,
                    _sa_func.date(_t_inventory_records.c.created_at) == day,
                    *r_scope,
                ))
            ).scalar() or 0
            in_data.append(in_total)

            out_total = sa_conn.execute(
                select(_sa_func.coalesce(_sa_func.sum(_t_inventory_records.c.quantity), 0))
                .where(and_(
                    _t_inventory_records.c.material_id == material_id,
                    _t_inventory_records.c.type == RecordType.OUT.value,
                    _sa_func.date(_t_inventory_records.c.created_at) == day,
                    *r_scope,
                ))
            ).scalar() or 0
            out_data.append(out_total)

        return WeeklyTrend(dates=dates, in_data=in_data, out_data=out_data)


@app.get("/api/materials/product-records", response_model=PaginatedProductRecordsResponse)
def get_product_records(
    name: str = Query(..., description="产品名称"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=10, le=100, description="每页条数"),
    warehouse_id: Optional[int] = Query(None, description="仓库ID"),
    current_user: CurrentUser = Depends(require_permission(Resource.MATERIALS, Action.READ))
):
    """获取单个产品的出入库记录（分页）— Phase 2e: SA Core read."""
    if not name:
        raise HTTPException(status_code=400, detail="缺少产品名称参数")

    wh_id = resolve_warehouse_id(current_user, warehouse_id)
    r_scope = list(build_authorized_scope_predicates(_t_inventory_records, current_user, wh_id))
    m_scope = list(build_authorized_scope_predicates(_t_materials, current_user, wh_id))

    with get_engine().connect() as sa_conn:
        product = sa_conn.execute(
            select(_t_materials.c.id).where(and_(_t_materials.c.name == name, *m_scope))
        ).fetchone()
        if not product:
            raise HTTPException(status_code=404, detail="产品不存在")

        material_id = product.id
        base_preds = [_t_inventory_records.c.material_id == material_id, *r_scope]

        total = sa_conn.execute(
            select(_sa_func.count()).select_from(_t_inventory_records).where(and_(*base_preds))
        ).scalar() or 0

        offset = (page - 1) * page_size
        j = _t_inventory_records.outerjoin(_t_batches, _t_inventory_records.c.batch_id == _t_batches.c.id)
        rows = sa_conn.execute(
            select(
                _t_inventory_records.c.type, _t_inventory_records.c.quantity,
                _t_inventory_records.c.operator, _t_inventory_records.c.reason_category,
                _t_inventory_records.c.reason_note, _t_inventory_records.c.created_at,
                _t_batches.c.variant, _t_batches.c.batch_no,
            ).select_from(j)
            .where(and_(*base_preds))
            .order_by(_t_inventory_records.c.created_at.desc())
            .limit(page_size).offset(offset)
        ).fetchall()

        items = [
            ProductRecord(
                type=row.type,
                quantity=row.quantity,
                operator=row.operator,
                reason_category=row.reason_category,
                reason_note=row.reason_note,
                created_at=row.created_at.strftime('%Y-%m-%d %H:%M:%S') if isinstance(row.created_at, datetime) else row.created_at,
                variant=row.variant or '',
                batch_no=row.batch_no or '',
            )
            for row in rows
        ]

        total_pages = math.ceil(total / page_size) if total > 0 else 1

        return PaginatedProductRecordsResponse(
            items=items,
            page=page,
            page_size=page_size,
            total=total,
            total_pages=total_pages
        )


@app.get("/api/reason-categories")
def get_reason_categories(current_user: CurrentUser = Depends(require_permission(Resource.INVENTORY, Action.READ))):
    """获取出入库原因分类列表"""
    return {
        "in": [{"key": k, "label": REASON_CATEGORY_LABELS[k]} for k in REASON_CATEGORIES["in"]],
        "out": [{"key": k, "label": REASON_CATEGORY_LABELS[k]} for k in REASON_CATEGORIES["out"]],
    }


@app.get("/api/inventory/records")
def get_inventory_records_paginated(
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=10, le=100, description="每页条数"),
    start_date: Optional[str] = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
    product_name: Optional[str] = Query(None, description="产品名称/SKU模糊搜索"),
    category: Optional[str] = Query(None, description="商品类型/分类"),
    record_type: Optional[str] = Query(None, description="记录类型: in/out"),
    status: Optional[str] = Query(None, description="状态(逗号分隔: normal,warning,danger,disabled)"),
    contact_id: Optional[int] = Query(None, description="联系方ID筛选"),
    operator_user_id: Optional[int] = Query(None, description="操作员用户ID筛选"),
    reason_category: Optional[str] = Query(None, description="原因分类筛选"),
    reason: Optional[str] = Query(None, description="原因/备注关键词搜索"),
    sort_by: str = Query("created_at", description="排序字段: created_at/quantity/material_name"),
    sort_order: str = Query("desc", description="排序方向: asc/desc"),
    format: Optional[str] = Query(None, description="brief时精简返回"),
    warehouse_id: Optional[int] = Query(None, description="仓库ID"),
    current_user: CurrentUser = Depends(require_permission(Resource.INVENTORY, Action.READ))
):
    """获取所有进出库记录（分页+筛选）— Phase 2e: SA Core read."""
    wh_id = resolve_warehouse_id(current_user, warehouse_id)
    r_scope = list(build_authorized_scope_predicates(_t_inventory_records, current_user, wh_id))

    # 解析状态筛选
    status_filter = status.split(',') if status else None

    # 构建过滤谓词（不含状态筛选 — 该项需在 Python 端二次过滤）
    preds = list(r_scope)
    if start_date:
        preds.append(_sa_func.date(_t_inventory_records.c.created_at) >= start_date)
    if end_date:
        preds.append(_sa_func.date(_t_inventory_records.c.created_at) <= end_date)
    if product_name:
        like = f'%{product_name}%'
        preds.append(or_(_t_materials.c.name.like(like), _t_materials.c.sku.like(like)))
    if category:
        preds.append(_t_materials.c.category == category)
    if record_type:
        preds.append(_t_inventory_records.c.type == record_type)
    if contact_id:
        preds.append(_t_inventory_records.c.contact_id == contact_id)
    if operator_user_id:
        preds.append(_t_inventory_records.c.operator_user_id == operator_user_id)
    if reason_category:
        preds.append(_t_inventory_records.c.reason_category == reason_category)
    if reason:
        rl = f'%{reason}%'
        preds.append(or_(
            _t_inventory_records.c.reason_note.like(rl),
            _t_inventory_records.c.reason_category.like(rl),
        ))

    # 单一真相源：active batches 聚合，避免 N+1 与 materials.quantity 脏值
    batch_sum = (
        select(
            _t_batches.c.material_id.label('material_id'),
            _sa_func.coalesce(_sa_func.sum(_t_batches.c.quantity), 0).label('qty'),
        )
        .where(_t_batches.c.is_exhausted == 0)
        .group_by(_t_batches.c.material_id)
        .subquery()
    )

    # 主查询 join
    j = (
        _t_inventory_records
        .join(_t_materials, _t_inventory_records.c.material_id == _t_materials.c.id)
        .outerjoin(_t_contacts, _t_inventory_records.c.contact_id == _t_contacts.c.id)
        .outerjoin(_t_batches, _t_inventory_records.c.batch_id == _t_batches.c.id)
        .outerjoin(_t_users, _t_inventory_records.c.operator_user_id == _t_users.c.id)
        .outerjoin(_t_warehouses, _t_inventory_records.c.warehouse_id == _t_warehouses.c.id)
        .outerjoin(batch_sum, batch_sum.c.material_id == _t_materials.c.id)
    )

    sort_column_map = {
        'created_at': _t_inventory_records.c.created_at,
        'quantity': _t_inventory_records.c.quantity,
        'material_name': _t_materials.c.name,
    }
    sort_col = sort_column_map.get(sort_by, _t_inventory_records.c.created_at)
    sort_expr = sort_col.asc() if sort_order.lower() == 'asc' else sort_col.desc()

    cols = [
        _t_inventory_records.c.id,
        _t_materials.c.name.label('material_name'),
        _t_materials.c.sku.label('material_sku'),
        _t_materials.c.category.label('category'),
        _t_inventory_records.c.type,
        _t_inventory_records.c.quantity,
        _t_inventory_records.c.operator,
        _t_inventory_records.c.operator_user_id,
        _t_inventory_records.c.reason_category,
        _t_inventory_records.c.reason_note,
        _t_inventory_records.c.created_at,
        _sa_func.coalesce(batch_sum.c.qty, 0).label('current_quantity'),
        _t_materials.c.safe_stock,
        _t_materials.c.is_disabled,
        _t_inventory_records.c.contact_id,
        _t_contacts.c.name.label('contact_name'),
        _t_inventory_records.c.batch_id,
        _t_batches.c.batch_no,
        _t_batches.c.variant,
        _t_users.c.display_name.label('operator_display_name'),
        _t_users.c.username.label('operator_username'),
        _t_inventory_records.c.warehouse_id,
        _t_warehouses.c.name.label('warehouse_name'),
    ]

    count_join = _t_inventory_records.join(_t_materials, _t_inventory_records.c.material_id == _t_materials.c.id)

    offset = (page - 1) * page_size

    with get_engine().connect() as sa_conn:
        # 获取总数
        count_stmt = select(_sa_func.count()).select_from(count_join)
        if preds:
            count_stmt = count_stmt.where(and_(*preds))
        total = sa_conn.execute(count_stmt).scalar() or 0

        # 主查询
        main_stmt = select(*cols).select_from(j)
        if preds:
            main_stmt = main_stmt.where(and_(*preds))
        main_stmt = main_stmt.order_by(sort_expr).limit(page_size).offset(offset)
        rows = sa_conn.execute(main_stmt).fetchall()

        result = []
        for row in rows:
            quantity = row.current_quantity
            safe_stock = row.safe_stock
            is_disabled = bool(row.is_disabled)

            if is_disabled:
                material_status = 'disabled'
            elif safe_stock is not None:
                if quantity >= safe_stock:
                    material_status = 'normal'
                elif quantity >= safe_stock * 0.5:
                    material_status = 'warning'
                else:
                    material_status = 'danger'
            else:
                material_status = 'normal'

            if status_filter and material_status not in status_filter:
                continue

            batch_details = None
            record_id = row.id
            record_type_val = row.type
            ca = row.created_at
            if isinstance(ca, datetime):
                ca = ca.strftime('%Y-%m-%d %H:%M:%S')

            if record_type_val == RecordType.OUT.value:
                cj = _t_batch_consumptions.join(_t_batches, _t_batch_consumptions.c.batch_id == _t_batches.c.id)
                consumptions = sa_conn.execute(
                    select(_t_batches.c.batch_no, _t_batch_consumptions.c.quantity)
                    .select_from(cj)
                    .where(_t_batch_consumptions.c.record_id == record_id)
                    .order_by(_t_batches.c.created_at.asc())
                ).fetchall()
                if consumptions:
                    details = [f"{c.batch_no}×{c.quantity}" for c in consumptions]
                    batch_details = ', '.join(details)

            operator_name = row.operator_display_name or row.operator_username or row.operator

            if format == "brief":
                result.append({
                    "id": record_id,
                    "material_name": row.material_name,
                    "type": record_type_val,
                    "quantity": row.quantity,
                    "created_at": ca,
                })
            else:
                result.append(InventoryRecordItem(
                    id=record_id,
                    material_name=row.material_name,
                    material_sku=row.material_sku,
                    category=row.category,
                    type=record_type_val,
                    quantity=row.quantity,
                    operator=row.operator,
                    operator_user_id=row.operator_user_id,
                    operator_name=operator_name,
                    reason_category=row.reason_category,
                    reason_note=row.reason_note,
                    created_at=ca,
                    material_status=material_status,
                    is_disabled=is_disabled,
                    contact_id=row.contact_id,
                    contact_name=row.contact_name,
                    batch_id=row.batch_id,
                    batch_no=row.batch_no,
                    batch_details=batch_details,
                    variant=row.variant or '',
                    warehouse_id=row.warehouse_id,
                    warehouse_name=row.warehouse_name,
                ))

        # 如果有状态筛选，需要重新计算总数（与原逻辑一致：仅 start/end_date、product_name、record_type 参与该子查询）
        if status_filter:
            count_preds = list(r_scope)
            if start_date:
                count_preds.append(_sa_func.date(_t_inventory_records.c.created_at) >= start_date)
            if end_date:
                count_preds.append(_sa_func.date(_t_inventory_records.c.created_at) <= end_date)
            if product_name:
                like = f'%{product_name}%'
                count_preds.append(or_(_t_materials.c.name.like(like), _t_materials.c.sku.like(like)))
            if record_type:
                count_preds.append(_t_inventory_records.c.type == record_type)

            stat_join = count_join.outerjoin(
                batch_sum, batch_sum.c.material_id == _t_materials.c.id
            )
            stat_stmt = select(
                _sa_func.coalesce(batch_sum.c.qty, 0).label('quantity'),
                _t_materials.c.safe_stock, _t_materials.c.is_disabled
            ).select_from(stat_join)
            if count_preds:
                stat_stmt = stat_stmt.where(and_(*count_preds))
            all_rows = sa_conn.execute(stat_stmt).fetchall()
            total = 0
            for rr in all_rows:
                qty = rr.quantity
                ss = rr.safe_stock
                dis = bool(rr.is_disabled)
                if dis:
                    s = 'disabled'
                elif ss is not None:
                    if qty >= ss:
                        s = 'normal'
                    elif qty >= ss * 0.5:
                        s = 'warning'
                    else:
                        s = 'danger'
                else:
                    s = 'normal'
                if s in status_filter:
                    total += 1

    total_pages = math.ceil(total / page_size) if total > 0 else 1

    if format == "brief":
        return {"items": result, "page": page, "page_size": page_size, "total": total, "total_pages": total_pages}
    return PaginatedRecordsResponse(
        items=result,
        page=page,
        page_size=page_size,
        total=total,
        total_pages=total_pages
    )


# ============ Stock Operation APIs (for MCP) ============

@app.post("/api/materials/stock-in", response_model=StockInResponse)
async def stock_in(
    request: StockOperationRequest,
    current_user: CurrentUser = Depends(require_permission(Resource.INVENTORY, Action.WRITE))
):
    """入库操作（需要operate权限）- 自动创建批次，支持模糊匹配"""
    product_name = request.product_name
    quantity = request.quantity
    reason_category = request.reason_category
    reason_note = request.reason_note
    operator = request.operator if request.operator and request.operator != "MCP系统" else current_user.get_operator_name()
    operator_user_id = current_user.id
    resolved_from = None
    resolved_material_id = None
    wh_id = require_warehouse_id(current_user, request.warehouse_id)

    if quantity <= 0:
        return StockInResponse(
            success=False,
            error="入库数量必须大于0",
            message=f"入库失败：数量 {quantity} 无效"
        )

    check_warehouse_access(None, current_user, wh_id)
    # 校验 contact_id 跨租户归属（防止恶意/前端误传）
    ensure_contact_tenant(None, current_user, request.contact_id,
                          resolve_tenant_id_for_write(current_user, wh_id))

    m_scope = build_authorized_scope_predicates(_t_materials, current_user, wh_id)

    # 查询产品（先精确匹配，按仓库过滤；排除已禁用物料）
    with get_engine().connect() as sa_conn:
        row = sa_conn.execute(
            select(_t_materials.c.id, _t_materials.c.name, _t_materials.c.unit).where(
                and_(
                    or_(_t_materials.c.name == product_name, _t_materials.c.sku == product_name),
                    _t_materials.c.is_disabled == 0,
                    *m_scope,
                )
            )
        ).first()
        if row and row.name != product_name:
            resolved_from = product_name
            product_name = row.name

    # 模糊匹配
    if not row and request.fuzzy:
        matcher = get_fuzzy_matcher()
        result = matcher.resolve(product_name, entity_type="material",
                                 tenant_id=current_user.tenant_id, warehouse_id=wh_id)

        if result['confident'] and result['best_match']:
            resolved_from = product_name
            best = result['best_match']
            extra = best.get('extra') or {}
            resolved_material_id = best.get('entity_id')
            product_name = extra.get('canonical_name') or best['name']
            with get_engine().connect() as sa_conn:
                row_preds = [_t_materials.c.is_disabled == 0, *m_scope]
                if resolved_material_id is not None:
                    row_preds.append(_t_materials.c.id == resolved_material_id)
                else:
                    row_preds.append(_t_materials.c.name == product_name)
                row = sa_conn.execute(
                    select(_t_materials.c.id, _t_materials.c.name, _t_materials.c.unit).where(
                        and_(*row_preds)
                    )
                ).first()
                if row:
                    product_name = row.name
        elif result['candidates']:
            names = [c['name'] for c in result['candidates'][:5]]
            return StockInResponse(
                success=False,
                error="ambiguous_name",
                message=f"无法确定产品 '{product_name}'，候选：{', '.join(names)}",
                candidates=result['candidates'],
            )

    if not row:
        return StockInResponse(
            success=False,
            error=f"产品不存在: {product_name}",
            message=f"入库失败：未找到产品 '{product_name}'"
        )

    material_id = row.id
    unit = row.unit

    record_tenant_id = resolve_tenant_id_for_write(current_user, wh_id)
    batch_no = request.batch_no.strip() if request.batch_no and request.batch_no.strip() else generate_batch_no(material_id, warehouse_id=wh_id)
    now_dt = datetime.now()

    with get_engine().begin() as sa_conn:
        # 单一真相源：从 active batches 聚合读取入库前库存（不再写 materials.quantity）。
        old_quantity = int(sa_conn.execute(
            select(_sa_func.coalesce(_sa_func.sum(_t_batches.c.quantity), 0))
            .where(and_(
                _t_batches.c.material_id == material_id,
                _t_batches.c.is_exhausted == 0,
            ))
        ).scalar() or 0)
        new_quantity = old_quantity + quantity

        ins_batch = sa_conn.execute(
            insert(_t_batches).values(
                batch_no=batch_no, material_id=material_id, quantity=quantity,
                initial_quantity=quantity, contact_id=request.contact_id,
                location=request.location, variant=request.variant,
                warehouse_id=wh_id, tenant_id=record_tenant_id, created_at=now_dt,
            )
        )
        batch_id = ins_batch.inserted_primary_key[0]

        sa_conn.execute(
            insert(_t_inventory_records).values(
                material_id=material_id, type=RecordType.IN.value, quantity=quantity,
                operator=operator, operator_user_id=operator_user_id,
                reason_category=reason_category, reason_note=reason_note,
                contact_id=request.contact_id, batch_id=batch_id,
                warehouse_id=wh_id, tenant_id=record_tenant_id, created_at=now_dt,
            )
        )

    # R5: stock-in only affects material partition (name+variant entries)
    get_fuzzy_matcher().invalidate_cache(
        entity_type="material", tenant_id=record_tenant_id, warehouse_id=wh_id,
    )

    audit_log("STOCK_IN", current_user.id, current_user.username, {
        "product": product_name,
        "quantity": quantity,
        "batch_no": batch_no,
        "old_qty": old_quantity,
        "new_qty": new_quantity,
        "resolved_from": resolved_from,
    })

    return StockInResponse(
        success=True,
        operation="stock_in",
        product=StockOperationProduct(
            name=product_name,
            old_quantity=old_quantity,
            in_quantity=quantity,
            new_quantity=new_quantity,
            unit=unit
        ),
        batch=BatchInfo(batch_no=batch_no, batch_id=batch_id, quantity=quantity, variant=request.variant),
        message=f"入库成功：{product_name} 入库 {quantity} {unit}（批次 {batch_no}），库存从 {old_quantity} 更新到 {new_quantity} {unit}",
        resolved_from=resolved_from,
    )


@app.post("/api/materials/stock-out", response_model=StockOutResponse)
@limiter.limit("60/minute")
async def stock_out(
    request: Request,
    stock_data: StockOperationRequest,
    current_user: CurrentUser = Depends(require_permission(Resource.INVENTORY, Action.WRITE))
):
    """出库操作（需要operate权限）- FIFO批次消耗，支持模糊匹配、指定批次。"""
    product_name = stock_data.product_name
    quantity = stock_data.quantity
    reason_category = stock_data.reason_category
    reason_note = stock_data.reason_note
    operator = stock_data.operator if stock_data.operator and stock_data.operator != "MCP系统" else current_user.get_operator_name()
    operator_user_id = current_user.id
    resolved_from = None
    resolved_variant = None
    resolved_material_id = None
    wh_id = require_warehouse_id(current_user, stock_data.warehouse_id)

    if quantity <= 0:
        return StockOutResponse(success=False, error="出库数量必须大于0",
                                message=f"出库失败：数量 {quantity} 无效")

    check_warehouse_access(None, current_user, wh_id)
    ensure_contact_tenant(None, current_user, stock_data.contact_id,
                          resolve_tenant_id_for_write(current_user, wh_id))

    m_scope = build_authorized_scope_predicates(_t_materials, current_user, wh_id)
    b_scope = build_authorized_scope_predicates(_t_batches, current_user, wh_id)

    with get_engine().connect() as sa_conn:
        row = sa_conn.execute(
            select(
                _t_materials.c.id, _t_materials.c.name,
                _t_materials.c.unit, _t_materials.c.safe_stock,
            ).where(
                and_(
                    or_(_t_materials.c.name == product_name, _t_materials.c.sku == product_name),
                    _t_materials.c.is_disabled == 0,
                    *m_scope,
                )
            )
        ).first()
        if row and row.name != product_name:
            resolved_from = product_name
            product_name = row.name

    if not row and stock_data.fuzzy:
        matcher = get_fuzzy_matcher()
        result = matcher.resolve(product_name, entity_type="material",
                                 tenant_id=current_user.tenant_id, warehouse_id=wh_id)
        if result['confident'] and result['best_match']:
            resolved_from = product_name
            best = result['best_match']
            extra = best.get('extra') or {}
            resolved_material_id = best.get('entity_id')
            resolved_variant = extra.get('variant')
            resolved_name = extra.get('canonical_name') or best['name']
            if resolved_variant:
                resolved_name = resolved_name.replace(f" {resolved_variant}", "").strip()
            product_name = resolved_name
            with get_engine().connect() as sa_conn:
                row_preds = [_t_materials.c.is_disabled == 0, *m_scope]
                if resolved_material_id is not None:
                    row_preds.append(_t_materials.c.id == resolved_material_id)
                else:
                    row_preds.append(_t_materials.c.name == product_name)
                row = sa_conn.execute(
                    select(
                        _t_materials.c.id, _t_materials.c.name,
                        _t_materials.c.unit, _t_materials.c.safe_stock,
                    ).where(
                        and_(*row_preds)
                    )
                ).first()
                if row:
                    product_name = row.name
        elif result['candidates']:
            names = [c['name'] for c in result['candidates'][:5]]
            return StockOutResponse(
                success=False, error="ambiguous_name",
                message=f"无法确定产品 '{product_name}'，候选：{', '.join(names)}",
                candidates=result['candidates'],
            )

    if not row:
        return StockOutResponse(success=False,
                                error=f"产品不存在: {product_name}",
                                message=f"出库失败：未找到产品 '{product_name}'")

    material_id = row.id
    unit = row.unit
    safe_stock = row.safe_stock

    effective_variant = stock_data.variant or resolved_variant
    effective_location = stock_data.location

    if stock_data.location_fuzzy and effective_location:
        loc_result = get_fuzzy_matcher().resolve_location_in_scope(
            material_id, wh_id, effective_location)
        if loc_result['confident'] and loc_result['best_match']:
            effective_location = loc_result['best_match']['name']
        elif loc_result['candidates']:
            names = [c['name'] for c in loc_result['candidates'][:5]]
            return StockOutResponse(
                success=False, error="location_ambiguous",
                message=f"库位 '{stock_data.location}' 在该产品下匹配多个：{', '.join(names)}",
                candidates=loc_result['candidates'],
            )
        else:
            with get_engine().connect() as sa_conn:
                avail_rows = sa_conn.execute(
                    select(_t_batches.c.location).where(
                        and_(
                            _t_batches.c.material_id == material_id,
                            _t_batches.c.warehouse_id == wh_id,
                            _t_batches.c.is_exhausted == 0,
                            _t_batches.c.quantity > 0,
                            _t_batches.c.location.is_not(None),
                            _t_batches.c.location != '',
                        )
                    ).distinct()
                ).all()
            avail = [r.location for r in avail_rows]
            return StockOutResponse(
                success=False, error="location_not_found",
                message=f"该产品在此仓库下没有匹配 '{stock_data.location}' 的库位。"
                        f"可用库位：{', '.join(avail) if avail else '（无）'}",
            )

    record_tenant_id = resolve_tenant_id_for_write(current_user, wh_id)
    now_dt = datetime.now()

    # ─── 分支 A：指定批次精确扣减 ───
    if stock_data.batch_no:
        with get_engine().begin() as sa_conn:
            batch = sa_conn.execute(
                select(
                    _t_batches.c.id, _t_batches.c.batch_no, _t_batches.c.quantity,
                    _t_batches.c.location, _t_batches.c.variant,
                    _t_batches.c.material_id, _t_batches.c.warehouse_id,
                ).where(and_(_t_batches.c.batch_no == stock_data.batch_no, *b_scope))
            ).first()
            if not batch or batch.material_id != material_id:
                return StockOutResponse(
                    success=False, error="batch_not_found",
                    message=f"批次 '{stock_data.batch_no}' 不存在或不属于当前产品/仓库")

            if effective_location and effective_location != (batch.location or ''):
                return StockOutResponse(
                    success=False, error="batch_field_mismatch",
                    message=f"批次 {batch.batch_no} 实际位于库位 "
                            f"'{batch.location or '（未设置）'}'，与指定的 '{effective_location}' 不符")
            if effective_variant and effective_variant != (batch.variant or ''):
                return StockOutResponse(
                    success=False, error="batch_field_mismatch",
                    message=f"批次 {batch.batch_no} 实际变体 "
                            f"'{batch.variant}'，与指定的 '{effective_variant}' 不符")

            # 单一真相源：从 active batches 聚合读取出库前库存（不再写 materials.quantity）。
            old_quantity = int(sa_conn.execute(
                select(_sa_func.coalesce(_sa_func.sum(_t_batches.c.quantity), 0))
                .where(and_(
                    _t_batches.c.material_id == material_id,
                    _t_batches.c.is_exhausted == 0,
                ))
            ).scalar() or 0)
            new_quantity = old_quantity - quantity

            if batch.quantity < quantity:
                shortfall = quantity - batch.quantity
                # 其他批次合计可用（用于 can_fallback 判断和后续 FIFO）
                # 必须带 *b_scope —— 否则在多租户/多仓场景下，
                # 预检会把其他 scope 的批次数量算进 can_fallback，让 wrapper
                # 生成"其他批次可补"的 speak_ask 骗用户答"是"，然后真扣时
                # FIFO 的 *b_scope 又把那些批次拒之门外，导致 409 回滚。
                other_avail = int(sa_conn.execute(
                    select(_sa_func.coalesce(_sa_func.sum(_t_batches.c.quantity), 0))
                    .where(and_(
                        _t_batches.c.material_id == material_id,
                        _t_batches.c.id != batch.id,
                        _t_batches.c.is_exhausted == 0,
                        _t_batches.c.quantity > 0,
                        *b_scope,
                    ))
                ).scalar() or 0)
                can_fallback = other_avail >= shortfall

                if not stock_data.allow_partial_fallback:
                    # 未授权 fallback：返回结构化失败，含 can_fallback 让 MCP 询问用户
                    return StockOutResponse(
                        success=False, error="batch_insufficient_stock",
                        batch_no_requested=batch.batch_no,
                        batch_available=batch.quantity,
                        shortfall=shortfall,
                        can_fallback=can_fallback,
                        fallback_total_available=other_avail,
                        message=(
                            f"批次 {batch.batch_no} 余量 {batch.quantity} {unit}，"
                            f"不足以出库 {quantity} {unit}，"
                            f"还差 {shortfall} {unit}。"
                            + (f"其他批次合计 {other_avail} {unit} 可补差额，是否确认？"
                               if can_fallback else
                               f"其他批次合计仅 {other_avail} {unit}，也不足补差额。")
                        ),
                    )

                # 已授权 fallback：同事务内"先扣指定批次全部 + FIFO 补差额"。
                if not can_fallback:
                    return StockOutResponse(
                        success=False, error="batch_insufficient_stock",
                        batch_no_requested=batch.batch_no,
                        batch_available=batch.quantity,
                        shortfall=shortfall,
                        can_fallback=False,
                        fallback_total_available=other_avail,
                        message=f"无法补足：指定批次 {batch.batch_no} 仅 {batch.quantity} {unit}，"
                                f"其他批次合计 {other_avail} {unit}，仍缺 {shortfall - other_avail} {unit}。",
                    )

                # 进入"先扣完指定批次 + FIFO 补差额"事务路径
                ins_rec = sa_conn.execute(
                    insert(_t_inventory_records).values(
                        material_id=material_id, type=RecordType.OUT.value, quantity=quantity,
                        operator=operator, operator_user_id=operator_user_id,
                        reason_category=reason_category, reason_note=reason_note,
                        contact_id=stock_data.contact_id, warehouse_id=wh_id,
                        tenant_id=record_tenant_id, created_at=now_dt,
                    )
                )
                record_id = ins_rec.inserted_primary_key[0]

                batch_consumptions = []

                # ① 先把指定批次的余量全扣完
                consumed_from_specified = batch.quantity
                spec_upd = sa_conn.execute(
                    update(_t_batches)
                    .where(and_(
                        _t_batches.c.id == batch.id,
                        _t_batches.c.quantity >= consumed_from_specified,
                        _t_batches.c.is_exhausted == 0,
                    ))
                    .values(quantity=0, is_exhausted=1)
                )
                if spec_upd.rowcount != 1:
                    raise HTTPException(status_code=409, detail="批次并发冲突，请重试")
                sa_conn.execute(
                    insert(_t_batch_consumptions).values(
                        record_id=record_id, batch_id=batch.id,
                        quantity=consumed_from_specified, created_at=now_dt,
                    )
                )
                batch_consumptions.append(BatchConsumption(
                    batch_no=batch.batch_no, batch_id=batch.id,
                    quantity=consumed_from_specified, remaining=0, variant=batch.variant,
                ))

                # ② FIFO 在其他批次（排除指定批次）中扣 shortfall
                remaining_to_consume = shortfall
                fifo_stmt = (
                    select(
                        _t_batches.c.id, _t_batches.c.batch_no, _t_batches.c.quantity,
                        _t_batches.c.variant,
                    )
                    .where(and_(
                        _t_batches.c.material_id == material_id,
                        _t_batches.c.id != batch.id,
                        _t_batches.c.is_exhausted == 0,
                        _t_batches.c.quantity > 0,
                        *b_scope,
                    ))
                    .order_by(_t_batches.c.created_at.asc())
                    .with_for_update()
                )
                for fb in sa_conn.execute(fifo_stmt).all():
                    if remaining_to_consume <= 0:
                        break
                    take = min(fb.quantity, remaining_to_consume)
                    upd = sa_conn.execute(
                        update(_t_batches)
                        .where(and_(
                            _t_batches.c.id == fb.id,
                            _t_batches.c.quantity >= take,
                            _t_batches.c.is_exhausted == 0,
                        ))
                        .values(
                            quantity=_t_batches.c.quantity - take,
                            is_exhausted=case(
                                (_t_batches.c.quantity - take <= 0, 1),
                                else_=0,
                            ),
                        )
                    )
                    if upd.rowcount != 1:
                        raise HTTPException(status_code=409, detail="批次并发冲突，请重试")
                    sa_conn.execute(
                        insert(_t_batch_consumptions).values(
                            record_id=record_id, batch_id=fb.id,
                            quantity=take, created_at=now_dt,
                        )
                    )
                    batch_consumptions.append(BatchConsumption(
                        batch_no=fb.batch_no, batch_id=fb.id,
                        quantity=take, remaining=max(fb.quantity - take, 0),
                        variant=fb.variant,
                    ))
                    remaining_to_consume -= take

                if remaining_to_consume > 0:
                    # 事务内 raise，前面所有扣减回滚
                    raise HTTPException(
                        status_code=409,
                        detail=f"出库失败：可用批次不足，仍缺 {remaining_to_consume} {unit}",
                    )

                # 出库不改 materials.name / batches.variant，fuzzy material 索引
                # 不需要失效（codex 复审 a6a98bcad2d766c5b 已确认，索引也不按
                # is_exhausted 过滤所以 exhausted 也不影响命中）。

                audit_log("STOCK_OUT", current_user.id, current_user.username, {
                    "product": product_name, "quantity": quantity,
                    "old_qty": old_quantity, "new_qty": new_quantity,
                    "resolved_from": resolved_from,
                    "specified_batch": batch.batch_no,
                    "partial_fallback": True,
                    "batches": [bc.batch_no for bc in batch_consumptions],
                })

                warning = ""
                if safe_stock is not None and new_quantity < safe_stock:
                    if new_quantity < safe_stock * 0.5:
                        warning = f"⚠️ 警告：库存告急！当前库存 {new_quantity} {unit}，低于安全库存 {safe_stock} {unit} 的50%"
                    else:
                        warning = f"⚠️ 提醒：库存偏低，当前库存 {new_quantity} {unit}，低于安全库存 {safe_stock} {unit}"

                details = "、".join(f"{bc.batch_no}×{bc.quantity}" for bc in batch_consumptions)
                return StockOutResponse(
                    success=True, operation="stock_out",
                    product=StockOperationProduct(
                        name=product_name, old_quantity=old_quantity,
                        out_quantity=quantity, new_quantity=new_quantity,
                        unit=unit, safe_stock=safe_stock,
                    ),
                    batch_consumptions=batch_consumptions,
                    message=f"出库成功（指定批次 {batch.batch_no} 不足，已从其他批次 FIFO 补足）："
                            f"{product_name} 共出 {quantity} {unit}（{details}），"
                            f"库存 {old_quantity}→{new_quantity} {unit}",
                    warning=warning if warning else None,
                    resolved_from=resolved_from,
                )

            ins_rec = sa_conn.execute(
                insert(_t_inventory_records).values(
                    material_id=material_id, type=RecordType.OUT.value, quantity=quantity,
                    operator=operator, operator_user_id=operator_user_id,
                    reason_category=reason_category, reason_note=reason_note,
                    contact_id=stock_data.contact_id, warehouse_id=wh_id,
                    tenant_id=record_tenant_id, created_at=now_dt,
                )
            )
            record_id = ins_rec.inserted_primary_key[0]

            consume_qty = quantity
            batch_upd = sa_conn.execute(
                update(_t_batches)
                    .where(and_(
                        _t_batches.c.id == batch.id,
                        _t_batches.c.quantity >= consume_qty,
                        _t_batches.c.is_exhausted == 0,
                    ))
                    .values(
                        quantity=_t_batches.c.quantity - consume_qty,
                        is_exhausted=case(
                            (_t_batches.c.quantity - consume_qty <= 0, 1),
                            else_=0,
                        ),
                    )
            )
            if batch_upd.rowcount != 1:
                raise HTTPException(status_code=409, detail="批次并发冲突，请重试")

            sa_conn.execute(
                insert(_t_batch_consumptions).values(
                    record_id=record_id, batch_id=batch.id,
                    quantity=consume_qty, created_at=now_dt,
                )
            )

            remaining_qty = max(batch.quantity - consume_qty, 0)
            batch_consumptions = [BatchConsumption(
                batch_no=batch.batch_no, batch_id=batch.id,
                quantity=consume_qty, remaining=remaining_qty, variant=batch.variant,
            )]

        # 出库不改索引（同上）。

        audit_log("STOCK_OUT", current_user.id, current_user.username, {
            "product": product_name, "quantity": quantity,
            "old_qty": old_quantity, "new_qty": new_quantity,
            "resolved_from": resolved_from,
            "specified_batch": batch.batch_no,
        })

        warning = ""
        if safe_stock is not None and new_quantity < safe_stock:
            if new_quantity < safe_stock * 0.5:
                warning = f"⚠️ 警告：库存告急！当前库存 {new_quantity} {unit}，低于安全库存 {safe_stock} {unit} 的50%"
            else:
                warning = f"⚠️ 提醒：库存偏低，当前库存 {new_quantity} {unit}，低于安全库存 {safe_stock} {unit}"

        return StockOutResponse(
            success=True, operation="stock_out",
            product=StockOperationProduct(
                name=product_name, old_quantity=old_quantity,
                out_quantity=quantity, new_quantity=new_quantity,
                unit=unit, safe_stock=safe_stock,
            ),
            batch_consumptions=batch_consumptions,
            message=f"出库成功：{product_name} 从指定批次 {batch.batch_no} "
                    f"出库 {quantity} {unit}，库存 {old_quantity}→{new_quantity} {unit}",
            warning=warning if warning else None,
            resolved_from=resolved_from,
        )

    # ─── 分支 B：FIFO（支持 location / variant 过滤） ───
    # 单一真相源：无条件前置 active batches 聚合校验，库存不足直接 409，不进事务。
    precheck_preds = [
        _t_batches.c.material_id == material_id,
        _t_batches.c.is_exhausted == 0,
        _t_batches.c.quantity > 0,
        *b_scope,
    ]
    if effective_variant:
        precheck_preds.append(_t_batches.c.variant == effective_variant)
    if effective_location:
        precheck_preds.append(_t_batches.c.location == effective_location)
    with get_engine().connect() as sa_conn:
        avail_qty = int(sa_conn.execute(
            select(_sa_func.coalesce(_sa_func.sum(_t_batches.c.quantity), 0))
                .where(and_(*precheck_preds))
        ).scalar() or 0)
    if avail_qty < quantity:
        scope = []
        if effective_location:
            scope.append(f"位置 '{effective_location}'")
        if effective_variant:
            scope.append(f"变体 '{effective_variant}'")
        scope_msg = f"在 {'、'.join(scope)}" if scope else ""
        return StockOutResponse(
            success=False, error="库存不足",
            message=f"出库失败：{product_name}{scope_msg} "
                    f"的可用库存为 {avail_qty} {unit}，需要出库 {quantity} {unit}")

    with get_engine().begin() as sa_conn:
        # old_quantity 取 material 总库存（不带 location/variant 过滤），与历史响应语义保持一致
        old_quantity = int(sa_conn.execute(
            select(_sa_func.coalesce(_sa_func.sum(_t_batches.c.quantity), 0))
            .where(and_(
                _t_batches.c.material_id == material_id,
                _t_batches.c.is_exhausted == 0,
            ))
        ).scalar() or 0)
        new_quantity = old_quantity - quantity

        ins_rec = sa_conn.execute(
            insert(_t_inventory_records).values(
                material_id=material_id, type=RecordType.OUT.value, quantity=quantity,
                operator=operator, operator_user_id=operator_user_id,
                reason_category=reason_category, reason_note=reason_note,
                contact_id=stock_data.contact_id, warehouse_id=wh_id,
                tenant_id=record_tenant_id, created_at=now_dt,
            )
        )
        record_id = ins_rec.inserted_primary_key[0]

        batch_consumptions = []
        remaining_to_consume = quantity
        fifo_preds = [
            _t_batches.c.material_id == material_id,
            _t_batches.c.is_exhausted == 0,
            _t_batches.c.quantity > 0,
            *b_scope,
        ]
        if effective_variant:
            fifo_preds.append(_t_batches.c.variant == effective_variant)
        if effective_location:
            fifo_preds.append(_t_batches.c.location == effective_location)
        fifo_stmt = (
            select(
                _t_batches.c.id, _t_batches.c.batch_no, _t_batches.c.quantity,
                _t_batches.c.variant, _t_batches.c.location,
            )
            .where(and_(*fifo_preds))
            .order_by(_t_batches.c.created_at.asc())
            .with_for_update()
        )
        fifo_rows = sa_conn.execute(fifo_stmt).all()

        for b in fifo_rows:
            if remaining_to_consume <= 0:
                break
            consume_qty = min(b.quantity, remaining_to_consume)
            remaining_to_consume -= consume_qty
            batch_upd = sa_conn.execute(
                update(_t_batches)
                    .where(and_(
                        _t_batches.c.id == b.id,
                        _t_batches.c.quantity >= consume_qty,
                        _t_batches.c.is_exhausted == 0,
                    ))
                    .values(
                        quantity=_t_batches.c.quantity - consume_qty,
                        is_exhausted=case(
                            (_t_batches.c.quantity - consume_qty <= 0, 1),
                            else_=0,
                        ),
                    )
            )
            if batch_upd.rowcount != 1:
                raise HTTPException(status_code=409, detail="批次并发冲突，请重试")
            sa_conn.execute(
                insert(_t_batch_consumptions).values(
                    record_id=record_id, batch_id=b.id,
                    quantity=consume_qty, created_at=now_dt,
                )
            )
            remaining_qty = max(b.quantity - consume_qty, 0)
            batch_consumptions.append(BatchConsumption(
                batch_no=b.batch_no, batch_id=b.id,
                quantity=consume_qty, remaining=remaining_qty, variant=b.variant,
            ))

        if remaining_to_consume > 0:
            # FIFO did not cover the requested quantity (e.g. batch table
            # under-counts vs materials.quantity, or scope filtered too
            # aggressively). Raise inside the txn so the materials decrement
            # and inventory_records insert above are rolled back.
            raise HTTPException(
                status_code=409,
                detail=f"出库失败：{product_name} 可用批次不足，仍缺 {remaining_to_consume} {unit}，请检查批次/库位/变体筛选条件",
            )

    # 出库不改索引（同上）。

    audit_log("STOCK_OUT", current_user.id, current_user.username, {
        "product": product_name, "quantity": quantity,
        "old_qty": old_quantity, "new_qty": new_quantity,
        "resolved_from": resolved_from,
        "batches": [bc.batch_no for bc in batch_consumptions],
    })

    warning = ""
    if safe_stock is not None and new_quantity < safe_stock:
        if new_quantity < safe_stock * 0.5:
            warning = f"⚠️ 警告：库存告急！当前库存 {new_quantity} {unit}，低于安全库存 {safe_stock} {unit} 的50%"
        else:
            warning = f"⚠️ 提醒：库存偏低，当前库存 {new_quantity} {unit}，低于安全库存 {safe_stock} {unit}"

    batch_details = ""
    if batch_consumptions:
        details = [f"{bc.batch_no}×{bc.quantity}" for bc in batch_consumptions]
        batch_details = f"（消耗批次: {', '.join(details)}）"

    return StockOutResponse(
        success=True, operation="stock_out",
        product=StockOperationProduct(
            name=product_name, old_quantity=old_quantity,
            out_quantity=quantity, new_quantity=new_quantity,
            unit=unit, safe_stock=safe_stock,
        ),
        batch_consumptions=batch_consumptions if batch_consumptions else None,
        message=f"出库成功：{product_name} 出库 {quantity} {unit}{batch_details}，"
                f"库存从 {old_quantity} 更新到 {new_quantity} {unit}",
        warning=warning if warning else None,
        resolved_from=resolved_from,
    )


@app.post("/api/materials/batches/move-location", response_model=BatchMoveResponse)
async def move_batch_location(
    request: BatchMoveRequest,
    current_user: CurrentUser = Depends(require_permission(Resource.INVENTORY, Action.WRITE))
):
    """批次库位移动。

    - quantity 留空或等于批次余量 → 整批移位（仅更新 location）
    - quantity 小于批次余量 → 拆分：源批次扣减 quantity，并在目标库位创建同物料/变体的新批次
    - 不改变物料总库存，仅改变批次的 location/数量分布
    """
    wh_id = require_warehouse_id(current_user, request.warehouse_id)
    check_warehouse_access(None, current_user, wh_id)

    new_location = (request.new_location or '').strip()
    if not new_location:
        return BatchMoveResponse(
            success=False, error="empty_location",
            message="移位失败：目标库位不能为空"
        )

    batch_no_norm = (request.batch_no or '').strip()
    if not batch_no_norm:
        return BatchMoveResponse(
            success=False, error="missing_batch_no",
            message="移位失败：必须指定批次号"
        )

    operator = request.operator if request.operator and request.operator != "MCP系统" else current_user.get_operator_name()
    record_tenant_id = resolve_tenant_id_for_write(current_user, wh_id)

    b_scope = list(build_authorized_scope_predicates(_t_batches, current_user, wh_id))
    operator_user_id = current_user.id

    with get_engine().begin() as sa_conn:
        # 先用 FOR UPDATE 锁住源批次行（参考 stock_out FIFO 路径 app.py:5226），
        # 防止并发拆分移位导致的 lost update / 总库存膨胀。
        # 为了能拿到 material name/unit，单独查 materials（无需锁）。
        locked = sa_conn.execute(
            select(
                _t_batches.c.id, _t_batches.c.material_id, _t_batches.c.quantity,
                _t_batches.c.location, _t_batches.c.variant, _t_batches.c.contact_id,
                _t_batches.c.batch_no,
            ).where(and_(
                _t_batches.c.batch_no == batch_no_norm,
                _t_batches.c.is_exhausted == 0,
                *b_scope,
            )).with_for_update()
        ).first()

        if not locked:
            return BatchMoveResponse(
                success=False, error="batch_not_found",
                message=f"移位失败：批次 '{batch_no_norm}' 不存在或已耗尽"
            )

        mat = sa_conn.execute(
            select(_t_materials.c.name, _t_materials.c.unit)
            .where(_t_materials.c.id == locked.material_id)
        ).first()
        material_name = mat.name if mat else ''
        unit = mat.unit if mat else ''

        # 用 namespace dict 方便下面统一引用
        class _B:
            pass
        batch = _B()
        batch.id = locked.id
        batch.material_id = locked.material_id
        batch.quantity = locked.quantity
        batch.location = locked.location
        batch.variant = locked.variant
        batch.contact_id = locked.contact_id
        batch.batch_no = locked.batch_no
        batch.material_name = material_name
        batch.unit = unit

        if request.product_name and request.product_name.strip() and request.product_name.strip() != batch.material_name:
            return BatchMoveResponse(
                success=False, error="product_mismatch",
                message=(f"移位失败：批次 '{batch_no_norm}' 属于 '{batch.material_name}'，"
                         f"与指定的 '{request.product_name}' 不符")
            )

        cur_loc = (batch.location or '').strip()
        if request.from_location is not None:
            req_from = request.from_location.strip()
            if cur_loc != req_from:
                return BatchMoveResponse(
                    success=False, error="from_location_mismatch",
                    message=(f"移位失败：批次 '{batch_no_norm}' 当前位置为 "
                             f"'{cur_loc or '（未设置）'}'，与指定的 '{req_from}' 不符")
                )

        if new_location == cur_loc:
            return BatchMoveResponse(
                success=False, error="same_location",
                message=f"移位失败：批次 '{batch_no_norm}' 已经位于 '{new_location}'"
            )

        move_qty = request.quantity if request.quantity is not None else batch.quantity
        if move_qty <= 0:
            return BatchMoveResponse(
                success=False, error="invalid_quantity",
                message=f"移位失败：移位数量 {move_qty} 无效（必须大于 0）"
            )
        if move_qty > batch.quantity:
            return BatchMoveResponse(
                success=False, error="insufficient_quantity",
                message=(f"移位失败：批次 '{batch_no_norm}' 当前余量 {batch.quantity} "
                         f"{batch.unit}，无法移位 {move_qty} {batch.unit}"),
            )

        now_dt = datetime.now()
        full_move = (move_qty == batch.quantity)

        if full_move:
            upd = sa_conn.execute(
                _t_batches.update()
                .where(and_(
                    _t_batches.c.id == batch.id,
                    _t_batches.c.quantity == batch.quantity,
                    _t_batches.c.is_exhausted == 0,
                ))
                .values(location=new_location)
            )
            if upd.rowcount != 1:
                raise HTTPException(status_code=409, detail="批次并发冲突，请重试")
            target_batch_no = batch.batch_no
            target_batch_id = batch.id
            source_remaining = 0  # 整批迁移：源库位已无该批次
        else:
            # 防御性：拆分扣减用条件 UPDATE，rowcount=0 表示别的事务先动过
            source_remaining = batch.quantity - move_qty
            upd = sa_conn.execute(
                _t_batches.update()
                .where(and_(
                    _t_batches.c.id == batch.id,
                    _t_batches.c.quantity == batch.quantity,
                    _t_batches.c.is_exhausted == 0,
                ))
                .values(quantity=source_remaining)
            )
            if upd.rowcount != 1:
                raise HTTPException(status_code=409, detail="批次并发冲突，请重试")

            # generate_batch_no 走独立连接看不到未提交行，拆分 INSERT 撞 unique
            # (batches.batch_no UNIQUE) 时重试，最多 5 次。
            target_batch_id = None
            target_batch_no = None
            for _attempt in range(5):
                candidate_no = generate_batch_no(batch.material_id, warehouse_id=wh_id)
                try:
                    ins = sa_conn.execute(
                        insert(_t_batches).values(
                            batch_no=candidate_no,
                            material_id=batch.material_id,
                            quantity=move_qty,
                            initial_quantity=move_qty,
                            contact_id=batch.contact_id,
                            location=new_location,
                            variant=batch.variant,
                            warehouse_id=wh_id,
                            tenant_id=record_tenant_id,
                            created_at=now_dt,
                        )
                    )
                    target_batch_id = ins.inserted_primary_key[0]
                    target_batch_no = candidate_no
                    break
                except IntegrityError:
                    continue
            if target_batch_no is None:
                raise HTTPException(status_code=409, detail="批次号生成冲突，请重试")

        # 库存台账：仅在**拆分**移位时写 inventory_records，记录"源批次扣减 + 新批次入账"
        # 便于按 batch_id 重放溯源。整批移位时只是同一批次换 location，
        # audit_log 已足够，不再额外往 inventory_records 写双条（避免污染当日 in/out 报表）。
        # reason_category 复用 transfer_out/transfer_in（已存在于 REASON_CATEGORIES），
        # reason_note 用 BATCH_RELOCATE 前缀标识本次是库内库位移动而非跨仓调拨。
        if not full_move:
            relocate_note = f"BATCH_RELOCATE: {cur_loc or '（未设置）'} -> {new_location}"
            sa_conn.execute(
                insert(_t_inventory_records).values(
                    material_id=batch.material_id, type=RecordType.OUT.value,
                    quantity=move_qty, operator=operator, operator_user_id=operator_user_id,
                    reason_category='transfer_out', reason_note=relocate_note,
                    contact_id=batch.contact_id, batch_id=batch.id,
                    warehouse_id=wh_id, tenant_id=record_tenant_id, created_at=now_dt,
                )
            )
            sa_conn.execute(
                insert(_t_inventory_records).values(
                    material_id=batch.material_id, type=RecordType.IN.value,
                    quantity=move_qty, operator=operator, operator_user_id=operator_user_id,
                    reason_category='transfer_in', reason_note=relocate_note,
                    contact_id=batch.contact_id, batch_id=target_batch_id,
                    warehouse_id=wh_id, tenant_id=record_tenant_id, created_at=now_dt,
                )
            )

    # 移位不动 materials 表（既不改 name 也不改 variant），fuzzy matcher 的
    # material 名字索引不需要失效。location 是 batch 字段，根本不在 material
    # 索引里，旧代码这里调 invalidate_cache(entity_type="material") 是纯浪费
    # （codex 复审已指出）。

    audit_log("BATCH_MOVE_LOCATION", current_user.id, current_user.username, {
        "batch_no": batch_no_norm,
        "material": batch.material_name,
        "from_location": cur_loc,
        "to_location": new_location,
        "moved_quantity": move_qty,
        "full_move": full_move,
        "target_batch_no": target_batch_no,
        "source_batch_id": batch.id,
        "target_batch_id": target_batch_id,
        "operator": operator,
    })

    kind = "整批移位" if full_move else "拆分移位"
    extra = "" if full_move else f"（新批次 {target_batch_no}，原批次余 {source_remaining} {batch.unit}）"
    return BatchMoveResponse(
        success=True,
        operation="move_batch_location",
        moved_quantity=move_qty,
        from_location=cur_loc,
        to_location=new_location,
        full_move=full_move,
        source_batch=BatchInfo(
            batch_no=batch.batch_no, batch_id=batch.id,
            quantity=source_remaining, variant=batch.variant,
        ),
        target_batch=BatchInfo(
            batch_no=target_batch_no, batch_id=target_batch_id,
            quantity=move_qty, variant=batch.variant,
        ),
        message=(f"{kind}成功：{batch.material_name} 批次 {batch_no_norm} "
                 f"从 '{cur_loc or '（未设置）'}' 移动 {move_qty} {batch.unit} "
                 f"到 '{new_location}'{extra}"),
    )


# ============ Excel Import/Export APIs ============

@app.get("/api/materials/export-excel")
def export_materials_excel(
    name: Optional[str] = Query(None, description="名称/SKU模糊搜索"),
    category: Optional[str] = Query(None, description="分类"),
    status: Optional[str] = Query(None, description="状态(逗号分隔)"),
    warehouse_id: Optional[int] = Query(None, description="仓库ID"),
    current_user: CurrentUser = Depends(require_permission(Resource.MATERIALS, Action.READ))
):
    """导出库存数据为Excel — 一行一批次，含批次号、位置、联系方 — Phase 3f: SA Core read."""
    wh_id = resolve_warehouse_id(current_user, warehouse_id)

    # 解析状态筛选
    status_filter = status.split(',') if status else None

    preds = list(build_authorized_scope_predicates(_t_materials, current_user, wh_id))
    if not status_filter or 'disabled' not in status_filter:
        preds.append(_t_materials.c.is_disabled == 0)
    if name:
        like = f'%{name}%'
        preds.append(or_(_t_materials.c.name.like(like), _t_materials.c.sku.like(like)))
    if category:
        preds.append(_t_materials.c.category == category)

    mat_stmt = select(
        _t_materials.c.id, _t_materials.c.name, _t_materials.c.sku,
        _t_materials.c.category,
        _sa_func.coalesce(
            select(_sa_func.sum(_t_batches.c.quantity))
            .where(and_(
                _t_batches.c.material_id == _t_materials.c.id,
                _t_batches.c.is_exhausted == 0,
            ))
            .scalar_subquery(),
            0,
        ).label('quantity'),
        _t_materials.c.unit,
        _t_materials.c.safe_stock, _t_materials.c.location, _t_materials.c.is_disabled,
    ).order_by(_t_materials.c.name.asc())
    if preds:
        mat_stmt = mat_stmt.where(and_(*preds))

    with get_engine().connect() as sa_conn:
        rows = sa_conn.execute(mat_stmt).fetchall()

        # 构建导出行（一行一批次）
        export_rows = []
        for row in rows:
            quantity = int(row.quantity or 0)
            safe_stock = row.safe_stock
            is_disabled = bool(row.is_disabled)

            # 计算状态
            if is_disabled:
                item_status = 'disabled'
            elif safe_stock is not None:
                if quantity >= safe_stock:
                    item_status = 'normal'
                elif quantity >= safe_stock * 0.5:
                    item_status = 'warning'
                else:
                    item_status = 'danger'
            else:
                item_status = 'normal'

            # 状态筛选
            if status_filter and item_status not in status_filter:
                continue

            material_base = {
                'name': row.name,
                'sku': row.sku,
                'category': row.category,
                'unit': row.unit,
                'safe_stock': row.safe_stock,
            }

            # 查询活跃批次
            batch_stmt = (
                select(
                    _t_batches.c.batch_no, _t_batches.c.quantity,
                    _t_batches.c.location, _t_batches.c.variant,
                    _t_contacts.c.name.label('contact_name'),
                )
                .select_from(
                    _t_batches.outerjoin(_t_contacts, _t_batches.c.contact_id == _t_contacts.c.id)
                )
                .where(and_(
                    _t_batches.c.material_id == row.id,
                    _t_batches.c.is_exhausted == 0,
                ))
                .order_by(_t_batches.c.created_at.asc())
            )
            batches = sa_conn.execute(batch_stmt).fetchall()

            if batches:
                for batch in batches:
                    export_rows.append({
                        **material_base,
                        'batch_no': batch.batch_no,
                        'quantity': batch.quantity,
                        'location': batch.location or '',
                        'contact_name': batch.contact_name or '',
                        'variant': batch.variant or '',
                    })
            else:
                # 无活跃批次（库存为0的物料）
                export_rows.append({
                    **material_base,
                    'batch_no': '',
                    'quantity': 0,
                    'location': row.location or '',
                    'contact_name': '',
                    'variant': '',
                })

    wb = Workbook()
    ws = wb.active
    ws.title = "库存数据"

    # 表头
    headers = ['物料名称', '规格', '物料编码(SKU)', '分类', '单位', '安全库存', '批次号', '库存', '存放位置', '联系方']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 数据
    for row_idx, item in enumerate(export_rows, 2):
        ws.cell(row=row_idx, column=1, value=item['name'])
        ws.cell(row=row_idx, column=2, value=item['variant'])
        ws.cell(row=row_idx, column=3, value=item['sku'])
        ws.cell(row=row_idx, column=4, value=item['category'])
        ws.cell(row=row_idx, column=5, value=item['unit'])
        ws.cell(row=row_idx, column=6, value=item['safe_stock'])
        ws.cell(row=row_idx, column=7, value=item['batch_no'])
        ws.cell(row=row_idx, column=8, value=item['quantity'])
        ws.cell(row=row_idx, column=9, value=item['location'])
        ws.cell(row=row_idx, column=10, value=item['contact_name'])

    # 设置列宽
    column_widths = [22, 10, 18, 14, 8, 12, 18, 10, 16, 16]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def extract_variants(names: list) -> tuple:
    """从同SKU的多个名称中提取公共前缀和变体。

    返回 (common_name, [variant_or_None, ...])。
    如果所有名称相同，variant 为 None。
    """
    unique_names = set(names)
    if len(unique_names) <= 1:
        return names[0], [None] * len(names)

    prefix = os.path.commonprefix(list(unique_names))
    min_len = min(len(n) for n in unique_names)

    # 公共前缀太短（<30%），不做拆分
    if len(prefix) < min_len * 0.3:
        return names[0], [n if n != names[0] else None for n in names]

    common_name = prefix.rstrip()
    variants = [n[len(prefix):].strip() or None for n in names]
    return common_name, variants


@app.post("/api/materials/import-excel/preview", response_model=ExcelImportPreviewResponse)
@limiter.limit("10/minute")  # Excel导入速率限制
async def preview_import_excel(
    request: Request,
    file: UploadFile = File(...),
    warehouse_id: Optional[int] = Query(None, description="仓库ID"),
    current_user: CurrentUser = Depends(require_permission(Resource.MATERIALS, Action.WRITE))
):
    """预览Excel导入内容，自动检测简化模式/批次模式"""
    wh_id = resolve_warehouse_id(current_user, warehouse_id)
    def _error_resp(msg):
        return ExcelImportPreviewResponse(
            success=False, preview=[], new_skus=[], total_in=0, total_out=0, total_new=0, message=msg
        )

    # 文件大小检查
    contents = await file.read()
    file_size_mb = len(contents) / (1024 * 1024)
    if file_size_mb > MAX_UPLOAD_SIZE_MB:
        return _error_resp(f"文件大小 ({file_size_mb:.1f}MB) 超过限制 ({MAX_UPLOAD_SIZE_MB}MB)")

    try:
        wb = load_workbook(filename=BytesIO(contents))
        ws = wb.active
    except Exception as e:
        return _error_resp(f"文件解析失败: {str(e)}")

    # 读取表头，自动识别列位置
    header_row = [str(cell).strip() if cell else "" for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    header_set = set(header_row)
    if {'记录类型', '操作人', '时间'}.issubset(header_set):
        return _error_resp("Excel格式错误：这是出入库记录导出文件，不能作为库存导入模板。请从“库存列表”导出 inventory_*.xlsx 后再导入。")

    col_mapping = {
        'name': None, 'sku': None, 'category': None, 'quantity': None,
        'unit': None, 'safe_stock': None, 'location': None,
        'batch_no': None, 'contact_name': None, 'variant': None,
    }

    for idx, header in enumerate(header_row):
        header_lower = header.lower()
        if '名称' in header or 'name' in header_lower:
            col_mapping['name'] = idx
        elif 'sku' in header_lower or '编码' in header:
            col_mapping['sku'] = idx
        elif '分类' in header or 'category' in header_lower:
            col_mapping['category'] = idx
        elif '库存' in header or 'quantity' in header_lower or '数量' in header:
            if '安全' not in header and '批次' not in header:
                col_mapping['quantity'] = idx
        elif '单位' in header or 'unit' in header_lower:
            col_mapping['unit'] = idx
        elif '安全库存' in header or 'safe' in header_lower:
            col_mapping['safe_stock'] = idx
        elif '位置' in header or 'location' in header_lower:
            col_mapping['location'] = idx
        elif '批次' in header or 'batch' in header_lower:
            col_mapping['batch_no'] = idx
        elif '联系方' in header or 'contact' in header_lower or '供应商' in header:
            col_mapping['contact_name'] = idx
        elif '变体' in header or '规格' in header or 'variant' in header_lower:
            col_mapping['variant'] = idx

    if col_mapping['sku'] is None:
        return _error_resp("Excel格式错误：找不到SKU/物料编码列")
    if col_mapping['quantity'] is None:
        return _error_resp("Excel格式错误：找不到库存/数量列")

    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    is_batch_mode = False
    if col_mapping['batch_no'] is not None:
        batch_idx = col_mapping['batch_no']
        has_batch_values = any(
            batch_idx < len(row) and row[batch_idx] not in (None, '')
            for row in data_rows
        )
        # 旧版库存导出模板包含空的“批次号”和“变体”列，语义是整库快照；
        # 普通批次模板只要有“批次号”列，即使单行为空，也表示导入为新批次。
        is_legacy_empty_batch_snapshot = not has_batch_values and col_mapping['variant'] is not None
        is_batch_mode = has_batch_values or not is_legacy_empty_batch_snapshot

    def _read_cell(row, key):
        ci = col_mapping[key]
        if ci is None or ci >= len(row) or row[ci] is None:
            return None
        return str(row[ci]).strip()

    def _read_int(row, key, default=0):
        ci = col_mapping[key]
        if ci is None or ci >= len(row) or row[ci] is None:
            return default
        return int(row[ci])

    preview_items = []
    new_skus = []
    new_contacts_set = set()
    seen_skus_simple = set()  # 简化模式下追踪已见SKU，检测同SKU多行
    sku_excel_names = {}  # SKU → [(preview_item_index, excel_name), ...] 用于后处理提取 variant
    has_variant_col = col_mapping['variant'] is not None
    total_in = 0
    total_out = 0
    total_new = 0
    row_count = 0
    duplicate_rows = 0
    seen_import_rows = set()

    with get_engine().connect() as sa_conn:

        # 联系方为租户级（不绑定仓库），用 tenant 单独构造 scope
        contact_tenant_id = resolve_tenant_id_for_write(current_user, wh_id) if wh_id is not None else current_user.tenant_id
        contact_preds_base = list(build_scope_predicates(_t_contacts, contact_tenant_id, None))
        mat_scope_preds = list(build_authorized_scope_predicates(_t_materials, current_user, wh_id))

        # 联系方解析辅助
        def resolve_contact(name):
            if not name:
                return None, None
            stmt = select(_t_contacts.c.id).where(and_(
                _t_contacts.c.name == name,
                _t_contacts.c.is_disabled == 0,
                *contact_preds_base,
            ))
            r = sa_conn.execute(stmt).first()
            if r:
                return r.id, name
            new_contacts_set.add(name)
            return None, name

        for idx, row in enumerate(data_rows, start=2):
            if not row[col_mapping['sku']]:
                continue

            row_count += 1
            if row_count > MAX_IMPORT_ROWS:
                return _error_resp(f"数据行数 ({row_count}) 超过限制 ({MAX_IMPORT_ROWS}行)")

            name = _read_cell(row, 'name') or ""
            sku = _read_cell(row, 'sku')
            category = _read_cell(row, 'category') or "未分类"
            unit = _read_cell(row, 'unit') or "个"
            location = _read_cell(row, 'location') or ""
            batch_no_val = _read_cell(row, 'batch_no') or ""
            contact_name_val = _read_cell(row, 'contact_name') or ""

            try:
                import_qty = _read_int(row, 'quantity', 0)
            except (ValueError, TypeError):
                return _error_resp(f"第 {idx} 行【库存数量】格式错误：需要整数，当前值为 '{row[col_mapping['quantity']]}'")

            try:
                safe_stock = _read_int(row, 'safe_stock', None)
            except (ValueError, TypeError):
                return _error_resp(f"第 {idx} 行【安全库存】格式错误：需要整数，当前值为 '{row[col_mapping['safe_stock']]}'")

            variant_val = _read_cell(row, 'variant') or ""
            contact_id, contact_name = resolve_contact(contact_name_val) if contact_name_val else (None, None)
            row_key = (
                sku, name, category, unit, safe_stock, location,
                batch_no_val, variant_val, contact_name_val, import_qty
            )
            if row_key in seen_import_rows:
                duplicate_rows += 1
                continue
            seen_import_rows.add(row_key)

            # 查询物料（按仓库过滤）— 单一真相源：quantity 来自 active batches sum
            material = sa_conn.execute(
                select(
                    _t_materials.c.id,
                    _t_materials.c.name,
                    _sa_func.coalesce(
                        select(_sa_func.sum(_t_batches.c.quantity))
                        .where(and_(
                            _t_batches.c.material_id == _t_materials.c.id,
                            _t_batches.c.is_exhausted == 0,
                        ))
                        .scalar_subquery(),
                        0,
                    ).label('quantity'),
                )
                .where(and_(_t_materials.c.sku == sku, *mat_scope_preds))
            ).first()

            if is_batch_mode:
                # === 批次模式：每行 = 一个批次 ===
                if material:
                    material_id = material.id
                    if batch_no_val:
                        # 查找已有批次
                        batch = sa_conn.execute(
                            select(_t_batches.c.id, _t_batches.c.quantity)
                            .where(and_(_t_batches.c.batch_no == batch_no_val,
                                        _t_batches.c.material_id == material_id))
                        ).first()
                        if batch:
                            current_qty = batch.quantity
                            difference = import_qty - current_qty
                            if difference > 0:
                                operation = RecordType.IN.value
                                total_in += difference
                            elif difference < 0:
                                operation = RecordType.OUT.value
                                total_out += abs(difference)
                            else:
                                operation = 'none'
                            preview_items.append(ImportPreviewItem(
                                sku=sku, name=material.name, category=category, unit=unit,
                                safe_stock=safe_stock, location=location,
                                current_quantity=current_qty, import_quantity=import_qty,
                                difference=difference, operation=operation,
                                batch_no=batch_no_val, contact_name=contact_name, contact_id=contact_id,
                                variant=variant_val or None,
                            ))
                        else:
                            # 批次号不存在，视为新批次导入
                            total_in += import_qty
                            preview_items.append(ImportPreviewItem(
                                sku=sku, name=material.name, category=category, unit=unit,
                                safe_stock=safe_stock, location=location,
                                current_quantity=0, import_quantity=import_qty,
                                difference=import_qty, operation=RecordType.IN.value,
                                batch_no=batch_no_val, is_batch_new=True,
                                contact_name=contact_name, contact_id=contact_id,
                                variant=variant_val or None,
                            ))
                    else:
                        # 新批次（无批次号）
                        total_in += import_qty
                        preview_items.append(ImportPreviewItem(
                            sku=sku, name=material.name, category=category, unit=unit,
                            safe_stock=safe_stock, location=location,
                            current_quantity=0, import_quantity=import_qty,
                            difference=import_qty, operation=RecordType.IN.value,
                            is_batch_new=True, contact_name=contact_name, contact_id=contact_id,
                            variant=variant_val or None,
                        ))
                else:
                    # 新SKU + 新批次
                    total_new += 1
                    total_in += import_qty
                    new_item = ImportPreviewItem(
                        sku=sku, name=name, category=category, unit=unit,
                        safe_stock=safe_stock, location=location,
                        current_quantity=None, import_quantity=import_qty,
                        difference=import_qty, operation='new', is_new=True,
                        is_batch_new=True, contact_name=contact_name, contact_id=contact_id,
                        variant=variant_val or None,
                    )
                    preview_items.append(new_item)
                    new_skus.append(new_item)
            else:
                # === 简化模式 ===
                # 同一 SKU 出现多次 → 每行作为新批次（不同位置/联系方）
                sku_is_duplicate = sku in seen_skus_simple
                seen_skus_simple.add(sku)

                # 显式 variant 列优先
                explicit_variant = variant_val if has_variant_col and variant_val else None

                if sku_is_duplicate:
                    # 重复的 SKU 行：作为新批次入库
                    mat_name = material.name if material else name
                    total_in += import_qty
                    preview_items.append(ImportPreviewItem(
                        sku=sku, name=mat_name, category=category, unit=unit,
                        safe_stock=safe_stock, location=location,
                        current_quantity=0, import_quantity=import_qty,
                        difference=import_qty, operation=RecordType.IN.value,
                        is_batch_new=True,
                        contact_name=contact_name, contact_id=contact_id,
                        variant=explicit_variant,
                    ))
                elif material:
                    current_qty = material.quantity
                    difference = import_qty - current_qty
                    if difference > 0:
                        operation = RecordType.IN.value
                        total_in += difference
                    elif difference < 0:
                        operation = RecordType.OUT.value
                        total_out += abs(difference)
                    else:
                        operation = 'none'
                    preview_items.append(ImportPreviewItem(
                        sku=sku, name=material.name, category=category, unit=unit,
                        safe_stock=safe_stock, location=location,
                        current_quantity=current_qty, import_quantity=import_qty,
                        difference=difference, operation=operation,
                        contact_name=contact_name, contact_id=contact_id,
                        variant=explicit_variant,
                    ))
                else:
                    total_new += 1
                    new_item = ImportPreviewItem(
                        sku=sku, name=name, category=category, unit=unit,
                        safe_stock=safe_stock, location=location,
                        current_quantity=None, import_quantity=import_qty,
                        difference=import_qty, operation='new', is_new=True,
                        contact_name=contact_name, contact_id=contact_id,
                        variant=explicit_variant,
                    )
                    preview_items.append(new_item)
                    new_skus.append(new_item)

                # 记录 Excel 原始名称，用于后处理自动提取 variant
                if not has_variant_col:
                    sku_excel_names.setdefault(sku, []).append((len(preview_items) - 1, name))

        # 后处理：自动从同SKU不同名称中提取 variant
        if not has_variant_col:
            for sku, entries in sku_excel_names.items():
                names = [e[1] for e in entries]
                if len(set(names)) <= 1:
                    continue  # 所有名称相同，无需提取
                common_name, variants = extract_variants(names)
                for (item_idx, _), variant in zip(entries, variants):
                    preview_items[item_idx].variant = variant
                    preview_items[item_idx].name = common_name

        # 查找缺失的SKU（系统中有但导入文件中没有的，且未被禁用的）
        # 单一真相源：current_quantity 取 active batches sum
        import_skus = {item.sku for item in preview_items}
        _batch_sum_sub = (
            select(
                _t_batches.c.material_id.label('material_id'),
                _sa_func.coalesce(_sa_func.sum(_t_batches.c.quantity), 0).label('qty'),
            )
            .where(_t_batches.c.is_exhausted == 0)
            .group_by(_t_batches.c.material_id)
            .subquery()
        )
        all_sys_stmt = (
            select(
                _t_materials.c.sku, _t_materials.c.name,
                _t_materials.c.category,
                _sa_func.coalesce(_batch_sum_sub.c.qty, 0).label('quantity'),
            )
            .select_from(
                _t_materials.outerjoin(_batch_sum_sub, _batch_sum_sub.c.material_id == _t_materials.c.id)
            )
            .where(and_(_t_materials.c.is_disabled == 0, *mat_scope_preds))
        )
        all_system_skus = sa_conn.execute(all_sys_stmt).fetchall()

        missing_skus = []
        for row in all_system_skus:
            if row.sku not in import_skus:
                missing_skus.append(MissingSkuItem(
                    sku=row.sku, name=row.name,
                    category=row.category or '未分类', current_quantity=row.quantity
                ))

        total_missing = len(missing_skus)

    mode_label = "批次模式" if is_batch_mode else "简化模式"
    duplicate_msg = f"，已跳过 {duplicate_rows} 条重复行" if duplicate_rows else ""
    return ExcelImportPreviewResponse(
        success=True,
        preview=preview_items,
        new_skus=new_skus,
        missing_skus=missing_skus,
        total_in=total_in,
        total_out=total_out,
        total_new=total_new,
        total_missing=total_missing,
        is_batch_mode=is_batch_mode,
        new_contacts=sorted(new_contacts_set),
        message=f'[{mode_label}] 共解析 {len(preview_items)} 条记录，其中新增 {total_new} 条'
                + duplicate_msg
                + (f'，有 {total_missing} 个SKU不在导入文件中' if total_missing > 0 else '')
                + (f'，将创建 {len(new_contacts_set)} 个新联系方' if new_contacts_set else '')
    )


@app.post("/api/materials/import-excel/confirm", response_model=ExcelImportResponse)
async def confirm_import_excel(
    request: ExcelImportConfirm,
    current_user: CurrentUser = Depends(require_permission(Resource.MATERIALS, Action.WRITE))
):
    """确认导入，执行变更单（需要operate权限）— 统一创建批次"""
    wh_id = require_warehouse_id(current_user, request.warehouse_id)
    in_count = 0
    out_count = 0
    new_count = 0
    records_created = 0
    warnings = []
    operator_user_id = current_user.id
    operator = request.operator if request.operator else current_user.get_operator_name()
    now_dt = datetime.now()

    # check_warehouse_access ignores its conn parameter (Phase 2b SA Core).
    check_warehouse_access(None, current_user, wh_id)

    # 校验所有显式传入的 contact_id 都属于本租户（防止跨租户写入引用）— uses its own SA read.
    target_tenant_for_contacts = resolve_tenant_id_for_write(current_user, wh_id)
    seen_contact_ids = {item.contact_id for item in request.changes if item.contact_id}
    for cid in seen_contact_ids:
        ensure_contact_tenant(None, current_user, cid, target_tenant_for_contacts)

    with get_engine().begin() as sa_conn:
        # 收集导入文件中的所有SKU
        import_skus = set(item.sku for item in request.changes)

        # 将不在导入文件中的SKU标记为禁用（需显式确认，仅限当前仓库）
        mat_scope_preds = list(build_authorized_scope_predicates(_t_materials, current_user, wh_id))
        if import_skus:
            if request.confirm_disable_missing_skus:
                sa_conn.execute(
                    update(_t_materials)
                    .where(and_(_t_materials.c.sku.notin_(list(import_skus)), *mat_scope_preds))
                    .values(is_disabled=1)
                )
            else:
                warnings.append("已跳过禁用导入文件之外的SKU，如需禁用请勾选确认选项后重试。")
            sa_conn.execute(
                update(_t_materials)
                .where(and_(_t_materials.c.sku.in_(list(import_skus)), *mat_scope_preds))
                .values(is_disabled=0)
            )

        # 前置：创建新联系方（联系方为租户级，不绑定仓库）
        contact_tenant_id = resolve_tenant_id_for_write(current_user, wh_id)
        contact_scope_preds = list(build_scope_predicates(_t_contacts, contact_tenant_id, None))
        contact_name_to_id = {}
        for item in request.changes:
            if item.contact_name and not item.contact_id:
                if item.contact_name not in contact_name_to_id:
                    existing = sa_conn.execute(
                        select(_t_contacts.c.id).where(and_(
                            _t_contacts.c.name == item.contact_name,
                            _t_contacts.c.is_disabled == 0,
                            *contact_scope_preds,
                        ))
                    ).first()
                    if existing:
                        contact_name_to_id[item.contact_name] = existing.id
                    else:
                        ins_res = sa_conn.execute(
                            insert(_t_contacts).values(
                                name=item.contact_name, is_supplier=1, warehouse_id=None,
                                tenant_id=contact_tenant_id, created_at=now_dt,
                            )
                        )
                        contact_name_to_id[item.contact_name] = ins_res.inserted_primary_key[0]

        def _get_contact_id(item):
            if item.contact_id:
                return item.contact_id
            if item.contact_name and item.contact_name in contact_name_to_id:
                return contact_name_to_id[item.contact_name]
            return None

        wh_tenant_id = resolve_tenant_id_for_write(current_user, wh_id)
        batch_scope_preds = list(build_scope_predicates(_t_batches, wh_tenant_id, wh_id))

        # In-session batch_no allocator. generate_batch_no(material_id) without a
        # cursor only sees committed rows, so consecutive calls inside one txn
        # would collide. Track allocated numbers in-memory and bump the suffix
        # until unique.
        today_prefix = datetime.now().strftime('%Y%m%d')
        allocated_batch_nos = set()

        def _alloc_batch_no(material_id):
            candidate = generate_batch_no(material_id, warehouse_id=wh_id)
            if candidate not in allocated_batch_nos:
                allocated_batch_nos.add(candidate)
                return candidate
            # Bump suffix until unique within this session.
            try:
                last_seq = int(candidate.split('-')[-1])
            except (ValueError, IndexError):
                last_seq = 0
            seq = last_seq + 1
            while True:
                candidate = f'{today_prefix}-{seq:03d}'
                if candidate not in allocated_batch_nos:
                    allocated_batch_nos.add(candidate)
                    return candidate
                seq += 1

        def _create_batch(material_id, quantity, location, contact_id, variant=None, batch_no=None):
            """创建新批次并返回 batch_id"""
            bn = batch_no or _alloc_batch_no(material_id)
            ins_res = sa_conn.execute(
                insert(_t_batches).values(
                    batch_no=bn, material_id=material_id, quantity=quantity,
                    initial_quantity=quantity, contact_id=contact_id,
                    location=location, variant=variant,
                    warehouse_id=wh_id, tenant_id=wh_tenant_id, created_at=now_dt,
                )
            )
            return ins_res.inserted_primary_key[0]

        def _create_record(material_id, rec_type, quantity, item_reason_category, reason_suffix, batch_id=None, contact_id=None):
            """创建出入库记录"""
            # reason_category 从每行 item 读取，reason_note 从全局备注 + 后缀拼接
            category = item_reason_category or ('purchase' if rec_type == RecordType.IN.value else 'sell')
            note_parts = []
            if request.reason_note:
                note_parts.append(request.reason_note)
            if reason_suffix:
                note_parts.append(reason_suffix.strip(' ()（）'))
            note = '; '.join(note_parts) if note_parts else None
            ins_res = sa_conn.execute(
                insert(_t_inventory_records).values(
                    material_id=material_id, type=rec_type, quantity=quantity,
                    operator=operator, operator_user_id=operator_user_id,
                    reason_category=category, reason_note=note,
                    contact_id=contact_id, batch_id=batch_id,
                    warehouse_id=wh_id, tenant_id=wh_tenant_id, created_at=now_dt,
                )
            )
            return ins_res.inserted_primary_key[0]

        def _lookup_material_id(sku):
            row = sa_conn.execute(
                select(_t_materials.c.id).where(and_(_t_materials.c.sku == sku, *mat_scope_preds))
            ).first()
            return row.id if row else None

        def _fifo_consume(material_id, quantity, record_id):
            """Consume `quantity` from the oldest batches of `material_id` and
            write batch_consumptions rows linked to `record_id`. Returns the
            unfilled remainder (0 on success). Caller decides how to react to
            a non-zero remainder (rollback / error response / raise).
            """
            remaining = quantity
            avail = sa_conn.execute(
                select(_t_batches.c.id, _t_batches.c.quantity)
                .where(and_(
                    _t_batches.c.material_id == material_id,
                    _t_batches.c.is_exhausted == 0,
                    _t_batches.c.quantity > 0,
                    *batch_scope_preds,
                ))
                .order_by(_t_batches.c.created_at.asc())
                .with_for_update()
            ).fetchall()
            for batch in avail:
                if remaining <= 0:
                    break
                consume = min(batch.quantity, remaining)
                new_batch_qty = batch.quantity - consume
                remaining -= consume
                # 并发门控：rowcount=0 表示批次已被其它事务改动，回滚整个事务
                upd_res = sa_conn.execute(
                    update(_t_batches)
                    .where(and_(
                        _t_batches.c.id == batch.id,
                        _t_batches.c.quantity >= consume,
                        _t_batches.c.is_exhausted == 0,
                    ))
                    .values(quantity=new_batch_qty,
                            is_exhausted=1 if new_batch_qty == 0 else 0)
                )
                if upd_res.rowcount != 1:
                    raise HTTPException(status_code=409, detail="批次并发冲突，请重试")
                sa_conn.execute(
                    insert(_t_batch_consumptions).values(
                        record_id=record_id, batch_id=batch.id,
                        quantity=consume, created_at=now_dt,
                    )
                )
            return remaining

        if request.is_batch_mode:
            # === 批次模式 ===
            for item in request.changes:
                if item.operation == 'none':
                    # 无变动，仅更新 batch location
                    if item.batch_no and item.location:
                        mid = _lookup_material_id(item.sku)
                        if mid is not None:
                            sa_conn.execute(
                                update(_t_batches)
                                .where(and_(_t_batches.c.batch_no == item.batch_no,
                                            _t_batches.c.material_id == mid))
                                .values(location=item.location)
                            )
                    continue

                contact_id = _get_contact_id(item)

                if item.is_new:
                    # 新SKU + 新批次
                    if not request.confirm_new_skus:
                        continue
                    existing_id = _lookup_material_id(item.sku)
                    if existing_id is None:
                        ins_res = sa_conn.execute(
                            insert(_t_materials).values(
                                name=item.name, sku=item.sku,
                                category=item.category or '未分类', quantity=0,
                                unit=item.unit or '个', safe_stock=item.safe_stock,
                                location=item.location or '',
                                warehouse_id=wh_id, tenant_id=wh_tenant_id, created_at=now_dt,
                            )
                        )
                        material_id = ins_res.inserted_primary_key[0]
                        new_count += 1
                    else:
                        material_id = existing_id

                    if item.import_quantity > 0:
                        batch_id = _create_batch(material_id, item.import_quantity, item.location, contact_id, item.variant)
                        # 单一真相源：不再写 materials.quantity，库存由 batches 派生。
                        _create_record(material_id, RecordType.IN.value, item.import_quantity, item.reason_category, ' (新建物料)', batch_id, contact_id)
                        in_count += 1
                        records_created += 1
                elif item.is_batch_new:
                    # 已有SKU，新批次
                    material_id = _lookup_material_id(item.sku)
                    if material_id is None:
                        continue
                    batch_id = _create_batch(material_id, item.import_quantity, item.location, contact_id, item.variant, item.batch_no)
                    # 单一真相源：不再写 materials.quantity
                    _create_record(material_id, RecordType.IN.value, item.import_quantity, item.reason_category, ' (新批次)', batch_id, contact_id)
                    in_count += 1
                    records_created += 1
                else:
                    # 已有批次有变动
                    material_id = _lookup_material_id(item.sku)
                    if material_id is None:
                        continue
                    batch = sa_conn.execute(
                        select(_t_batches.c.id, _t_batches.c.quantity)
                        .where(and_(_t_batches.c.batch_no == item.batch_no,
                                    _t_batches.c.material_id == material_id,
                                    *batch_scope_preds))
                    ).first()
                    if not batch:
                        continue

                    diff = item.difference
                    # 复活已耗尽批次：用户用 Excel 给历史 is_exhausted=1 的批次写回正数时，
                    # 必须同步清掉 is_exhausted 标记，否则所有 WHERE is_exhausted=0 的读端
                    # 会无视这一行 → 库存静默丢失。反向：若新值 <= 0 则标记耗尽。
                    sa_conn.execute(
                        update(_t_batches).where(_t_batches.c.id == batch.id)
                        .values(quantity=item.import_quantity,
                                location=item.location or '', variant=item.variant,
                                is_exhausted=0 if item.import_quantity > 0 else 1)
                    )
                    # 单一真相源：不再写 materials.quantity（batch 写入已是真相）

                    rec_type = RecordType.IN.value if diff > 0 else RecordType.OUT.value
                    new_record_id = _create_record(
                        material_id, rec_type, abs(diff), item.reason_category,
                        '', batch.id, contact_id,
                    )
                    if diff < 0:
                        # Pair the OUT inventory_record with a batch_consumptions
                        # row so per-batch SUM matches materials.quantity.
                        sa_conn.execute(
                            insert(_t_batch_consumptions).values(
                                record_id=new_record_id, batch_id=batch.id,
                                quantity=abs(diff), created_at=now_dt,
                            )
                        )
                    if diff > 0:
                        in_count += 1
                    else:
                        out_count += 1
                    records_created += 1

                # 更新 materials.location 为最新
                if item.location:
                    sa_conn.execute(
                        update(_t_materials)
                        .where(and_(_t_materials.c.sku == item.sku, *mat_scope_preds))
                        .values(location=item.location)
                    )
        else:
            # === 简化模式（统一创建批次）===
            for item in request.changes:
                contact_id = _get_contact_id(item)

                if item.is_new:
                    if not request.confirm_new_skus:
                        continue

                    existing = sa_conn.execute(
                        select(_t_materials.c.id).where(and_(_t_materials.c.sku == item.sku, *mat_scope_preds))
                    ).first()

                    if existing is not None:
                        # SKU已存在，按已有物料处理
                        # 单一真相源：当前库存读自 active batches sum
                        material_id = existing.id
                        current_qty = int(sa_conn.execute(
                            select(_sa_func.coalesce(_sa_func.sum(_t_batches.c.quantity), 0))
                            .where(and_(
                                _t_batches.c.material_id == material_id,
                                _t_batches.c.is_exhausted == 0,
                            ))
                        ).scalar() or 0)
                        if item.import_quantity != current_qty:
                            diff = item.import_quantity - current_qty
                            # 单一真相源：不再写 materials.quantity
                            rec_type = RecordType.IN.value if diff > 0 else RecordType.OUT.value
                            if diff > 0:
                                batch_id = _create_batch(material_id, abs(diff), item.location, contact_id, item.variant)
                                _create_record(material_id, rec_type, abs(diff), item.reason_category, ' (SKU已存在，调整库存)', batch_id, contact_id)
                            else:
                                # Negative diff: route through FIFO so per-batch
                                # totals stay consistent with materials.quantity
                                # instead of writing a phantom OUT with no
                                # batch_consumptions linkage.
                                new_record_id = _create_record(
                                    material_id, rec_type, abs(diff),
                                    item.reason_category,
                                    ' (SKU已存在，调整库存)', None, contact_id,
                                )
                                remaining = _fifo_consume(material_id, abs(diff), new_record_id)
                                if remaining > 0:
                                    sa_conn.rollback()
                                    return ExcelImportResponse(
                                        success=False, in_count=in_count, out_count=out_count,
                                        new_count=new_count, records_created=records_created,
                                        message=(
                                            f"出库失败：SKU {item.sku} 在可用批次中仅消耗到 "
                                            f"{abs(diff) - remaining}，仍缺 {remaining}，已终止导入。"
                                        ),
                                    )
                            records_created += 1
                        new_count += 1
                        continue

                    ins_res = sa_conn.execute(
                        insert(_t_materials).values(
                            name=item.name, sku=item.sku,
                            category=item.category or '未分类',
                            quantity=0,
                            unit=item.unit or '个', safe_stock=item.safe_stock,
                            location=item.location or '',
                            warehouse_id=wh_id, tenant_id=wh_tenant_id, created_at=now_dt,
                        )
                    )
                    material_id = ins_res.inserted_primary_key[0]
                    if item.import_quantity > 0:
                        batch_id = _create_batch(material_id, item.import_quantity, item.location, contact_id)
                        _create_record(material_id, RecordType.IN.value, item.import_quantity, item.reason_category, ' (新建物料)', batch_id, contact_id)
                        records_created += 1
                    new_count += 1
                else:
                    material = sa_conn.execute(
                        select(_t_materials.c.id)
                        .where(and_(_t_materials.c.sku == item.sku, *mat_scope_preds))
                    ).first()
                    if not material:
                        continue

                    material_id = material.id
                    # 单一真相源：当前库存读自 active batches sum
                    current_qty = int(sa_conn.execute(
                        select(_sa_func.coalesce(_sa_func.sum(_t_batches.c.quantity), 0))
                        .where(and_(
                            _t_batches.c.material_id == material_id,
                            _t_batches.c.is_exhausted == 0,
                        ))
                    ).scalar() or 0)

                    # 更新基本信息（含 variant 提取后可能变更的物料名称）
                    sa_conn.execute(
                        update(_t_materials).where(_t_materials.c.id == material_id)
                        .values(
                            name=item.name, safe_stock=item.safe_stock,
                            category=item.category or '未分类',
                            unit=item.unit or '个', location=item.location or '',
                        )
                    )

                    if item.operation == 'none':
                        continue

                    abs_diff = abs(item.difference)

                    if item.operation == RecordType.IN.value:
                        batch_id = _create_batch(material_id, abs_diff, item.location, contact_id, item.variant)
                        # 单一真相源：不再写 materials.quantity
                        _create_record(material_id, RecordType.IN.value, abs_diff, item.reason_category, '', batch_id, contact_id)
                        in_count += 1
                        records_created += 1
                    elif item.operation == RecordType.OUT.value:
                        if current_qty - abs_diff < 0:
                            # Roll back partial mutations from this batch import
                            # (preserves original sqlite get_db() semantics where
                            # an early return without conn.commit() discarded
                            # all in-progress writes).
                            sa_conn.rollback()
                            return ExcelImportResponse(
                                success=False, in_count=in_count, out_count=out_count,
                                new_count=new_count, records_created=records_created,
                                message=f"出库失败：SKU {item.sku} 出库 {abs_diff} 超过当前库存 {current_qty}，已终止导入。"
                            )
                        # 单一真相源：不再写 materials.quantity；FIFO 在 _fifo_consume 中扣 batches
                        out_reason = item.reason_category or 'sell'
                        out_ins = sa_conn.execute(
                            insert(_t_inventory_records).values(
                                material_id=material_id, type=RecordType.OUT.value, quantity=abs_diff,
                                operator=operator, operator_user_id=operator_user_id,
                                reason_category=out_reason, reason_note=request.reason_note,
                                contact_id=contact_id,
                                warehouse_id=wh_id, tenant_id=wh_tenant_id, created_at=now_dt,
                            )
                        )
                        record_id = out_ins.inserted_primary_key[0]

                        remaining = _fifo_consume(material_id, abs_diff, record_id)
                        if remaining > 0:
                            # Materials.quantity says we have stock but no batch
                            # rows back it (e.g. orphan adjustments). Roll back
                            # the partial txn rather than commit a phantom OUT.
                            sa_conn.rollback()
                            return ExcelImportResponse(
                                success=False, in_count=in_count, out_count=out_count,
                                new_count=new_count, records_created=records_created,
                                message=(
                                    f"出库失败：SKU {item.sku} 在可用批次中仅消耗到 "
                                    f"{abs_diff - remaining}，仍缺 {remaining}，已终止导入。"
                                ),
                            )

                        out_count += 1
                        records_created += 1

    # R5: import only writes to materials/batches → material partition only
    get_fuzzy_matcher().invalidate_cache(entity_type="material")

    warning_text = f" {' '.join(warnings)}" if warnings else ""
    return ExcelImportResponse(
        success=True,
        in_count=in_count,
        out_count=out_count,
        new_count=new_count,
        records_created=records_created,
        message=f'导入完成：{in_count}条入库，{out_count}条出库，{new_count}条新增物料。{warning_text}'.strip()
    )


@app.get("/api/materials/import-excel/sample")
async def download_sample_excel(
    current_user: CurrentUser = Depends(require_permission(Resource.MATERIALS, Action.READ))
):
    """下载导入示例文件"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "库存数据"

    CANONICAL = [
        "物料名称", "规格", "物料编码(SKU)", "分类", "单位",
        "安全库存", "批次号", "库存", "存放位置", "联系方",
    ]
    SAMPLES = [
        ["M3 螺丝", "黑色 10mm", "SKU-0001", "紧固件", "个", 100, "B20260501-001", 500, "A-01-03", "深圳XX五金"],
        ["M3 螺丝", "银色 8mm", "SKU-0002", "紧固件", "个", 80, "B20260502-001", 300, "A-01-04", "深圳XX五金"],
        ["M6 螺母", "", "SKU-0003", "紧固件", "个", 200, "B20260503-001", 1000, "A-02-01", "东莞YY金属"],
        ["钢板 Q235", "200x200x5mm", "SKU-0004", "板材", "张", 50, "", 150, "B-03-01", "佛山ZZ钢材"],
    ]
    WIDTHS = [16, 14, 20, 12, 8, 12, 20, 10, 14, 18]

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for c, h in enumerate(CANONICAL, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for r, row in enumerate(SAMPLES, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border

    for i, w in enumerate(WIDTHS, 1):
        col_letter = chr(64 + i)
        ws.column_dimensions[col_letter].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=warehouse_import_sample.xlsx"},
    )


@app.get("/api/inventory/export-excel")
def export_inventory_records(
    start_date: Optional[str] = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
    product_name: Optional[str] = Query(None, description="产品名称"),
    record_type: Optional[str] = Query(None, description="记录类型(in/out)"),
    warehouse_id: Optional[int] = Query(None, description="仓库ID"),
    current_user: CurrentUser = Depends(require_permission(Resource.INVENTORY, Action.READ))
):
    """导出出入库记录为Excel（支持筛选，含批次信息）— Phase 3f: SA Core read."""
    wh_id = resolve_warehouse_id(current_user, warehouse_id)

    preds = list(build_authorized_scope_predicates(_t_inventory_records, current_user, wh_id))
    if start_date:
        preds.append(_sa_func.date(_t_inventory_records.c.created_at) >= start_date)
    if end_date:
        preds.append(_sa_func.date(_t_inventory_records.c.created_at) <= end_date)
    if product_name:
        preds.append(_t_materials.c.name.like(f'%{product_name}%'))
    if record_type and record_type != 'all':
        preds.append(_t_inventory_records.c.type == record_type)

    rec_stmt = (
        select(
            _t_inventory_records.c.id,
            _t_materials.c.name,
            _t_materials.c.sku,
            _t_materials.c.category,
            _t_inventory_records.c.type,
            _t_inventory_records.c.quantity,
            _t_inventory_records.c.operator,
            _t_inventory_records.c.operator_user_id,
            _t_inventory_records.c.reason_category,
            _t_inventory_records.c.reason_note,
            _t_inventory_records.c.created_at,
            _t_contacts.c.name.label('contact_name'),
            _t_inventory_records.c.batch_id,
            _t_batches.c.batch_no,
            _t_batches.c.variant,
            _t_users.c.display_name.label('operator_display_name'),
            _t_users.c.username.label('operator_username'),
        )
        .select_from(
            _t_inventory_records
                .join(_t_materials, _t_inventory_records.c.material_id == _t_materials.c.id)
                .outerjoin(_t_contacts, _t_inventory_records.c.contact_id == _t_contacts.c.id)
                .outerjoin(_t_batches, _t_inventory_records.c.batch_id == _t_batches.c.id)
                .outerjoin(_t_users, _t_inventory_records.c.operator_user_id == _t_users.c.id)
        )
        .order_by(_t_inventory_records.c.created_at.desc())
    )
    if preds:
        rec_stmt = rec_stmt.where(and_(*preds))

    with get_engine().connect() as sa_conn:
        records = sa_conn.execute(rec_stmt).fetchall()

        # 为出库记录获取批次消耗详情
        batch_details_map = {}
        out_record_ids = [r.id for r in records if r.type == RecordType.OUT.value]
        if out_record_ids:
            cons_stmt = (
                select(
                    _t_batch_consumptions.c.record_id,
                    _t_batches.c.batch_no,
                    _t_batch_consumptions.c.quantity,
                    _t_batches.c.created_at,
                )
                .select_from(
                    _t_batch_consumptions.join(_t_batches, _t_batch_consumptions.c.batch_id == _t_batches.c.id)
                )
                .where(_t_batch_consumptions.c.record_id.in_(out_record_ids))
                .order_by(_t_batch_consumptions.c.record_id, _t_batches.c.created_at.asc())
            )
            for cons in sa_conn.execute(cons_stmt).fetchall():
                batch_details_map.setdefault(cons.record_id, []).append(
                    f"{cons.batch_no}×{cons.quantity}"
                )
            batch_details_map = {k: ', '.join(v) for k, v in batch_details_map.items()}

    wb = Workbook()
    ws = wb.active
    ws.title = "出入库记录"

    headers = ['物料名称', '规格', '物料编码', '商品类型', '记录类型', '数量', '批次', '联系方', '操作人', '原因类别', '备注', '时间']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for row_idx, record in enumerate(records, 2):
        # 批次信息：入库显示批次号，出库显示消耗详情
        batch_info = ''
        if record.type == RecordType.IN.value and record.batch_no:
            batch_info = record.batch_no
        elif record.type == RecordType.OUT.value:
            batch_info = batch_details_map.get(record.id, '')

        created_at_str = (record.created_at.strftime('%Y-%m-%d %H:%M:%S')
                          if isinstance(record.created_at, datetime) else record.created_at)

        ws.cell(row=row_idx, column=1, value=record.name)
        ws.cell(row=row_idx, column=2, value=record.variant or '')
        ws.cell(row=row_idx, column=3, value=record.sku)
        ws.cell(row=row_idx, column=4, value=record.category)
        ws.cell(row=row_idx, column=5, value='入库' if record.type == RecordType.IN.value else '出库')
        ws.cell(row=row_idx, column=6, value=record.quantity)
        ws.cell(row=row_idx, column=7, value=batch_info)
        ws.cell(row=row_idx, column=8, value=record.contact_name or '')
        # 操作员：优先使用用户表中的显示名称，否则回退到旧的operator字段
        operator_name = record.operator_display_name or record.operator_username or record.operator
        ws.cell(row=row_idx, column=9, value=operator_name)
        ws.cell(row=row_idx, column=10, value=REASON_CATEGORY_LABELS.get(record.reason_category, record.reason_category or ''))
        ws.cell(row=row_idx, column=11, value=record.reason_note or '')
        ws.cell(row=row_idx, column=12, value=created_at_str)

    # 设置列宽
    column_widths = [22, 10, 18, 14, 12, 10, 28, 16, 14, 14, 24, 22]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"inventory_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.post("/api/inventory/add-record")
async def add_inventory_record(
    http_request: Request,
    request: ManualRecordRequest,
    current_user: CurrentUser = Depends(require_permission(Resource.INVENTORY, Action.WRITE))
):
    """手动新增出入库记录（需要operate权限）- 返回StockInResponse或StockOutResponse"""
    # 使用请求中的operator，如果为空则使用当前用户名
    operator = request.operator if request.operator else current_user.get_operator_name()

    if request.type == RecordType.IN.value:
        result = await stock_in(
            StockOperationRequest(
                product_name=request.product_name,
                quantity=request.quantity,
                reason_category=request.reason_category,
                reason_note=request.reason_note,
                operator=operator,
                contact_id=request.contact_id,
                location=request.location,
                batch_no=request.batch_no,
                variant=request.variant,
                warehouse_id=request.warehouse_id,
            ),
            current_user
        )
        # 入库成功且填写了库位时，更新产品汇总库位（必须限定到 stock_in 实际写入的仓库/租户，
        # 否则同名 SKU 在其他租户/仓库的 location 会被一起改掉）
        if result.success and request.location:
            wh_id = require_warehouse_id(current_user, request.warehouse_id)
            scope_preds = build_scope_predicates(
                _t_materials,
                resolve_tenant_id_for_write(current_user, wh_id),
                wh_id,
            )
            with get_engine().begin() as sa_conn:
                sa_conn.execute(
                    update(_t_materials)
                    .where(and_(_t_materials.c.name == request.product_name, *scope_preds))
                    .values(location=request.location)
                )
        return result
    elif request.type == RecordType.OUT.value:
        return await stock_out(
            http_request,
            StockOperationRequest(
                product_name=request.product_name,
                quantity=request.quantity,
                reason_category=request.reason_category,
                reason_note=request.reason_note,
                operator=operator,
                contact_id=request.contact_id,
                location=request.location,
                batch_no=request.batch_no,
                variant=request.variant,
                warehouse_id=request.warehouse_id,
            ),
            current_user
        )
    else:
        return StockOperationResponse(
            success=False,
            error="无效的操作类型",
            message="类型必须是 'in' 或 'out'"
        )


# ============ MCP 连接管理 ============

# MCP process manager singleton lives on ``app.state.mcp_manager`` (one per
# worker process). Routes resolve it through ``Depends(get_mcp_manager)`` in
# ``routers/mcp_admin.py``; ``deps.get_mcp_manager`` lazily falls back to a
# fresh instance under FastAPI ``TestClient`` (which skips lifespan events).


@app.on_event("startup")
async def _run_migrations():
    """Run Alembic migrations on startup. Idempotent.

    Replaces the legacy module-level ``init_database()`` call so that the
    deployment works on both SQLite (dev/local) and MySQL (production).
    Also seeds mock data after the schema is in place when INIT_MOCK_DATA is
    enabled.
    """
    from alembic.config import Config as AlembicConfig
    from alembic import command as alembic_command
    cfg = AlembicConfig(os.path.join(os.path.dirname(__file__), "alembic.ini"))
    cfg.set_main_option(
        "script_location", os.path.join(os.path.dirname(__file__), "alembic")
    )
    alembic_command.upgrade(cfg, "head")

    # 幂等补种：确保 tenant 1 和默认仓库存在，不依赖 init_database()。
    # Docker 部署走纯 Alembic 路径，Alembic 只建表不插数据，需在此补齐。
    _seed_base_data()

    # Schema 不变式校验 — 防止 metadata.py 加列忘写迁移（曾踩坑两次）。
    # 故意让异常向上抛：缺列时拒绝启动，不该被吞掉。
    _validate_schema_matches_metadata()

    # DEPLOY_MODE 不变式校验 — 启动阶段就把不一致状态拦掉，避免运行时 UI/业务半对半错。
    # 故意让异常向上抛：违反不变式应阻止启动，不该被吞掉。
    _validate_deploy_mode_invariants()

    # Boot-time route audit (warning-only). Bugs here must NEVER crash
    # startup, so we swallow any exception and log it.
    try:
        _audit_routes()
    except Exception as e:  # noqa: BLE001
        logger.warning(f"_audit_routes() skipped: {e}")

    if INIT_MOCK_DATA:
        try:
            generate_mock_data()
        except Exception as e:
            logger.warning(f"generate_mock_data() skipped: {e}")


@app.on_event("startup")
async def startup_mcp_manager():
    """启动时恢复 auto_start 的 MCP 连接"""
    # Create the per-worker singleton on app.state.
    mcp_manager = MCPProcessManager()
    app.state.mcp_manager = mcp_manager
    await mcp_manager.start_monitor()

    # 恢复 auto_start 的连接 — Phase 3e: subprocess/network 在 engine.begin() 之外
    try:
        with get_engine().connect() as sa_conn:
            rows = sa_conn.execute(
                select(
                    _t_mcp_connections.c.id,
                    _t_mcp_connections.c.name,
                    _t_mcp_connections.c.mcp_endpoint,
                    _t_mcp_connections.c.api_key,
                    _t_mcp_connections.c.debug_mode,
                ).where(_t_mcp_connections.c.auto_start == 1)
            ).all()

        for row in rows:
            logger.info(f"Auto-starting MCP connection: {row.name}")
            await mcp_manager.start_connection(
                row.id, row.mcp_endpoint, row.api_key,
                debug_mode=bool(row.debug_mode)
            )
            # 更新数据库状态（status 列为 String，updated_at 为 String(32) ISO 字符串）
            status_info = mcp_manager.get_connection_status(row.id)
            with get_engine().begin() as sa_conn:
                sa_conn.execute(
                    update(_t_mcp_connections)
                    .where(_t_mcp_connections.c.id == row.id)
                    .values(status=status_info['status'], updated_at=datetime.now().isoformat())
                )
    except Exception as e:
        logger.error(f"Failed to restore MCP connections: {e}")


@app.on_event("shutdown")
async def shutdown_mcp_manager():
    """关闭时停止所有 MCP 连接"""
    mcp_manager = getattr(app.state, "mcp_manager", None)
    if mcp_manager is not None:
        await mcp_manager.stop_all()


# MCP connection admin routes moved to backend/routers/mcp_admin.py
# (Phase 3 split, task #7). The router is included below; the
# MCPProcessManager singleton is now resolved per-request via
# Depends(get_mcp_manager) (see deps.py) and lives on app.state.
from routers.mcp_admin import router as mcp_admin_router
app.include_router(mcp_admin_router)


# ============ ERP Provider 管理 APIs ============
# ERP routes moved to backend/routers/erp.py (Phase 2 split, task #6).
# /api/system/mode below remains here — it is not part of the erp domain.
from routers.erp import router as erp_router
app.include_router(erp_router)


@app.get("/api/system/mode")
async def get_system_mode():
    """查询系统当前运行模式（self_owned / external_erp）及部署模式（single_tenant / multi_tenant）。

    无需登录：deploy_mode 是部署元信息，前端需要它在登录前就能决定 UI 形态（多租户/单租户），
    要求登录会让首屏一直走 single_tenant 默认值，触发模式判断分裂。system_mode 同理（self_owned
    vs external_erp 只决定 UI 走向，不暴露任何业务数据）。
    """
    with get_engine().connect() as sa_conn:
        row = sa_conn.execute(
            select(_t_system_settings.c.value).where(_t_system_settings.c.key == 'system_mode')
        ).first()
    mode = row.value if row else 'self_owned'
    return {"mode": mode, "deploy_mode": get_deploy_mode()}


@app.put("/api/system/mode")
async def set_system_mode(
    request: Request,
    current_user: CurrentUser = Depends(require_permission(Resource.SYSTEM, Action.ADMIN))
):
    """切换系统运行模式"""
    body = await request.json()
    mode = body.get('mode', '')
    if mode not in ('self_owned', 'external_erp'):
        raise HTTPException(status_code=400, detail="mode 必须是 self_owned 或 external_erp")

    # 切换到外部ERP模式时，必须先有激活的Provider（按当前租户范围）
    if mode == 'external_erp':
        preds = [_t_erp_providers.c.is_active == 1]
        preds.extend(build_scope_predicates(_t_erp_providers, current_user.tenant_id, None))
        with get_engine().connect() as sa_conn:
            row = sa_conn.execute(
                select(_t_erp_providers.c.id).where(and_(*preds))
            ).first()
        if not row:
            raise HTTPException(status_code=400, detail="切换到外部ERP模式前，请先激活一个 Provider")

    # Upsert system_mode — Phase 3e: SA Core
    with get_engine().begin() as sa_conn:
        upd = sa_conn.execute(
            update(_t_system_settings)
            .where(_t_system_settings.c.key == 'system_mode')
            .values(value=mode, updated_at=datetime.now())
        )
        if upd.rowcount == 0:
            sa_conn.execute(
                insert(_t_system_settings).values(key='system_mode', value=mode)
            )

    logger.info(f"系统模式切换为: {mode}，操作人: {current_user.display_name}")
    return {"mode": mode}


# ERP test/activate/deactivate/status moved to backend/routers/erp.py
# (Phase 2 split, task #6).


# ============ Face Recognition Management APIs ============
# Moved to backend/routers/face.py (Phase 1 split, task #5).
from routers.face import router as face_router
app.include_router(face_router)


# ============ Factory 设备代理 API ============

@app.get("/factory/devices")
async def factory_devices(
    request: Request,
    page: int = Query(1, ge=1, description="页码，最小值 1"),
    pageSize: int = Query(20, ge=1, le=100, description="每页数量，范围 1~100"),
    query: Optional[str] = Query(None, description="搜索关键词（如设备 ID）"),
):
    """代理小智开发者设备查询接口，透传原始 JSON 结构。"""
    # 验证 X-Factory-Key
    incoming_key = request.headers.get("X-Factory-Key", "")
    if not FACTORY_API_KEY or incoming_key != FACTORY_API_KEY:
        return JSONResponse(
            status_code=401,
            content={"success": False, "message": "Unauthorized"},
        )

    upstream_url = f"{FACTORY_API_BASE_URL}/factory/devices"
    params = {"page": page, "pageSize": pageSize}
    if query:
        params["query"] = query

    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.get(upstream_url, params=params, headers={"X-Factory-Key": FACTORY_API_KEY})
            return JSONResponse(content=resp.json(), status_code=resp.status_code)
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="Upstream API timeout")
    except httpx.RequestError:
        raise HTTPException(status_code=502, detail="Upstream API unavailable")


# ============ 前端静态文件（all-in-one 部署）============

STATIC_DIR = os.environ.get('STATIC_DIR', '')
if not STATIC_DIR:
    # 自动检测：Docker 环境 /app/static 或开发环境 ../frontend/dist
    for candidate in ['/app/static', os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'frontend', 'dist')]:
        if os.path.isdir(candidate) and os.path.exists(os.path.join(candidate, 'index.html')):
            STATIC_DIR = candidate
            break

if STATIC_DIR and os.path.isdir(STATIC_DIR):
    from fastapi.staticfiles import StaticFiles
    from fastapi.responses import FileResponse

    _index_html = os.path.join(STATIC_DIR, 'index.html')

    # /assets 静态资源（带缓存）
    _assets_dir = os.path.join(STATIC_DIR, 'assets')
    if os.path.isdir(_assets_dir):
        app.mount("/assets", StaticFiles(directory=_assets_dir), name="static-assets")

    # SPA catch-all: 非 /api 非 /assets 的请求都返回 index.html
    @app.get("/{path:path}")
    async def serve_spa(path: str):
        # 先尝试精确匹配静态文件（如 favicon.ico）
        file_path = os.path.join(STATIC_DIR, path)
        if path and os.path.isfile(file_path):
            return FileResponse(file_path)
        # index.html 禁止缓存，确保部署后用户立即获取新版本
        return FileResponse(_index_html, headers={"Cache-Control": "no-cache, no-store, must-revalidate"})

    logger.info(f"Serving frontend from {STATIC_DIR}")


# ============ 启动配置 ============

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=int(os.environ.get('PORT', 2124)))
